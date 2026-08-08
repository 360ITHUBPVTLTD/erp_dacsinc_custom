# import frappe
# from frappe.model.document import Document

# @frappe.whitelist()
# def create_quotation_from_lead(lead_name):
#     """
#     Create a Quotation from a Lead and return the Quotation name.
#     """
#     # Fetch the Lead document
#     lead = frappe.get_doc('Lead', lead_name)
    
#     if not lead:
#         frappe.throw(f"No Lead found with name {lead_name}")
    
#     # Create a new Quotation
#     quotation = frappe.get_doc({
#         'doctype': 'Quotation',
#         'customer': lead.company_name or '',  # Set customer or other relevant fields
#         'customer_name': lead.company_name or lead.first_name + ' ' + (lead.middle_name or '') + ' ' + (lead.last_name or ''),
#         'quotation_to': 'Customer',
#         'contact_person': lead.custom_contact_person_name,
#         'contact_email': lead.email_id,
#         'contact_mobile': lead.mobile_no,
#         'address': lead.custom_address,
#         'gstin': lead.custom_gstin,
#         # Add other fields as required
#     })

#     # Save the Quotation document
#     quotation.insert(ignore_permissions=True)

#     return quotation.name


import frappe
from frappe import _  # Import the _ function for translations

def get_mapped_doc(source_doctype, source_name, mapping, target_doc=None, set_missing_values=None, ignore_permissions=False):
    """
    Map fields from a source document to a target document.

    :param source_doctype: The doctype of the source document.
    :param source_name: The name of the source document.
    :param mapping: A dictionary defining field mappings.
    :param target_doc: An optional target document instance.
    :param set_missing_values: An optional function to set additional values.
    :param ignore_permissions: Whether to ignore permissions checks.
    :return: The mapped target document.
    """
    
    # Load the source document
    source_doc = frappe.get_doc(source_doctype, source_name)

    # If no target document is provided, create a new one
    if not target_doc:
        target_doc = frappe.new_doc(mapping[source_doctype]["doctype"])

    # Map fields from source to target
    field_map = mapping[source_doctype].get("field_map", {})
    field_no_map = mapping[source_doctype].get("field_no_map", [])
    
    for src_field, tgt_field in field_map.items():
        if hasattr(source_doc, src_field) and tgt_field not in field_no_map:
            setattr(target_doc, tgt_field, getattr(source_doc, src_field))

    # Set missing values if provided
    if set_missing_values:
        set_missing_values(source_doc, target_doc)

    # Save the document if ignore_permissions is not True
    if not ignore_permissions:
        target_doc.insert()

    return target_doc

# @frappe.whitelist()
# def make_customer(source_name, target_doc=None):
#     # Call _make_customer to get the mapped document
#     target_doc = _make_customer(source_name, target_doc)
    
#     # Return the target document for preview
#     return target_doc.as_dict()

# def _make_customer(source_name, target_doc=None, ignore_permissions=False):
#     def set_missing_values(source, target):
#         if source.company_name:
#             target.customer_type = "Company"
#             target.customer_name = source.company_name
#         else:
#             target.customer_type = "Individual"
#             target.customer_name = source.lead_name
        
#         target.customer_group = frappe.db.get_default("Customer Group")

#     # Load the source document
#     source_doc = frappe.get_doc("Lead", source_name)

#     # Create a new target document
#     target_doc = frappe.new_doc("Customer")

#     # Map fields from source to target
#     field_map = {
#         "name": "lead_name",
#         "company_name": "customer_name",
#         "contact_no": "phone_1",
#         "fax": "fax_1",
#         "mobile_no": "custom_customer_mobile_no",
#         "email_id": "custom_customer_mail_id",
#         "custom_gstin": "gstin",
#         "custom_address": "custom_address",
#     }
#     field_no_map = ["disabled"]
    
#     for src_field, tgt_field in field_map.items():
#         if hasattr(source_doc, src_field) and tgt_field not in field_no_map:
#             setattr(target_doc, tgt_field, getattr(source_doc, src_field))

#     # Set missing values if provided
#     if set_missing_values:
#         set_missing_values(source_doc, target_doc)

#     return target_doc

@frappe.whitelist()
def make_customer(source_name, target_doc=None):
    # Call _make_customer to get the mapped document
    target_doc = _make_customer(source_name, target_doc)
    
    # Return the target document for preview
    return target_doc.as_dict()

def _make_customer(source_name, target_doc=None, ignore_permissions=False):
    def set_missing_values(source, target):
        if source.company_name:
            target.customer_type = "Company"
            target.customer_name = source.company_name
        else:
            target.customer_type = "Individual"
            target.customer_name = source.lead_name
        
        target.customer_group = frappe.db.get_default("Customer Group")

    # Load the source document
    source_doc = frappe.get_doc("Lead", source_name)

    # Check if a Customer with the same email ID or mobile number already exists
    existing_customer = frappe.get_all("Customer", filters={
        "custom_customer_mail_id": source_doc.email_id,
        "custom_customer_mobile_no": source_doc.mobile_no,
        
    }, fields=["name"])

    if existing_customer:
        # Raise an error if a customer with the same email ID or mobile number already exists
        frappe.throw(_("A customer with this email ID or mobile number already exists."))

    # Create a new target document
    target_doc = frappe.new_doc("Customer")

    # Map fields from source to target
    field_map = {
        "name": "lead_name",
        "company_name": "customer_name",
        "contact_no": "phone_1",
        "fax": "fax_1",
        "mobile_no": "custom_customer_mobile_no",
        "email_id": "custom_customer_mail_id",
        "custom_gstin": "gstin",
        "custom_address": "custom_address",
        "gender":"gender",
        "custom_reference_by_cid":"custom_reference_by_cid",
        "custom_address":"custom_address",
        "country":"custom_country",
        "custom_state_1":"custom_state",
        "custom_city_2":"custom_city",
        "custom_pin_code":"custom_pincode"
    }
    field_no_map = ["disabled"]
    
    for src_field, tgt_field in field_map.items():
        if hasattr(source_doc, src_field) and tgt_field not in field_no_map:
            setattr(target_doc, tgt_field, getattr(source_doc, src_field))

    # Set missing values if provided
    if set_missing_values:
        set_missing_values(source_doc, target_doc)

    return target_doc



@frappe.whitelist()
def make_quotation(source_name, target_doc=None, ignore_permissions=False):
    return _make_quotation(source_name, target_doc)

def _make_quotation(source_name, target_doc=None, ignore_permissions=False):
    def set_missing_values(source, target):
        target.quotation_to = "Lead"
        target.custom_customer_gstin = source.custom_gstin
        target.custom_customer_email_id = source.email_id
        target.custom_customer_mobile_no = source.mobile_no
        target.custom_address = source.custom_address

    # Get the source document
    source_doc = frappe.get_doc("Lead", source_name)

    # Create a new Quotation document
    target_doc = frappe.new_doc("Quotation")

    # Map fields from source to target
    field_map = {
        "name": "party_name",
        "contact_no": "contact_mobile",
        "email_id": "contact_email",
        "mobile_no": "contact_mobile",
    }
    field_no_map = ["disabled"]

    for src_field, tgt_field in field_map.items():
        if hasattr(source_doc, src_field) and tgt_field not in field_no_map:
            setattr(target_doc, tgt_field, getattr(source_doc, src_field))

    # Set additional values
    set_missing_values(source_doc, target_doc)

    # Return the target document as a dictionary for the client side
    return target_doc.as_dict()



import frappe
@frappe.whitelist()
def get_activities_for_lead(lead_name):
    """
    Fetches all To-dos and Events linked to a specific lead.
    Returns a dictionary with two keys: 'todos' and 'events'.
    Filters records based on the current user's permissions.
    """
    if not lead_name:
        return None

    user = frappe.session.user  # Get the current session user

    # Fetch all To-dos where the lead is the reference, considering user permissions
    todos = frappe.get_all(
        "ToDo",
        filters={
            "reference_type": "Lead",
            "reference_name": lead_name,
        },
        fields=["name", "description", "status", "date", "allocated_to"],
    )
    
    # Filter ToDos based on access permissions
    todos = [
        todo for todo in todos
        if frappe.has_permission("ToDo", "read", todo["name"])  # Check if user has 'read' permission on the ToDo
    ]

    # Fetch all Events where the lead is referenced, considering user permissions
    events = frappe.get_all(
        "Event",
        filters=[
            ["Event Participants", "reference_doctype", "=", "Lead"],
            ["Event Participants", "reference_docname", "=", lead_name]
        ],
        fields=["name", "subject", "starts_on", "status"]
    )
    
    # Filter Events based on access permissions
    events = [
        event for event in events
        if frappe.has_permission("Event", "read", event["name"])  # Check if user has 'read' permission on the Event
    ]

    return {
        "todos": todos,
        "events": events
    }
   

import frappe

@frappe.whitelist()
def get_tasks_for_lead(lead_name):
    return frappe.get_all(
        "Task",
        filters={
            "custom_lead_id": lead_name,
        },
        fields=["name", "subject", "status", "exp_end_date"],
        order_by="modified desc"
    )


@frappe.whitelist()
def get_lead_activity_status(lead_id):
    open_count = frappe.db.count("Event Activity", {"reference_type":"Lead","reference_name": lead_id, "status": "Open"})
    closed_count = frappe.db.count("Event Activity", {"reference_type":"Lead","reference_name": lead_id, "status": "Completed"})
    total_count = frappe.db.count("Event Activity", {"reference_type":"Lead","reference_name": lead_id})  # Total activities regardless of status

    return {"open": open_count, "closed": closed_count, "total": total_count}

from frappe.share import add, get_users, remove

# Event after insertion for Event
def after_insert_event(doc, method):
    if doc.custom_allocated_to:
        # remove_existing_shares(doc)
        share_event_with_user(doc)

# Event before saving for Event
def before_save_event(doc, method):
    if not doc.is_new():
        if doc.custom_allocated_to:
            if doc.get("name"):
                # remove_existing_shares(doc)
                share_event_with_user(doc)

# Lead after insertion
def after_insert_lead(doc, method):
    doc.db_set("custom_lead_id", doc.name)
    # print("After Insert Lead Triggered")  # Debugging line to check if the function is called
    if doc.lead_owner:
        # print(f"Lead Owner: {doc.lead_owner}")  # Debugging line to check if lead_owner is set
        # remove_existing_shares(doc)
        share_event_with_user(doc)
    else:
        print("No lead_owner set.")  # Debugging line if lead_owner is not set

import frappe
def on_update_lead(doc, method=None):
    """
    Handle sharing in on_update to avoid 'Document Modified' errors.
    """
    # Use flags to prevent infinite loops if your sharing logic triggers more saves
    if doc.flags.in_sharing_logic:
        return

    doc.flags.in_sharing_logic = True

    # Get the previous owner from the last history row we just added in validate
    if doc.get('custom_lead_owner_change_history'):
        last_change = doc.custom_lead_owner_change_history[-1]
        
        # If the latest history entry was created just now (within the last few seconds)
        # This confirms an owner change actually happened
        old_owner = last_change.from_lead_owner
        new_owner = last_change.to_lead_owner

        if old_owner != new_owner:
            # Update Share Permissions for Lead
            from erp_dacsinc_custom.order_flow_api import add_docshare
            if old_owner:
                add_docshare(doc.doctype, doc.name, old_owner, read=1, write=0, share=1, notify=0)

            if new_owner:
                add_docshare(doc.doctype, doc.name, new_owner, read=1, write=1, share=1, notify=0)

            # Update sharing for related Event Activities
            events = frappe.get_all("Event Activity", 
                filters={"reference_type": "Lead", "reference_name": doc.name}, 
                pluck="name"
            )
            for event in events:
                if new_owner:
                    add_docshare("Event Activity", event, new_owner, read=1, write=0, share=1, notify=0)

            # Update sharing for related Quotations
            quotations_by_lead_id = frappe.get_all("Quotation", filters={"custom_lead_id": doc.name}, pluck="name")
            quotations_by_party = frappe.get_all("Quotation", filters={"quotation_to": "Lead", "party_name": doc.name}, pluck="name")
            quotation_names = list(set(quotations_by_lead_id + quotations_by_party))

            for q_name in quotation_names:
                if new_owner:
                    add_docshare("Quotation", q_name, new_owner, read=1, write=1, share=1, notify=0)
                
                if old_owner:
                    add_docshare("Quotation", q_name, old_owner, read=1, write=0, share=1, notify=0)

    # Handle sharing for new owner if it's a new lead (Logic from your share_event_with_user)
    # Note: Only run if it hasn't been shared yet
    if doc.lead_owner:
        # Check if already shared to avoid redundant DB hits
        if not frappe.db.exists("DocShare", {"share_doctype": doc.doctype, "share_name": doc.name, "user": doc.lead_owner}):
            from erp_dacsinc_custom.order_flow_api import add_docshare
            add_docshare(doc.doctype, doc.name, doc.lead_owner, read=1, write=1, share=1, notify=0)



# Function to remove existing shares
def remove_existing_shares(doc):
    try:
        # Get all users who have access to this document
        shares = get_users(doc.doctype, doc.name)
        
        # Remove share for each user except the new one
        for share in shares:
            # Check to ensure that the current user is not the one we want to share with
            if (doc.doctype == "Event" and share.user != doc.custom_allocated_to) or \
               (doc.doctype == "Lead" and share.user != doc.lead_owner):
                remove(doc.doctype, doc.name, share.user)
        
        # frappe.msgprint(f"Removed all shares except the new user for {doc.doctype} {doc.name}")
    except Exception as e:
        frappe.log_error(f"Failed to remove shares for {doc.doctype} {doc.name}: {str(e)}")

# Function to share event or lead with the appropriate user
def share_event_with_user(doc):
    # print('sssssssssssssss',doc)
    try:
        # Check for Event or Lead and share accordingly
        if doc.doctype == "Event" and doc.custom_allocated_to:
            add(doc.doctype, doc.name, doc.custom_allocated_to, write=1, share=1, everyone=0)
            # frappe.msgprint(f"Event shared with {doc.custom_allocated_to}")
        elif doc.doctype == "Lead" and doc.lead_owner:
            add(doc.doctype, doc.name, doc.lead_owner, write=1, share=1, notify=0,everyone=0)
        elif doc.doctype == "Event Activity" and not doc.assigned_to:
            add(doc.doctype, doc.name, doc.assigned_to, write=1, share=1, notify=0,everyone=0)
            # frappe.msgprint(f"Lead shared with {doc.lead_owner}")
    except Exception as e:
        frappe.log_error(f"Failed to share {doc.doctype} {doc.name} with {doc.custom_allocated_to if doc.doctype == 'Event' else doc.lead_owner}: {str(e)}")





@frappe.whitelist()
def get_lead_details(id):
    lead = frappe.get_doc("Lead", id)
    return {
        "lead_name": lead.lead_name,
        "email_id": lead.email_id,
        "mobile_no": lead.mobile_no,
        "company_name": lead.company_name,
        "custom_lead_category": getattr(lead, "custom_lead_category", None),
        "link": f"/app/lead/{lead.name}"
    }

@frappe.whitelist()
def get_customer_details(id):
    customer = frappe.get_doc("Customer", id)
    return {
        "name": customer.customer_name,
        "email_id": customer.email_id,
        "mobile_no": customer.mobile_no,
        "company_name": customer.customer_group,  # or company_name if you have it
        "link": f"/app/customer/{customer.name}"
    }


@frappe.whitelist()
def get_supplier_details(id):
    supplier = frappe.get_doc("Supplier", id)
    return {
        "name": supplier.supplier_name,
        "email_id": supplier.email_id,
        "mobile_no": supplier.mobile_no,
        "company_name": supplier.supplier_type,  # or company_name if you have it
        "link": f"/app/supplier/{supplier.name}"
    }

@frappe.whitelist()
def get_business_contacts_details(id):
    business_contacts = frappe.get_doc("Business Contacts", id)
    return {
        "name": business_contacts.contact_name,
        # "email_id": business_contacts.email_id,
        "contact_type": business_contacts.contact_type,
        "mobile_no": business_contacts.mobile_number,
        "company_name": business_contacts.organization_name,  # or company_name if you have it
        "link": f"/app/business-contacts/{business_contacts.name}"
    }

############################################ lead Dashboard ###########################################################################


# import frappe
# from frappe.utils import getdate, today, add_days

# @frappe.whitelist()
# def get_user_roles():
#     return {"roles": frappe.get_roles(frappe.session.user)}

# @frappe.whitelist()
# def get_lead_dashboard_data(user=None):
#     filters = {}
#     if user:
#         filters["assign_to"] = user

#     leads = frappe.get_all(
#         "Lead",
#         filters=filters,
#         fields=["name", "lead_name", "creation", "custom_lead_category", "assign_to"]
#     )

#     return {"leads": leads, "is_admin": "DAC CRM Head" in frappe.get_roles()}

# @frappe.whitelist()
# def get_lead_activities(user=None):
#     filters = {}
#     if user:
#         filters["assigned_to"] = user

#     activities = frappe.get_all(
#         "Event Activity",
#         filters=filters,
#         fields=["name", "subject", "status", "starts_on", "due_date", "assigned_to", "reference_name"]
#     )

#     return {"activities": activities}



# file: your_app/your_app/doctype/lead_dashboard/lead_dashboard.py

# import frappe

# @frappe.whitelist()
# def get_lead_dashboard_data():
#     """
#     Returns:
#     {
#         "lead_counts": {
#             "Category A": 10,
#             "Category B": 5
#         },
#         "activity_counts": {
#             "Lead": {
#                 "Category A": 15,
#                 "Category B": 8
#             },
#             "Customer": {
#                 "Category A": 3,
#                 "Category B": 5
#             },
#             "Supplier": {
#                 "Category A": 2,
#                 "Category B": 1
#             }
#         }
#     }
#     """

#     # 1. Lead counts by custom_lead_category
#     lead_counts = frappe.db.sql("""
#         SELECT custom_lead_category, COUNT(name) as cnt
#         FROM `tabLead`
#         GROUP BY custom_lead_category
#     """, as_dict=True)

#     lead_counts_dict = {row.custom_lead_category or "Uncategorized": row.cnt for row in lead_counts}

#     # 2. Event Activity counts by reference_type and category
#     activity_counts = frappe.db.sql("""
#         SELECT la.reference_type, l.custom_lead_category, COUNT(la.name) as cnt
#         FROM `tabEvent Activity` la
#         LEFT JOIN `tabLead` l ON la.reference_name = l.name
#         WHERE la.reference_type IN ('Lead', 'Customer', 'Supplier')
#         GROUP BY la.reference_type, l.custom_lead_category
#     """, as_dict=True)

#     # Convert to nested dict: {reference_type: {category: count}}
#     activity_counts_dict = {}
#     for row in activity_counts:
#         reference_type = row.reference_type or "Unknown"
#         category = row.custom_lead_category or "Uncategorized"
#         if reference_type not in activity_counts_dict:
#             activity_counts_dict[reference_type] = {}
#         activity_counts_dict[reference_type][category] = row.cnt

#     return {
#         "lead_counts": lead_counts_dict,
#         "activity_counts": activity_counts_dict
#     }


import frappe
from frappe.utils import getdate, nowdate, add_days

# @frappe.whitelist()
# def get_lead_dashboard_data():
#     data = frappe._dict()

#     # Top summary cards
#     data.leads_without_activity = get_leads_without_activity_count()
#     data.total_follow_ups_today_upcoming = get_today_upcoming_follow_ups_count()
#     data.total_overdue_activities = get_overdue_activities_count()
#     data.total_open_activities = get_open_activities_count()
#     data.bc_without_activity = get_bc_without_activity_data()
#     # Lead counts by category
#     data.lead_counts = get_lead_category_counts()

#     return data


import frappe
from frappe.utils import getdate, nowdate, add_days, flt

@frappe.whitelist()
def get_lead_dashboard_data(from_date=None, to_date=None, selected_user=None):
    # Determine Role & User
    user_roles = frappe.get_roles(frappe.session.user)
    is_crm_head = "DAC CRM Head" in user_roles
    
    target_user = None
    if is_crm_head:
        target_user = selected_user if selected_user else None 
    else:
        target_user = frappe.session.user

    filters = frappe._dict({"from_date": from_date, "to_date": to_date, "user": target_user})
    data = frappe._dict()

    # Metrics
    data.leads_without_activity = get_filtered_leads_without_activity(filters)
    data.total_follow_ups_today_upcoming = get_filtered_activities(filters, timeframe="upcoming")
    data.total_overdue_activities = get_filtered_activities(filters, timeframe="overdue")
    data.total_open_activities = get_filtered_activities(filters)
    data.bc_status_counts = get_bc_status_counts(filters)
    data.bc_without_activity = get_bc_without_activity_data(filters)
    data.lead_counts = get_filtered_lead_category_counts(filters)

    return data

def get_filtered_leads_without_activity(filters):
    db_f = {"custom_is_activity_created": 0, "custom_lead_category": ["in", ["Enquiry", "Pipeline"]]}
    if filters.user: db_f["lead_owner"] = filters.user
    if filters.from_date: db_f["creation"] = ["between", [filters.from_date, filters.to_date]]
    return frappe.db.count("Lead", db_f)

def get_filtered_activities(filters, timeframe=None):
    db_f = {"status": "Open"}
    if filters.user: db_f["assigned_to"] = filters.user
    today = nowdate()
    if timeframe == "upcoming":
        db_f["starts_on"] = ["between", [f"{today} 00:00:00", f"{add_days(today, 3)} 23:59:59"]]
    elif timeframe == "overdue":
        db_f["starts_on"] = ["<", today]
    return frappe.db.count("Event Activity", db_f)

def get_bc_status_counts(filters):
    counts = {}
    for s in ["Open", "Converted to Lead", "Existing Customer"]:
        db_f = {"status": s}
        if filters.user: db_f["assign_to"] = filters.user
        if filters.from_date: db_f["creation"] = ["between", [filters.from_date, filters.to_date]]
        counts[s] = frappe.db.count("Business Contacts", db_f)
    return counts

def get_filtered_lead_category_counts(filters):
    res = {}
    for cat in ['Enquiry', 'Pipeline', 'Order', 'Lost Enquiry', 'Lost Pipeline']:
        db_f = {"custom_lead_category": cat}
        if filters.user: db_f["lead_owner"] = filters.user
        if filters.from_date: db_f["creation"] = ["between", [filters.from_date, filters.to_date]]
        
        recs = frappe.db.get_all("Lead", filters=db_f, fields=["custom_expected_revenue", "custom_po_value"])
        rev_sum = sum([flt(r.custom_po_value if cat == "Order" else r.custom_expected_revenue) for r in recs])
        res[cat] = {"count": len(recs), "revenue": rev_sum}
    return res

@frappe.whitelist()
def get_crm_users():
    """Returns users having 'DAC CRM' or 'DAC CRM Head' roles with their Full Names."""
    return frappe.db.sql("""
        SELECT DISTINCT
            u.name as user,
            u.full_name as user_name
        FROM
            `tabUser` u
        INNER JOIN
            `tabHas Role` hr ON u.name = hr.parent
        WHERE
            hr.role IN ('DAC CRM', 'DAC CRM Head')
            AND u.enabled = 1
            AND u.name NOT IN ('Administrator', 'Guest')
        ORDER BY u.full_name ASC
    """, as_dict=True)
# -----------------------
# Utility: role-based filters
# -----------------------
def get_filters_for(doctype):
    """Return role-based filters for the given doctype."""
    if "DAC CRM Head" in frappe.get_roles(frappe.session.user):
        return {}  # no restrictions

    if doctype == "Lead":
        return {"lead_owner": frappe.session.user}

    if doctype == "Event Activity":
        return {"assigned_to": frappe.session.user}

    return {}


# def get_bc_without_activity_count():
#     """Returns count of Open Business Contacts with no linked Event Activity."""
#     is_crm_head = "DAC CRM Head" in frappe.get_roles()
#     user = frappe.session.user

#     # Logic: Status is 'Open' and Name does NOT exist in Event Activity's reference_name
#     query = """
#         SELECT count(name) 
#         FROM `tabBusiness Contacts` bc
#         WHERE bc.status = 'Open'
#     """
    
#     # Filter by owner if not CRM Head
#     if not is_crm_head:
#         # Assuming Business Contact has an 'owner' or 'lead_owner' field. 
#         # Using 'owner' (the creator) as standard. Change to specific field if different.
#         query += f" AND bc.assign_to = '{user}'"

#     query += """
#         AND bc.name NOT IN (
#             SELECT DISTINCT reference_name 
#             FROM `tabEvent Activity` 
#             WHERE reference_type = 'Business Contact'
#         )
#     """

#     count = frappe.db.sql(query)[0][0]
#     return count
import frappe
import json

@frappe.whitelist()
def get_bc_without_activity_data(filters=None):
    """
    Returns Business Contacts with status 'Open' that have 
    ZERO entries in the Event Activity table (neglected contacts).
    """
    # 1. Parse Filters
    if isinstance(filters, str):
        filters = frappe._dict(json.loads(filters))
    elif not filters:
        filters = frappe._dict()

    # 2. Determine User Scope
    user_roles = frappe.get_roles(frappe.session.user)
    is_crm_head = "DAC CRM Head" in user_roles
    
    target_user = filters.get("user")
    if not target_user and not is_crm_head:
        target_user = frappe.session.user

    # 3. Build Query Conditions
    # We only look for Business Contacts currently in 'Open' status
    conditions = ["bc.status = 'Open'"]
    query_params = []

    # Filter by Assignment
    if target_user:
        conditions.append("bc.assign_to = %s")
        query_params.append(target_user)

    # Dashboard Date Filtering (usually based on when the BC was created)
    if filters.get("from_date") and filters.get("to_date"):
        conditions.append("bc.creation BETWEEN %s AND %s")
        query_params.extend([filters.from_date, filters.to_date])

    where_clause = " AND ".join(conditions)

    # 4. The Logic: NOT EXISTS
    # This finds BCs where there is NO record in tabEvent Activity 
    # regardless of that activity's status.
    sql_query = f"""
        SELECT bc.name 
        FROM `tabBusiness Contacts` bc 
        WHERE {where_clause}
        AND NOT EXISTS (
            SELECT 1 
            FROM `tabEvent Activity` ea 
            WHERE ea.reference_type = 'Business Contacts' 
            AND ea.reference_name = bc.name
        )
    """
    
    names = frappe.db.sql(sql_query, tuple(query_params), pluck=True)
    
    return {
        "count": len(names) or 0,
        "names": list(names) if names else []
    }
# -----------------------
# Counts
# -----------------------
import frappe
import frappe

@frappe.whitelist()
def get_leads_without_activity_count():
    """Return count of Leads that do NOT have any associated Event Activity."""

    # Get the current session user
    user = frappe.session.user
    roles = frappe.get_roles(user)

    filters = {
        "custom_lead_category": ["in", ["Enquiry", "Pipeline"]]
    }

    # If user has DAC CRM role, restrict to leads they own
    if "DAC CRM" in roles:
        filters["lead_owner"] = user

    # Get all leads that are referenced in Event Activity
    leads_with_activity = frappe.db.get_all(
        "Event Activity",
        fields=["reference_name"],
        filters={"reference_type": "Lead"},
        pluck="reference_name"
    )

    # If there are leads with activity, exclude them
    if leads_with_activity:
        filters["name"] = ["not in", leads_with_activity]

    # Count leads with the applied filters
    count = frappe.db.count("Lead", filters=filters)

    return count




def get_today_upcoming_follow_ups_count():
    """Counts Open Event Activities for today + next 3 days."""
    today = getdate(nowdate())
    end_date = add_days(today, 3)

    filters = get_filters_for("Event Activity")
    filters.update({
        "status": "Open",
        "starts_on": ["between", [f"{today} 00:00:00", f"{end_date} 23:59:59"]]
    })

    return frappe.db.count("Event Activity", filters=filters)


def get_overdue_activities_count():
    """Counts Open Event Activities that are overdue (starts_on < today)."""
    today = getdate(nowdate())

    filters = get_filters_for("Event Activity")
    filters.update({
        "status": "Open",
        "starts_on": ["<", f"{today} 00:00:00"]
    })

    return frappe.db.count("Event Activity", filters=filters)

def get_open_activities_count():
    """Counts Lead Activities with a status of 'Open'."""
    filters = get_filters_for("Event Activity")
    filters["status"] = "Open"

    return frappe.db.count("Event Activity", filters=filters)


# def get_lead_category_counts():
#     """Counts Leads grouped by their custom_lead_category and sums custom_expected_revenue."""
#     categories = [
#         'Enquiry', 'Pipeline', 'Order', 'Lost Enquiry', 'Lost Pipeline'
#     ]
#     counts = {}

#     for category in categories:
#         filters = get_filters_for("Lead")
#         filters["custom_lead_category"] = category

#         # Count the leads
#         lead_count = frappe.db.count("Lead", filters=filters)

#         # Sum the custom_expected_revenue
#         total_revenue = frappe.db.get_all(
#             "Lead",
#             filters=filters,
#             fields=["custom_expected_revenue"]
#         )
#         revenue_sum = sum([float(l.custom_expected_revenue or 0) for l in total_revenue])

#         counts[category] = {
#             "count": lead_count,
#             "revenue": revenue_sum
#         }

#     return counts

import frappe

@frappe.whitelist()
def get_lead_category_counts():
    """Counts Leads grouped by their custom_lead_category and sums revenue fields.
       For 'Order' category → sum custom_po_value
       For others → sum custom_expected_revenue
    """

    categories = ['Enquiry', 'Pipeline', 'Order', 'Lost Enquiry', 'Lost Pipeline']
    counts = {}

    for category in categories:
        filters = get_filters_for("Lead")
        filters["custom_lead_category"] = category

        # Count total leads in this category
        lead_count = frappe.db.count("Lead", filters=filters)

        # Choose revenue field dynamically
        revenue_field = "custom_po_value" if category == "Order" else "custom_expected_revenue"

        # Fetch and sum revenue
        total_revenue_records = frappe.db.get_all(
            "Lead",
            filters=filters,
            fields=[revenue_field]
        )

        revenue_sum = sum([float(l[revenue_field] or 0) for l in total_revenue_records])

        counts[category] = {
            "count": lead_count,
            "revenue": revenue_sum
        }

    return counts



@frappe.whitelist()
def get_followup_summary(session_user=None):
    user = session_user or frappe.session.user
    roles = frappe.get_roles(user)
    is_admin = "DAC CRM Head" in roles

    # Get today's start and end datetime
    today = frappe.utils.today()
    start = f"{today} 00:00:00"
    end = f"{today} 23:59:59"

    filters = {
        "starts_on": ["between", [start, end]],
        "status": "Open"
    }

    if not is_admin:
        filters["assigned_to"] = user

    activities = frappe.get_all(
        "Event Activity",
        filters=filters,
        fields=["category", "assigned_to"]
    )

    summary = {}
    all_categories = set()

    for a in activities:
        # Convert user id → full name
        u = frappe.utils.get_fullname(a.assigned_to) if a.assigned_to else "Unassigned"
        cat = a.category or "Uncategorized"
        all_categories.add(cat)

        if u not in summary:
            summary[u] = {}
        summary[u][cat] = summary[u].get(cat, 0) + 1

    return {
        "users": list(summary.keys()),          # full names now
        "categories": sorted(c for c in all_categories if c),
        "matrix": summary                       # keyed by full name
    }
import frappe
from frappe.utils import nowdate

@frappe.whitelist()
def get_followup_report(from_date=None, to_date=None):
    """
    Returns user-wise summary of Event Activity with Open/Closed counts,
    filterable by custom date range.
    """
    try:
        user = frappe.session.user
        roles = frappe.get_roles(user)
        is_admin = "DAC CRM Head" in roles

        # Default filters
        filters = {}
        if from_date and to_date:
            filters["created_on"] = ["between", [from_date, to_date]]

        # Non-admin users see only their assigned activities
        if not is_admin:
            filters["assigned_to"] = user

        activities = frappe.get_all(
            "Event Activity",
            filters=filters,
            fields=["category", "assigned_to", "status"]
        )

        summary = {}
        all_categories = set()

        for a in activities:
            u = a.assigned_to or "Unassigned"
            cat = a.category or "Uncategorized"
            all_categories.add(cat)

            if u not in summary:
                full_name = frappe.db.get_value("User", u, "full_name") if u != "Unassigned" else "Unassigned"
                summary[u] = {"name": full_name, "categories": {}}

            if cat not in summary[u]["categories"]:
                summary[u]["categories"][cat] = {"Open": 0, "Closed": 0}

            # ✅ Map Completed → Closed
            status = "Closed" if a.status == "Completed" else "Open"
            summary[u]["categories"][cat][status] += 1

        return {
            "users": [
                {"id": uid, "name": data["name"], "categories": data["categories"]}
                for uid, data in summary.items()
            ],
            "activity_categories": sorted(all_categories),
            "is_admin": is_admin
        }

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "get_followup_report Error")
        frappe.throw(f"Error fetching followup report: {str(e)}")

# import frappe
# from frappe.utils import getdate
# @frappe.whitelist()
# def get_lead_category_report(from_date=None, to_date=None):
#     user = frappe.session.user
#     roles = frappe.get_roles(user)
#     is_admin = "DAC CRM Head" in roles

#     lead_filters = {}
#     if from_date and to_date:
#         lead_filters["custom_created_at"] = ["between", [from_date, to_date]]

#     if not is_admin:

#         lead_filters["lead_owner"] = user

#     # --- Fetch Leads ---
#     leads = frappe.get_all(
#         "Lead",
#         filters=lead_filters,
#         fields=[
#             "name", "lead_owner", "lead_name", "status", "custom_lead_category",
#             "custom_lead_type", "custom_direction_type", "company_name",
#             "industry", "custom_expected_revenue", "custom_expected_closure_date", "mobile_no","custom_po_value",
#         ]

#     )

#     summary = {}
#     all_categories, all_lead_types, all_direction_types, all_industries, all_products, all_closure_months = set(), set(), set(), set(), set(), set()

#     for l in leads:
#         u = l.lead_owner
#         cat = l.custom_lead_category or "Uncategorized"
#         lead_type = l.custom_lead_type or "Unknown"
#         direction = l.custom_direction_type or "Unknown"
#         industry = l.industry or "Unknown"

#         if lead_type == "WON":
#             revenue = float(l.custom_po_value or 0)
#         else:
#             revenue = float(l.custom_expected_revenue or 0)

#         # ✅ Revenue logic
#         if cat == "Order":
#             revenue = float(l.custom_po_value or 0)
#         else:
#             revenue = float(l.custom_expected_revenue or 0)
#         closure_month = getdate(l.custom_expected_closure_date).strftime("%b %Y") if l.custom_expected_closure_date else "No Closure Date"

#         all_categories.add(cat)
#         all_lead_types.add(lead_type)
#         all_direction_types.add(direction)
#         all_industries.add(industry)
#         all_closure_months.add(closure_month)

#         if u not in summary:
#             full_name = frappe.db.get_value("User", u, "full_name") or "Unknown"
#             summary[u] = {
#                 "name": full_name,
#                 "categories": {},
#                 "lead_types": {},
#                 "direction_types": {},
#                 "industries": {},
#                 "products": {},
#                 "closures": {},
#                 "quotations": {}
#             }

#         def add_to_bucket(bucket, key):
#             if key not in summary[u][bucket]:
#                 summary[u][bucket][key] = {"count": 0, "revenue": 0, "leads": []}
#             summary[u][bucket][key]["count"] += 1
#             summary[u][bucket][key]["revenue"] += revenue
#             summary[u][bucket][key]["leads"].append({
#                 "Lead ID": l.name,
#                 "Full Name": l.lead_name,
#                 "Organization Name": l.company_name or "",
#                 "Lead Category": l.custom_lead_category or "",
#                 "Expected Revenue": revenue,
#                 "Mobile No": l.mobile_no or "",
#                 "Expected Closure Month": closure_month
#             })

#         add_to_bucket("categories", cat)
#         add_to_bucket("lead_types", lead_type)
#         add_to_bucket("direction_types", direction)
#         add_to_bucket("industries", industry)
#         add_to_bucket("closures", closure_month)

#         # --- Products child table ---
#         product_rows = frappe.get_all(
#             "Product Category Multiselect",
#             filters={"parent": l.name},
#             fields=["product_category"]
#         )
#         for row in product_rows:
#             if row.product_category:
#                 all_products.add(row.product_category)
#                 add_to_bucket("products", row.product_category)

#     # --- Fetch Quotations ---
#     quotation_filters = {"docstatus": 1}
#     if from_date and to_date:
#         quotation_filters["transaction_date"] = ["between", [from_date, to_date]]

#     quotations = frappe.get_all(
#         "Quotation",
#         filters=quotation_filters,
#         fields=[
#             "name", "owner", "quotation_to", "party_name",
#             "grand_total", "transaction_date", "status","customer_name"
#         ],
#         order_by="transaction_date desc"
#     )

#     all_quotation_labels = set()

#     for q in quotations:
#         owner = q.owner or "Unknown"
#         label = q.quotation_to or "Unknown"
#         revenue = float(q.grand_total or 0)
#         all_quotation_labels.add(label)

#         if owner not in summary:
#             full_name = frappe.db.get_value("User", owner, "full_name") or "Unknown"
#             summary[owner] = {
#                 "name": full_name,
#                 "categories": {},
#                 "lead_types": {},
#                 "direction_types": {},
#                 "industries": {},
#                 "products": {},
#                 "closures": {},
#                 "quotations": {}
#             }

#         if label not in summary[owner]["quotations"]:
#             summary[owner]["quotations"][label] = {"count": 0, "revenue": 0, "quotations": []}

#         summary[owner]["quotations"][label]["count"] += 1
#         summary[owner]["quotations"][label]["revenue"] += revenue
#         summary[owner]["quotations"][label]["quotations"].append({
#             "Quotation ID": q.name,
#             "Party Name": q.party_name or "",
#             "Customer Name":q.customer_name or "",
#             "Quotation Type": q.quotation_to or "",
#             "Grand Total": revenue,
#             "Status": q.status or "",
#             "Transaction Date": q.transaction_date.strftime("%d-%b-%Y") if q.transaction_date else ""
#         })
#     # print(summary)
#     # --- Sorting ---
#     def sort_by_order(values, order):
#         return sorted(values, key=lambda x: order.index(x) if x in order else 99)

#     lead_category_order = ["Enquiry", "Pipeline", "Order", "Lost Enquiry", "Lost Pipeline"]
#     lead_type_order = ["HOT", "WARM", "COLD", "WON", "LOST"]
#     direction_order = ["Inbound", "Outbound"]

#     return {
#         "users": [
#             {
#                 "id": uid,
#                 "name": data["name"],
#                 "categories": data["categories"],
#                 "lead_types": data["lead_types"],
#                 "direction_types": data["direction_types"],
#                 "industries": data["industries"],
#                 "products": data["products"],
#                 "closures": data["closures"],
#                 "quotations": data["quotations"]
#             }
#             for uid, data in summary.items()
#         ],
#         "lead_categories": sort_by_order(all_categories, lead_category_order),
#         "lead_types": sort_by_order(all_lead_types, lead_type_order),
#         "direction_types": sort_by_order(all_direction_types, direction_order),
#         "industries": sorted(all_industries),
#         "products": sorted(all_products),
#         "month_year_closure": sorted(
#             all_closure_months,
#             key=lambda x: getdate("01-" + x.split(" ")[0] + "-" + x.split(" ")[1])
#             if x != "No Closure Date" else getdate()
#         ),
#         "quotation_labels": sorted(all_quotation_labels),
#         "is_admin": is_admin
#     }

@frappe.whitelist()
def get_lead_category_report(from_date=None, to_date=None):
    user = frappe.session.user
    roles = frappe.get_roles(user)
    is_admin = "DAC CRM Head" in roles or "Administrator" in roles

    lead_filters = {}
    if from_date and to_date:
        lead_filters["custom_created_at"] = ["between", [from_date, to_date]]

    if not is_admin:
        lead_filters["lead_owner"] = user

    # --- Fetch Leads ---
    leads = frappe.get_all(
        "Lead",
        filters=lead_filters,
        fields=[
            "name", "lead_owner", "lead_name", "status", "custom_lead_category",
            "custom_lead_type", "custom_direction_type", "company_name",
            "industry", "custom_expected_revenue", "custom_expected_closure_date", "mobile_no","custom_po_value",
        ]
    )

    summary = {}
    all_categories, all_lead_types, all_direction_types, all_industries, all_products, all_closure_months = set(), set(), set(), set(), set(), set()

    # Define Category Groupings for the buckets
    OPEN_CATS = ["Enquiry", "Pipeline"]
    ORDER_CATS = ["Order"]
    LOST_CATS = ["Lost Enquiry", "Lost Pipeline"]

    for l in leads:
        u = l.lead_owner
        cat = l.custom_lead_category or "Uncategorized"
        lead_type = l.custom_lead_type or "Unknown"
        direction = l.custom_direction_type or "Unknown"
        industry = l.industry or "Unknown"

        # Determine the group_key for the bucket update
        if cat in OPEN_CATS: 
            group_key = "open"
        elif cat in ORDER_CATS: 
            group_key = "order"
        elif cat in LOST_CATS: 
            group_key = "lost"
        else: 
            group_key = "other"

        if lead_type == "WON":
            revenue = float(l.custom_po_value or 0)
        else:
            revenue = float(l.custom_expected_revenue or 0)

        # ✅ Revenue logic
        if cat == "Order":
            revenue = float(l.custom_po_value or 0)
        else:
            revenue = float(l.custom_expected_revenue or 0)
            
        closure_month = getdate(l.custom_expected_closure_date).strftime("%b %Y") if l.custom_expected_closure_date else "No Closure Date"

        all_categories.add(cat)
        all_lead_types.add(lead_type)
        all_direction_types.add(direction)
        all_industries.add(industry)
        all_closure_months.add(closure_month)

        if u not in summary:
            full_name = frappe.db.get_value("User", u, "full_name") or "Unknown"
            summary[u] = {
                "name": full_name,
                "categories": {},
                "lead_types": {},
                "direction_types": {},
                "industries": {},
                "products": {},
                "closures": {},
                "quotations": {}
            }

        def add_to_bucket(bucket, key, group_key, revenue):
            if key not in summary[u][bucket]:
                summary[u][bucket][key] = {
                    "open_count": 0, "open_revenue": 0,
                    "order_count": 0, "order_revenue": 0,
                    "lost_count": 0, "lost_revenue": 0,
                    "count": 0, "revenue": 0,
                    "leads": []
                }
            
            # Update Grouped Metrics
            if group_key != "other":
                summary[u][bucket][key][f"{group_key}_count"] += 1
                summary[u][bucket][key][f"{group_key}_revenue"] += revenue
            
            # Update Total Metrics
            summary[u][bucket][key]["count"] += 1
            summary[u][bucket][key]["revenue"] += revenue
            
            summary[u][bucket][key]["leads"].append({
                "Lead ID": l.name,
                "Full Name": l.lead_name,
                "Organization Name": l.company_name or "",
                "Lead Category": l.custom_lead_category or "",
                "Expected Revenue": revenue,
                "Mobile No": l.mobile_no or "",
                "Expected Closure Month": closure_month
            })

        add_to_bucket("categories", cat, group_key, revenue)
        add_to_bucket("lead_types", lead_type, group_key, revenue)
        add_to_bucket("direction_types", direction, group_key, revenue)
        add_to_bucket("industries", industry, group_key, revenue)
        add_to_bucket("closures", closure_month, group_key, revenue)

        # --- Products child table ---
        product_rows = frappe.get_all(
            "Product Category Multiselect",
            filters={"parent": l.name},
            fields=["product_category"]
        )
        for row in product_rows:
            if row.product_category:
                all_products.add(row.product_category)
                add_to_bucket("products", row.product_category, group_key, revenue)

    # --- Fetch Quotations ---
    quotation_filters = {"docstatus": 1}
    if from_date and to_date:
        quotation_filters["transaction_date"] = ["between", [from_date, to_date]]

    quotations = frappe.get_all(
        "Quotation",
        filters=quotation_filters,
        fields=[
            "name", "owner", "quotation_to", "party_name",
            "grand_total", "transaction_date", "status","customer_name"
        ],
        order_by="transaction_date desc"
    )

    all_quotation_labels = set()

    for q in quotations:
        owner = q.owner or "Unknown"
        label = q.quotation_to or "Unknown"
        revenue = float(q.grand_total or 0)
        all_quotation_labels.add(label)

        if owner not in summary:
            full_name = frappe.db.get_value("User", owner, "full_name") or "Unknown"
            summary[owner] = {
                "name": full_name,
                "categories": {},
                "lead_types": {},
                "direction_types": {},
                "industries": {},
                "products": {},
                "closures": {},
                "quotations": {}
            }

        if label not in summary[owner]["quotations"]:
            summary[owner]["quotations"][label] = {"count": 0, "revenue": 0, "quotations": []}

        summary[owner]["quotations"][label]["count"] += 1
        summary[owner]["quotations"][label]["revenue"] += revenue
        summary[owner]["quotations"][label]["quotations"].append({
            "Quotation ID": q.name,
            "Party Name": q.party_name or "",
            "Customer Name":q.customer_name or "",
            "Quotation Type": q.quotation_to or "",
            "Grand Total": revenue,
            "Status": q.status or "",
            "Transaction Date": q.transaction_date.strftime("%d-%b-%Y") if q.transaction_date else ""
        })

    # --- Sorting ---
    def sort_by_order(values, order):
        return sorted(values, key=lambda x: order.index(x) if x in order else 99)

    lead_category_order = ["Enquiry", "Pipeline", "Order", "Lost Enquiry", "Lost Pipeline"]
    lead_type_order = ["HOT", "WARM", "COLD", "WON", "LOST"]
    direction_order = ["Inbound", "Outbound"]

    return {
        "users": [
            {
                "id": uid,
                "name": data["name"],
                "categories": data["categories"],
                "lead_types": data["lead_types"],
                "direction_types": data["direction_types"],
                "industries": data["industries"],
                "products": data["products"],
                "closures": data["closures"],
                "quotations": data["quotations"]
            }
            for uid, data in summary.items()
        ],
        "lead_categories": sort_by_order(all_categories, lead_category_order),
        "lead_types": sort_by_order(all_lead_types, lead_type_order),
        "direction_types": sort_by_order(all_direction_types, direction_order),
        "industries": sorted(all_industries),
        "products": sorted(all_products),
        "month_year_closure": sorted(
            all_closure_months,
            key=lambda x: getdate("01-" + x.split(" ")[0] + "-" + x.split(" ")[1])
            if x != "No Closure Date" else getdate()
        ),
        "quotation_labels": sorted(all_quotation_labels),
        "is_admin": is_admin
    }

# # Separate PDF generation function
# import frappe
# from frappe.utils.pdf import get_pdf

# @frappe.whitelist()
# def generate_lead_report_pdf(from_date=None, to_date=None, quotation_from_date=None, quotation_to_date=None):
#     # Get the data using your existing function
#     report_data = get_lead_category_report(
#         from_date=from_date,
#         to_date=to_date,
#         quotation_from_date=quotation_from_date,
#         quotation_to_date=quotation_to_date
#     )
    
#     # Prepare HTML from the data (simple table example)
#     html = "<h2>Lead Report</h2><table border='1' cellspacing='0' cellpadding='4'>"
#     html += "<tr><th>User</th><th>Category</th><th>Count</th><th>Revenue</th></tr>"
#     for u in report_data["users"]:
#         for cat, val in u["categories"].items():
#             html += f"<tr><td>{u['name']}</td><td>{cat}</td><td>{val['count']}</td><td>{val['revenue']}</td></tr>"
#     html += "</table>"

#     # Generate PDF
#     pdf = get_pdf(html)
#     return pdf
# import frappe
# from frappe.utils import getdate

# @frappe.whitelist()
# def get_leads_by_user_and_label(user, label, fieldname):
#     filters = {"lead_owner": user}

#     lead_fields = [
#         "name", "lead_name", "status", "custom_lead_category",
#         "custom_expected_revenue", "custom_expected_closure_date"
#     ]

#     if fieldname == "categories":
#         filters["custom_lead_category"] = label
#     elif fieldname == "lead_types":
#         filters["custom_lead_type"] = label
#     elif fieldname == "industries":
#         filters["industry"] = label
#     elif fieldname == "direction_types":
#         filters["custom_direction_type"] = label
#     elif fieldname == "products":
#         # Fetch leads which have this product category in child table
#         product_links = frappe.get_all(
#             "Product Category Multiselect",
#             filters={"product_category": label},
#             fields=["parent"]
#         )
#         lead_names = [p["parent"] for p in product_links]
#         if lead_names:
#             filters = {"name": ["in", lead_names], "lead_owner": user}
#         else:
#             return []  # No leads found
#     else:
#         return []

#     leads = frappe.get_all(
#         "Lead",
#         filters=filters,
#         fields=lead_fields
#     )

#     for l in leads:
#         # Add products (child table)
#         l["products"] = frappe.get_all(
#             "Product Category Multiselect",
#             filters={"parent": l["name"]},
#             fields=["product_category"]
#         )

#         # Format expected closure month
#         closure_date = l.get("custom_expected_closure_date")
#         if closure_date:
#             # Convert string to date and format as "Mon YYYY"
#             l["closure_month"] = getdate(closure_date).strftime("%b %Y")
#         else:
#             l["closure_month"] = "No Closure Date"

#         # Fetch related Quotations for this lead (optional)
#         quotations = frappe.get_all(
#             "Quotation",
#             filters={"quotation_to": l.name, "docstatus": 1},
#             fields=["name", "grand_total", "transaction_date"]
#         )
#         l["quotations"] = quotations

#     return leads



@frappe.whitelist()
def get_lead_quotation_details(lead_name):
    """Quotations behind the drilldown's Quotations count.

    Deliberately separate from get_quotations_for_lead (which lead.js relies on):
    the drilldown counts quotations matching party_name OR custom_lead_id, so this
    must use the same filter or the number and the list it opens would disagree.
    """
    if not frappe.has_permission("Quotation", "read"):
        frappe.throw(_("You do not have permission to access Quotation records."),
                     frappe.PermissionError)
    rows = frappe.db.sql("""
        SELECT q.name, q.status, q.customer_name, q.party_name, q.currency,
               DATE_FORMAT(q.transaction_date, '%%d-%%b-%%Y') AS transaction_date,
               q.grand_total, q.valid_till
        FROM `tabQuotation` q
        WHERE q.docstatus < 2
          AND (q.party_name = %(lead)s OR q.custom_lead_id = %(lead)s)
        ORDER BY q.transaction_date DESC, q.name DESC
    """, {"lead": lead_name}, as_dict=1)
    for r in rows:
        r["items"] = [i.item_code for i in frappe.get_all(
            "Quotation Item", fields=["item_code"], filters={"parent": r["name"]})]
    return rows


@frappe.whitelist()
def get_quotations_for_lead(lead_name):
    if not frappe.has_permission('Quotation', 'read'):
        frappe.throw(_("You do not have permission to access Quotation records."), frappe.PermissionError)
    
    quotations = frappe.get_all(
        'Quotation',
        fields=['name', 'quotation_to', 'customer_name', 'transaction_date', 'grand_total','status'],
        filters={'custom_lead_id': lead_name}
    )
    
    quotation_data = []
    
    for quotation in quotations:
        items = frappe.get_all(
            'Quotation Item',
            fields=['item_code'],
            filters={'parent': quotation.name}
        )
        # print(items)
        # Append the data as a dictionary
        quotation_data.append({
            'quotation_id': quotation.name,
            'status': quotation.status,
            'customer_name': quotation.customer_name,
            'transaction_date': quotation.transaction_date,
            'grand_total': quotation.grand_total,
            'items': [item.item_code for item in items]
        })
    
    return quotation_data


@frappe.whitelist()
def update_expected_revenue_from_quotation(lead_name):
    """
    Update the Expected Revenue field of a Lead based on the most recent submitted quotation.
    Called from list view for existing leads that need revenue sync.
    """
    from frappe.utils import flt
    
    if not frappe.has_permission('Lead', 'write'):
        frappe.throw(_("You do not have permission to update Lead records."), frappe.PermissionError)
    
    quotations = frappe.get_all(
        'Quotation',
        fields=['name', 'grand_total', 'transaction_date'],
        filters={
            'custom_lead_id': lead_name,
            'docstatus': 1
        },
        order_by='transaction_date desc'
    )
    
    if quotations:
        latest_quotation = quotations[0]
        frappe.db.set_value("Lead", lead_name, "custom_expected_revenue", flt(latest_quotation.grand_total))
        return {"status": "success", "message": "Expected Revenue updated from Quotation " + latest_quotation.name}
    
    return {"status": "info", "message": "No submitted quotations found for this lead"}


@frappe.whitelist()
def update_all_leads_expected_revenue():
    """
    Bulk update all leads' expected revenue from their most recent quotations.
    Returns count of updated leads.
    """
    from frappe.utils import flt
    
    leads = frappe.get_all("Lead", filters={"custom_lead_category": ["in", ["Enquiry", "Pipeline"]]}, pluck="name")
    
    updated_count = 0
    for lead_name in leads:
        result = update_expected_revenue_from_quotation(lead_name)
        if result.get("status") == "success":
            updated_count += 1
    
    return updated_count


from frappe.utils.pdf import get_pdf
from frappe.utils import now_datetime
import frappe

@frappe.whitelist()
def generate_combined_report_pdf(from_date=None, to_date=None):
    try:
        # Render HTML using existing functions
        followup_html = render_followup_html(from_date, to_date)
        lead_html = render_lead_html(from_date, to_date)

        # Full HTML
        full_html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: 'Segoe UI', sans-serif; font-size: 11pt; }}
                .summary-header {{ font-size: 16pt; font-weight: 600; margin: 15px 0; color: #333; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
                th {{ background: #1976d2; color: white; }}
                tbody tr:nth-child(even) {{ background: #f9f9f9; }}
                tbody tr:hover {{ background: #f1faff; }}
                tfoot td {{ font-weight: bold; background: #e0e0e0; }}
            </style>
        </head>
        <body>
            <h2>Event Activities Report</h2>
            {followup_html}
            <h2>Lead Report</h2>
            {lead_html}
        </body>
        </html>
        """

        pdf = get_pdf(full_html)

        # Save PDF
        filename = f"Lead_Followup_Report_{now_datetime().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = frappe.get_site_path("public", "files", filename)
        with open(file_path, "wb") as f:
            f.write(pdf)

        # Return public URL
        return {"pdf_file": f"/files/{filename}"}

    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "generate_combined_report_pdf Error")
        frappe.throw(f"Error generating PDF: {str(e)}")


def render_followup_html(from_date=None, to_date=None):
    data = get_followup_report(from_date, to_date)
    users = data.get("users", [])
    categories = data.get("activity_categories", [])

    if not users:
        return "<p>No Event Activities found.</p>"

    html = "<table><thead><tr><th>User</th>"
    for c in categories:
        html += f"<th>{c}</th>"
    html += "<th>Total</th></tr></thead><tbody>"

    for u in users:
        row = f"<tr><td>{u['name']}</td>"
        total = 0
        for c in categories:
            val = u["categories"].get(c, {"Open": 0, "Closed": 0})
            count = val.get("Open", 0) + val.get("Closed", 0)
            total += count
            row += f"<td>{count}</td>"
        row += f"<td>{total}</td></tr>"
        html += row

    html += "</tbody></table>"
    return html


def render_lead_html(from_date=None, to_date=None):
    data = get_lead_category_report(from_date, to_date)
    users = data.get("users", [])
    categories = data.get("lead_categories", [])

    if not users:
        return "<p>No Lead records found.</p>"

    html = "<table><thead><tr><th>User</th>"
    for c in categories:
        html += f"<th>{c}</th>"
    html += "<th>Total</th></tr></thead><tbody>"

    for u in users:
        row = f"<tr><td>{u['name']}</td>"
        total = 0
        for c in categories:
            val = u["categories"].get(c, {"count": 0})
            total += val.get("count", 0)
            row += f"<td>{val.get('count', 0)}</td>"
        row += f"<td>{total}</td></tr>"
        html += row

    html += "</tbody></table>"
    return html



import frappe
from frappe import _

@frappe.whitelist(allow_guest=True)
def create_duplicate_lead(lead_name):
    """
    Duplicates the lead, resets Workflow/Status to 'Enquiry', 
    and links the correct Customer.
    """
    if not lead_name:
        return

    # 1. Fetch the original Lead
    original_doc = frappe.get_doc("Lead", lead_name)
    
    # 2. Determine the Customer to set
    customer_id = original_doc.custom_lead_customer
    if not customer_id:
        customer_id = frappe.db.get_value("Customer", {"lead_name": lead_name}, "name")

    # 3. Create the copy (this copies 'Order' status too!)
    new_doc = frappe.copy_doc(original_doc)
    
    # --- FIX START ---
    
    # 4. Force Reset Status / Category to Initial State
    # Change "Enquiry" to whatever your Workflow's actual Start State is (e.g., "Pipeline", "Open")
    initial_state = "Enquiry" 
    
    new_doc.custom_lead_category = initial_state
    new_doc.custom_duplicated_lead = 1
    new_doc.custom_duplicate_lead_id = original_doc.name
    new_doc.custom_lead_type = 'HOT'
    new_doc.custom_po_value = ''
    # IMPORTANT: You must also reset the specific Workflow State field
    # Standard field is 'workflow_state', but checking meta is safer
    # workflow_state_field = frappe.get_meta("Lead").get_workflow_state_field()
    # if workflow_state_field:
    #     new_doc.set(workflow_state_field, initial_state)

    # --- FIX END ---

    # 5. Set the determined customer
    new_doc.custom_lead_customer = customer_id

    # 6. Insert the new document
    new_doc.insert(ignore_permissions=True)
    
    return new_doc.name

import frappe
from erpnext.crm.doctype.lead.lead import make_quotation as original_make_quotation

@frappe.whitelist()
def make_quotation_custom(source_name, target_doc=None):
    """
    Custom mapper to create a Quotation.
    1. Calls standard ERPNext mapper.
    2. Checks if 'custom_duplicated_lead' is checked AND 'custom_lead_customer' exists.
    3. If yes, changes Quotation To -> Customer.
    4. Sets custom_lead_id on Quotation.
    """
    
    # 1. Create the Quotation object using standard ERPNext logic
    doc = original_make_quotation(source_name, target_doc)

    # 2. Fetch Lead details (Added 'custom_duplicated_lead' to the fetch list)
    lead = frappe.db.get_value("Lead", source_name, ["custom_lead_customer", "custom_duplicated_lead"], as_dict=True)

    # 3. Logic to switch to Customer
    # Only proceed if customer exists AND the duplicate checkbox is checked (1)
    if lead and lead.custom_lead_customer and lead.custom_duplicated_lead:
        doc.quotation_to = "Customer"
        doc.party_name = lead.custom_lead_customer
        
        # Optional: Fetch and set the Customer Name for display purposes
        doc.customer_name = frappe.db.get_value("Customer", lead.custom_lead_customer, "customer_name")

    return doc

import datetime


# Every live Quotation status, in pipeline order. Cancelled is excluded because
# those documents are docstatus 2 and are not "created" work.
# Reports break the created total down across ALL of these so the parts always add
# up to the whole — showing only Draft/Open/Ordered hid Expired and Lost entirely.
QUOTATION_STATUSES = ["Draft", "Open", "Replied", "Partially Ordered", "Ordered", "Lost", "Expired"]


def _q_status_key(status):
    """Column key for a status, e.g. 'Partially Ordered' -> 'q_st_partially_ordered'."""
    return "q_st_" + status.lower().replace(" ", "_")


def _add_quotation_status_counts(target, where_sql, params):
    """
    Add one q_st_<status> count per live status onto `target`, from a single
    grouped query. The parts always sum to the created total, so the report can
    never show a total that its own breakdown contradicts.
    """
    for st in QUOTATION_STATUSES:
        target[_q_status_key(st)] = 0
    rows = frappe.db.sql(
        f"SELECT status, COUNT(*) FROM `tabQuotation` {where_sql} GROUP BY status", params) or []
    for status, n in rows:
        key = _q_status_key(status or "")
        if key in target:
            target[key] = n or 0
        else:
            target["q_st_other"] = target.get("q_st_other", 0) + (n or 0)
    return target



def _fmt_dmy(value):
    """dd-mmm-yyyy — the only date format used in the report e-mails."""
    if not value:
        return ""
    d = frappe.utils.getdate(value)
    return d.strftime("%d-%b-%Y")


def _fmt_dmy_day(value):
    """Sun, 02-Aug-2026 — a date the reader can place without counting back."""
    if not value:
        return ""
    d = frappe.utils.getdate(value)
    return "{}, {}".format(d.strftime("%a"), d.strftime("%d-%b-%Y"))


def report_period_range(report_type, on_date=None):
    """
    The window a report covers.

        daily    -> that day
        weekly   -> Sunday to Friday of that week (the working week here)
        monthly  -> first to last day of that month

    Returns (from_date, to_date) as date objects.
    """
    import datetime as _dt

    ref = frappe.utils.getdate(on_date or frappe.utils.nowdate())
    rt = (report_type or "daily").lower()

    if rt == "weekly":
        # weekday(): Mon=0 .. Sun=6. Step back to the most recent Sunday.
        days_since_sunday = (ref.weekday() + 1) % 7
        start = ref - _dt.timedelta(days=days_since_sunday)
        return start, start + _dt.timedelta(days=5)      # Sunday -> Friday

    if rt == "monthly":
        start = ref.replace(day=1)
        nxt = (start + _dt.timedelta(days=32)).replace(day=1)
        return start, nxt - _dt.timedelta(days=1)

    return ref, ref


def report_period_label(report_type, on_date=None):
    """'Sun, 02-Aug-2026 to Fri, 07-Aug-2026' — or a single day for daily."""
    frm, to = report_period_range(report_type, on_date)
    if frm == to:
        return _fmt_dmy_day(frm)
    return "{} to {}".format(_fmt_dmy_day(frm), _fmt_dmy_day(to))


def report_subject(report_type, on_date=None):
    """
    Subject line carrying the period and its dates, e.g.

        [DAC CRM] Weekly Report | Sun, 02-Aug-2026 to Fri, 07-Aug-2026

    The same subject is used for the team and for each member — the report is
    identified by its period, not by who is reading it.
    """
    return "[DAC CRM] {} Report | {}".format(
        (report_type or "daily").title(), report_period_label(report_type, on_date))


def _quotation_status_line(stats):
    """
    'Draft: 4 | Expired: 2 | Lost: 1' — only the statuses that actually occurred,
    so the caption stays short but never hides a status the way the fixed
    Draft/Open/Ordered caption did.
    """
    parts = []
    for st in QUOTATION_STATUSES:
        n = (stats or {}).get(_q_status_key(st), 0) or 0
        if n:
            parts.append("{}: {}".format(st, n))
    other = (stats or {}).get("q_st_other", 0) or 0
    if other:
        parts.append("Other: {}".format(other))
    return " | ".join(parts) if parts else "No quotations in this period"


@frappe.whitelist()
def get_crm_dashboard_metadata():
    """Return users, fiscal years, industries, sources, and current fiscal year for CRM Dashboard filters."""
    users = frappe.db.sql("""
        SELECT u.name, u.full_name
        FROM `tabUser` u
        JOIN `tabHas Role` hr ON hr.parent = u.name
        WHERE u.enabled = 1 AND u.user_type = 'System User'
          AND hr.role IN ('DAC CRM Head', 'DAC CRM')
        GROUP BY u.name
        ORDER BY u.full_name ASC
    """, as_dict=True)
    
    fiscal_years = frappe.get_all("Fiscal Year", filters={"disabled": 0}, fields=["name"], order_by="year_start_date desc")
    industries = frappe.get_all("Industry Type", fields=["name"], order_by="name asc")

    # Fetch distinct lead sources from Lead Source doctype
    sources = frappe.get_all("Lead Source", fields=["name"], order_by="name asc")
    # Also grab any sources used in leads but not yet in Lead Source
    db_sources = frappe.db.sql(
        "SELECT DISTINCT source FROM `tabLead` WHERE source IS NOT NULL AND source != '' ORDER BY source ASC",
        as_dict=1)
    source_names = {s.name for s in sources}
    for sr in db_sources:
        if sr.source and sr.source not in source_names:
            sources.append({"name": sr.source})
            source_names.add(sr.source)
    sources = sorted(sources, key=lambda x: x["name"])
    
    today_date = frappe.utils.today()
    current_fy = frappe.db.get_value("Fiscal Year", {"year_start_date": ["<=", today_date], "year_end_date": [">=", today_date], "disabled": 0}, "name")
    if not current_fy and fiscal_years:
        current_fy = fiscal_years[0].name

    all_users = frappe.db.get_all("User", fields=["name", "full_name"])
    user_map = {u.name: u.full_name for u in all_users if u.full_name}

    return {
        "users": users,
        "fiscal_years": fiscal_years,
        "industries": industries,
        "sources": sources,
        "current_fiscal_year": current_fy,
        "user_map": user_map
    }

@frappe.whitelist()
def get_target_vs_actual(from_date=None, to_date=None, fiscal_year=None, month=None, user=None):
    if fiscal_year and frappe.db.exists("Fiscal Year", fiscal_year):
        fy_doc = frappe.db.get_value("Fiscal Year", fiscal_year, ["year_start_date", "year_end_date"], as_dict=1)
        if fy_doc:
            from_date = str(fy_doc.year_start_date)
            to_date = str(fy_doc.year_end_date)
            if month:
                try:
                    import calendar
                    m_idx = int(month)
                    fy_start = getdate(fy_doc.year_start_date)
                    m_year = fy_start.year if m_idx >= fy_start.month else fy_start.year + 1
                    from_date = f"{m_year}-{m_idx:02d}-01"
                    last_day = calendar.monthrange(m_year, m_idx)[1]
                    to_date = f"{m_year}-{m_idx:02d}-{last_day}"
                except Exception:
                    pass

    if not from_date or not to_date:
        fy_doc = frappe.db.get_value("Fiscal Year", "2025-2026", ["year_start_date", "year_end_date"], as_dict=1)
        if not fy_doc:
            fy_doc = frappe.db.get_value("Fiscal Year", {"disabled": 0}, ["year_start_date", "year_end_date"], order_by="year_start_date desc", as_dict=1)
        if fy_doc:
            from_date = str(fy_doc.year_start_date)
            to_date = str(fy_doc.year_end_date)
            if month:
                try:
                    import calendar
                    m_idx = int(month)
                    fy_start = getdate(fy_doc.year_start_date)
                    m_year = fy_start.year if m_idx >= fy_start.month else fy_start.year + 1
                    from_date = f"{m_year}-{m_idx:02d}-01"
                    last_day = calendar.monthrange(m_year, m_idx)[1]
                    to_date = f"{m_year}-{m_idx:02d}-{last_day}"
                except Exception:
                    pass
        else:
            from_date = "2025-04-01"
            to_date = "2026-03-31"

    from_date_obj = getdate(from_date)
    to_date_obj = getdate(to_date)
    current_user = frappe.session.user
    
    # 1. Role-Based Filtering
    user_roles = frappe.get_roles(current_user)
    sp_filters = {"enabled": 1, "is_group": 0}
    
    if user:
        employee_id = frappe.db.get_value("Employee", {"user_id": user}, "name")
        if employee_id:
            sp_filters["employee"] = employee_id
        else:
            return {"year_start_date": str(from_date), "year_end_date": str(to_date), "data": []}
    elif "DAC CRM Head" not in user_roles and "Administrator" not in user_roles:
        employees = frappe.get_all("Employee", filters={"user_id": current_user}, fields=["name"])
        employee_names = [e.name for e in employees]
        if not employee_names:
            return {"year_start_date": str(from_date), "year_end_date": str(to_date), "data": []}
        sp_filters["employee"] = ["in", employee_names]

    sales_persons = frappe.get_all("Sales Person", fields=["name", "sales_person_name"], filters=sp_filters)
    if not sales_persons: 
        return {"year_start_date": str(from_date), "year_end_date": str(to_date), "data": []}

    data = []

    if user and not month and len(sales_persons) == 1:
        sp = sales_persons[0]
        months_to_show = get_months_in_range(from_date_obj, to_date_obj)
        sp_targets = frappe.db.get_all("Target Detail", filters={"parent": sp.name}, fields=["fiscal_year", "target_amount", "distribution_id"])
        
        active_fy = None
        for t in sp_targets:
            fy_doc = frappe.get_cached_doc("Fiscal Year", t.fiscal_year)
            if getdate(fy_doc.year_start_date) <= to_date_obj and getdate(fy_doc.year_end_date) >= from_date_obj:
                active_fy = fy_doc
                break
                
        fy_start_date = getdate(active_fy.year_start_date) if active_fy else from_date_obj
        
        running_bal = 0.0
        if from_date_obj > fy_start_date:
            hist_end = from_date_obj - datetime.timedelta(days=1)
            hist_months = get_months_in_range(fy_start_date, hist_end)
            for m_name in hist_months:
                m_idx = list(calendar_months().values()).index(m_name) + 1
                m_year = fy_start_date.year if m_idx >= fy_start_date.month else fy_start_date.year + 1
                m_start = datetime.date(m_year, m_idx, 1)
                m_t = get_specific_target(sp.name, [m_name], m_start, get_last_day(m_start))
                m_a = get_specific_actual(sp.name, m_start, get_last_day(m_start))
                running_bal += (m_a - m_t)
                if running_bal > 0: running_bal = 0.0
                
        for m_name in months_to_show:
            m_idx = list(calendar_months().values()).index(m_name) + 1
            m_year = fy_start_date.year if m_idx >= fy_start_date.month else fy_start_date.year + 1
            m_start = datetime.date(m_year, m_idx, 1)
            m_end = get_last_day(m_start)
            
            if m_start < from_date_obj: m_start = from_date_obj
            if m_end > to_date_obj: m_end = to_date_obj
            
            curr_target = get_specific_target(sp.name, [m_name], m_start, m_end)
            curr_actual = get_specific_actual(sp.name, m_start, m_end)
            
            ytd_months_list = get_months_in_range(fy_start_date, m_end)
            ytd_range_label = f"{fy_start_date.strftime('%b')} to {m_end.strftime('%b')}"
            ytd_target = get_specific_target(sp.name, ytd_months_list, fy_start_date, m_end)
            
            prev_deficit = running_bal
            curr_variance = curr_actual - curr_target
            curr_ach = (curr_actual / curr_target * 100) if curr_target > 0 else 0.0
            overall_variance = prev_deficit + curr_variance
            total_burden = curr_target + abs(prev_deficit)
            overall_ach = (curr_actual / total_burden * 100) if total_burden > 0 else 0.0
            
            running_bal += curr_variance
            if running_bal > 0: running_bal = 0.0
            
            data.append({
                "sales_person": f"{m_name} {m_year}",
                "sales_person_name": f"{m_name} {m_year}",
                "full_name": f"{m_name} {m_year}",
                "target": curr_target,
                "target_amount": curr_target,
                "monthly_target": curr_target,
                "annual_target": ytd_target,
                "actual": curr_actual,
                "actual_closed": curr_actual,
                "actual_revenue": curr_actual,
                "curr_variance": curr_variance,
                "variance_amount": curr_variance,
                "curr_achievement": curr_ach,
                "achievement_percent": curr_ach,
                "ytd_label": ytd_range_label,
                "ytd_target": ytd_target,
                "overall_variance": overall_variance,
                "overall_achievement": overall_ach
            })
            
    else:
        for sp in sales_persons:
            curr_actual = get_specific_actual(sp.name, from_date_obj, to_date_obj)
            sp_targets = frappe.db.get_all("Target Detail", filters={"parent": sp.name}, fields=["fiscal_year", "target_amount", "distribution_id"])
            
            active_fy = None
            curr_target = 0.0
            ytd_target = 0.0
            prev_deficit = 0.0
            ytd_range_label = ""
    
            for t in sp_targets:
                fy_doc = frappe.get_cached_doc("Fiscal Year", t.fiscal_year)
                if getdate(fy_doc.year_start_date) <= to_date_obj and getdate(fy_doc.year_end_date) >= from_date_obj:
                    active_fy = fy_doc
                    break
            
            if active_fy:
                fy_start_date = getdate(active_fy.year_start_date)
    
                if from_date_obj > fy_start_date:
                    hist_end = from_date_obj - datetime.timedelta(days=1)
                    hist_months = get_months_in_range(fy_start_date, hist_end)
                    running_bal = 0.0
                    for m_name in hist_months:
                        m_idx = list(calendar_months().values()).index(m_name) + 1
                        m_year = fy_start_date.year if m_idx >= fy_start_date.month else fy_start_date.year + 1
                        m_start = datetime.date(m_year, m_idx, 1)
                        m_t = get_specific_target(sp.name, [m_name], m_start, get_last_day(m_start))
                        m_a = get_specific_actual(sp.name, m_start, get_last_day(m_start))
                        running_bal += (m_a - m_t)
                        if running_bal > 0: running_bal = 0.0
                    prev_deficit = running_bal
    
                curr_months = get_months_in_range(from_date_obj, to_date_obj)
                curr_target = get_specific_target(sp.name, curr_months, from_date_obj, to_date_obj)
                ytd_months_list = get_months_in_range(fy_start_date, to_date_obj)
                ytd_range_label = f"{fy_start_date.strftime('%b')} to {to_date_obj.strftime('%b')}"
                ytd_target = get_specific_target(sp.name, ytd_months_list, fy_start_date, to_date_obj)
    
            curr_variance = curr_actual - curr_target
            curr_ach = (curr_actual / curr_target * 100) if curr_target > 0 else 0.0
            overall_variance = prev_deficit + curr_variance
            total_burden = curr_target + abs(prev_deficit)
            overall_ach = (curr_actual / total_burden * 100) if total_burden > 0 else 0.0
    
            data.append({
                "sales_person": sp.sales_person_name,
                "sales_person_name": sp.sales_person_name,
                "full_name": sp.sales_person_name,
                "target": curr_target,
                "target_amount": curr_target,
                "monthly_target": curr_target,
                "annual_target": ytd_target,
                "actual": curr_actual,
                "actual_closed": curr_actual,
                "actual_revenue": curr_actual,
                "curr_variance": curr_variance,
                "variance_amount": curr_variance,
                "curr_achievement": curr_ach,
                "achievement_percent": curr_ach,
                "ytd_label": ytd_range_label,
                "ytd_target": ytd_target,
                "overall_variance": overall_variance,
                "overall_achievement": overall_ach
            })

    return {
        "year_start_date": str(from_date),
        "year_end_date": str(to_date),
        "data": data
    }


@frappe.whitelist()
def get_monthly_target_vs_actual(fiscal_year=None, user=None):
    if not fiscal_year:
        fiscal_year = "2025-2026"
        
    fy_doc = frappe.db.get_value("Fiscal Year", fiscal_year, ["year_start_date", "year_end_date"], as_dict=1)
    if not fy_doc:
        fy_doc = frappe.db.get_value("Fiscal Year", {"disabled": 0}, ["year_start_date", "year_end_date"], order_by="year_start_date desc", as_dict=1)
        
    if not fy_doc:
        return []
        
    from_date = fy_doc.year_start_date
    to_date = fy_doc.year_end_date
    
    current_user = frappe.session.user
    user_roles = frappe.get_roles(current_user)
    sp_filters = {"enabled": 1, "is_group": 0}
    
    if user:
        employee_id = frappe.db.get_value("Employee", {"user_id": user}, "name")
        if employee_id:
            sp_filters["employee"] = employee_id
        else:
            return []
    elif "DAC CRM Head" not in user_roles and "Administrator" not in user_roles:
        employees = frappe.get_all("Employee", filters={"user_id": current_user}, fields=["name"])
        employee_names = [e.name for e in employees]
        if not employee_names:
            return []
        sp_filters["employee"] = ["in", employee_names]
        
    sales_persons = frappe.get_all("Sales Person", fields=["name", "sales_person_name"], filters=sp_filters)
    if not sales_persons:
        return []
        
    from_date_obj = getdate(from_date)
    to_date_obj = getdate(to_date)
    months_to_show = get_months_in_range(from_date_obj, to_date_obj)
    
    data = []
    running_target = 0.0
    running_actual = 0.0
    
    for m_name in months_to_show:
        m_idx = list(calendar_months().values()).index(m_name) + 1
        m_year = from_date_obj.year if m_idx >= from_date_obj.month else from_date_obj.year + 1
        m_start = datetime.date(m_year, m_idx, 1)
        m_end = get_last_day(m_start)
        
        month_target = 0.0
        month_actual = 0.0
        
        for sp in sales_persons:
            month_target += get_specific_target(sp.name, [m_name], m_start, m_end)
            month_actual += get_specific_actual(sp.name, m_start, m_end)
            
        month_variance = month_actual - month_target
        month_ach = (month_actual / month_target * 100) if month_target > 0 else 0.0
        
        running_target += month_target
        running_actual += month_actual
        cum_variance = running_actual - running_target
        cum_ach = (running_actual / running_target * 100) if running_target > 0 else 0.0
        
        data.append({
            "month": f"{m_name} {m_year}",
            "target": month_target,
            "actual": month_actual,
            "variance": month_variance,
            "ach_pct": month_ach,
            "cum_target": running_target,
            "cum_actual": running_actual,
            "cum_variance": cum_variance,
            "cum_ach_pct": cum_ach
        })
        
    return data


@frappe.whitelist()
def get_target_vs_actual_data(fiscal_year=None, from_date=None, to_date=None):
    """Whitelisted alias for Target vs Actual analysis endpoint."""
    return get_target_vs_actual(from_date=from_date, to_date=to_date, fiscal_year=fiscal_year)


def get_specific_target(sp_name, month_names, f_date, t_date):
    target_val = 0.0
    sp_targets = frappe.db.get_all("Target Detail", 
                                   filters={"parent": sp_name}, 
                                   fields=["fiscal_year", "target_amount", "distribution_id"])
    for t in sp_targets:
        fy_doc = frappe.get_cached_doc("Fiscal Year", t.fiscal_year)
        if getdate(fy_doc.year_start_date) <= t_date and getdate(fy_doc.year_end_date) >= f_date:
            if t.distribution_id:
                dist = frappe.get_cached_doc("Monthly Distribution", t.distribution_id)
                pct_sum = sum([flt(getattr(d, 'percentage_allocation', 0) or getattr(d, 'percentage', 0)) 
                              for d in dist.percentages if d.month in month_names])
                target_val += (flt(t.target_amount) * pct_sum) / 100.0
            else:
                target_val += (flt(t.target_amount) / 12.0) * len(month_names)
    return target_val

def get_specific_actual(sp_name, f_date, t_date):
    actual_query = """
        SELECT SUM(st.allocated_amount)
        FROM `tabSales Order` so
        INNER JOIN `tabSales Team` st ON st.parent = so.name
        WHERE so.docstatus = 1 AND st.sales_person = %s
        AND so.transaction_date BETWEEN %s AND %s
    """
    res = frappe.db.sql(actual_query, (sp_name, f_date, t_date))
    return flt(res[0][0]) if res else 0.0

def calendar_months():
    return {1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June", 
            7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December"}

def get_months_in_range(start_date, end_date):
    month_map = calendar_months()
    months = []
    current = start_date
    while current <= end_date:
        if month_map[current.month] not in months:
            months.append(month_map[current.month])
        if current.month == 12:
            current = datetime.date(current.year + 1, 1, 1)
        else:
            current = datetime.date(current.year, current.month + 1, 1)
    return months



def validate_lead(doc, method=None):

    # 1. Update Direction Type (Logic you asked for earlier)
    outbound_sources = ["Field Visit", "Cold Calling"]
    doc.custom_direction_type = "Outbound" if doc.source in outbound_sources else "Inbound"

    # 2. Track Changes (History Tracking)
    # We do this in validate because it's safe to append to child tables here
    if doc.is_new():
        # Duplicating a lead (the reopen flow, or the standard Duplicate button) goes
        # through frappe.copy_doc, which brings the source lead's history across with
        # its original timestamps. A brand new lead has no history of its own, so
        # anything sitting here is inherited and would make the dashboards count the
        # source lead's Order conversion a second time.
        doc.set('custom_lead_status_change_history', [])
        doc.set('custom_lead_owner_change_history', [])
    else:
        # Track Category Change
        old_status = frappe.db.get_value("Lead", doc.name, "custom_lead_category")
        if old_status and old_status != doc.custom_lead_category:
            doc.append('custom_lead_status_change_history', {
                'old_status': old_status,
                'new_status': doc.custom_lead_category,
                'changed_by': frappe.session.user,
                'updated_at': now_datetime(),
            })

        # Track Owner Change
        old_owner = frappe.db.get_value("Lead", doc.name, "lead_owner")
        if old_owner and old_owner != doc.lead_owner:
            doc.append('custom_lead_owner_change_history', {
                'from_lead_owner': old_owner,
                'to_lead_owner': doc.lead_owner,
                'changed_by': frappe.session.user,
                'change_on': now_datetime(),
            })








# import frappe
# from frappe.utils import flt

# @frappe.whitelist()
# def get_tabular_dashboard_data(from_date=None, to_date=None, user=None, industry=None):
#     lead_rows = []
    
#     # 1. Base Query Params
#     params = {}
#     filter_clause = " WHERE docstatus < 2 "
    
#     if user:
#         filter_clause += " AND lead_owner = %(user)s "
#         params["user"] = user
#     if industry:
#         # User defined label Industry Type, usually matches field name custom_industry_type
#         # Verify fieldname in Lead doctype. Replace 'custom_industry_type' if needed.
#         filter_clause += " AND industry = %(industry)s "
#         params["industry"] = industry
#     if from_date and to_date:
#         filter_clause += " AND creation BETWEEN %(sd)s AND %(ed)s "
#         params["sd"] = from_date
#         params["ed"] = to_date

#     # --- Lead Data Extraction ---
#     # Using specific Logic: Order -> PO Value | All others -> Expected Revenue
#     leads_sql = frappe.db.sql(f"""
#         SELECT 
#             MONTHNAME(creation) as m_name, YEAR(creation) as m_year, MONTH(creation) as m_num,
#             custom_lead_category as category,
#             COUNT(name) as count,
#             SUM(CASE 
#                 WHEN custom_lead_category = 'Order' THEN IFNULL(custom_po_value, 0) 
#                 ELSE IFNULL(custom_expected_revenue, 0) 
#             END) as value
#         FROM `tabLead` 
#         {filter_clause}
#         GROUP BY m_year, m_num, category
#         ORDER BY m_year ASC, m_num ASC
#     """, params, as_dict=1)

#     lead_map = {}
#     # Grand Totals accumulator
#     lt = {"enq_c":0,"enq_v":0,"pipe_c":0,"pipe_v":0,"ord_c":0,"ord_v":0,"lenq_c":0,"lenq_v":0,"lpipe_c":0,"lpipe_v":0}

#     for l in leads_sql:
#         m_key = f"{l.m_name} {l.m_year}"
#         if m_key not in lead_map:
#             lead_map[m_key] = {"month_label": m_key, "sort": f"{l.m_year}-{l.m_num:02}", "enq_cnt":0, "enq_val":0, "pipe_cnt":0, "pipe_val":0, "ord_cnt":0, "ord_val":0, "lenq_cnt":0, "lenq_val":0, "lpipe_cnt":0, "lpipe_val":0}
        
#         row = lead_map[m_key]
#         c, v = int(l.count), flt(l.value)
        
#         if l.category == 'Enquiry': row['enq_cnt'], row['enq_val'] = c, v; lt["enq_c"] += c; lt["enq_v"] += v
#         elif l.category == 'Pipeline': row['pipe_cnt'], row['pipe_val'] = c, v; lt["pipe_c"] += c; lt["pipe_v"] += v
#         elif l.category == 'Order': row['ord_cnt'], row['ord_val'] = c, v; lt["ord_c"] += c; lt["ord_v"] += v
#         elif l.category == 'Lost Enquiry': row['lenq_cnt'], row['lenq_val'] = c, v; lt["lenq_c"] += c; lt["lenq_v"] += v
#         elif l.category == 'Lost Pipeline': row['lpipe_cnt'], row['lpipe_val'] = c, v; lt["lpipe_c"] += c; lt["lpipe_v"] += v

#     lead_rows = sorted(lead_map.values(), key=lambda x: x['sort'])

#     # --- Contact Management Logic ---
#     bc_filters = " WHERE docstatus < 2 "
#     bc_params = {}
#     if user: bc_filters += " AND assign_to = %(u)s "; bc_params["u"] = user
#     if from_date and to_date: bc_filters += " AND creation BETWEEN %(s)s AND %(e)s "; bc_params["s"] = from_date; bc_params["e"] = to_date

#     contacts = frappe.db.sql(f"""
#         SELECT MONTHNAME(creation) as m_name, YEAR(creation) as m_year, status, count(name) as count
#         FROM `tabBusiness Contacts` {bc_filters}
#         GROUP BY m_year, m_name, status
#     """, bc_params, as_dict=1)

#     bc_map = {}
#     ct = {"o": 0, "c": 0, "e": 0} # Totals: Open, Converted, Existing
    
#     for c in contacts:
#         label = f"{c.m_name} {c.m_year}"
#         if label not in bc_map: bc_map[label] = {"label": label, "open": 0, "conv": 0, "existing": 0}
        
#         count = int(c.count)
#         if c.status == "Open": bc_map[label]["open"] = count; ct["o"] += count
#         elif c.status == "Converted to Lead": bc_map[label]["conv"] = count; ct["c"] += count
#         elif c.status == "Existing Customer": bc_map[label]["existing"] = count; ct["e"] += count

#     return {
#         "lead_data": lead_rows,
#         "lead_totals": lt,
#         "contact_data": list(bc_map.values()),
#         "contact_totals": ct
#     }

# import frappe
# from frappe.utils import flt, format_date

# @frappe.whitelist()
# def get_tabular_dashboard_data(from_date=None, to_date=None, user=None, industry=None):
#     # Parameters initialization
#     params = {"sd": from_date, "ed": to_date, "usr": user, "ind": industry}
    
#     # 1. Base Filter String logic
#     filters = " WHERE docstatus < 2 "
#     bc_filters = " WHERE docstatus < 2 "
    
#     if user:
#         filters += " AND lead_owner = %(usr)s "
#         bc_filters += " AND assign_to = %(usr)s "
#     if industry:
#         # Assuming field name is industry in both doctypes. Change if different.
#         filters += " AND industry = %(ind)s "
#         bc_filters += " AND industry = %(ind)s "
#     if from_date and to_date:
#         filters += " AND creation BETWEEN %(sd)s AND %(ed)s "
#         bc_filters += " AND creation BETWEEN %(sd)s AND %(ed)s "

#     # --- LEAD LOGIC ---
#     leads = frappe.db.sql(f"""
#         SELECT 
#             MONTHNAME(creation) as m_name, YEAR(creation) as m_year, MONTH(creation) as m_num,
#             custom_lead_category as cat, COUNT(name) as count,
#             SUM(CASE WHEN custom_lead_category = 'Order' THEN IFNULL(custom_po_value, 0) ELSE IFNULL(custom_expected_revenue, 0) END) as val
#         FROM `tabLead` {filters}
#         GROUP BY m_year, m_num, m_name, cat ORDER BY m_year ASC, m_num ASC
#     """, params, as_dict=1)

#     lead_map = {}
#     lt = {"enq_c":0,"enq_v":0,"pipe_c":0,"pipe_v":0,"ord_c":0,"ord_v":0,"lenq_c":0,"lenq_v":0,"lpipe_c":0,"lpipe_v":0}

#     for l in leads:
#         key = f"{l.m_name} {l.m_year}"
#         if key not in lead_map:
#             lead_map[key] = {"label": key, "sort": f"{l.m_year}-{l.m_num:02}", "enq_c":0,"enq_v":0,"pipe_c":0,"pipe_v":0,"ord_c":0,"ord_v":0,"lenq_c":0,"lenq_v":0,"lpipe_c":0,"lpipe_v":0}
        
#         row, c, v = lead_map[key], int(l.count), flt(l.val)
#         if l.cat == 'Enquiry': row['enq_c'], row['enq_v'] = c, v; lt['enq_c']+=c; lt['enq_v']+=v
#         elif l.cat == 'Pipeline': row['pipe_c'], row['pipe_v'] = c, v; lt['pipe_c']+=c; lt['pipe_v']+=v
#         elif l.cat == 'Order': row['ord_c'], row['ord_v'] = c, v; lt['ord_c']+=c; lt['ord_v']+=v
#         elif l.cat == 'Lost Enquiry': row['lenq_c'], row['lenq_v'] = c, v; lt['lenq_c']+=c; lt['lenq_v']+=v
#         elif l.cat == 'Lost Pipeline': row['lpipe_c'], row['lpipe_v'] = c, v; lt['lpipe_c']+=c; lt['lpipe_v']+=v

#     # --- CONTACT LOGIC ---
#     contacts = frappe.db.sql(f"""
#         SELECT MONTHNAME(creation) as m_name, YEAR(creation) as m_year, MONTH(creation) as m_num, status, COUNT(name) as count
#         FROM `tabBusiness Contacts` {bc_filters}
#         GROUP BY m_year, m_num, m_name, status ORDER BY m_year ASC, m_num ASC
#     """, params, as_dict=1)

#     bc_map = {}
#     ct = {"o": 0, "c": 0, "e": 0}
#     for b in contacts:
#         key = f"{b.m_name} {b.m_year}"
#         if key not in bc_map:
#             bc_map[key] = {"label": key, "o": 0, "c": 0, "e": 0, "sort": f"{b.m_year}-{b.m_num:02}"}
        
#         cnt = int(b.count)
#         if b.status == "Open": bc_map[key]["o"] = cnt; ct["o"] += cnt
#         elif b.status == "Converted to Lead": bc_map[key]["c"] = cnt; ct["c"] += cnt
#         elif b.status == "Existing Customer": bc_map[key]["e"] = cnt; ct["e"] += cnt

#     return {
#         "leads": sorted(lead_map.values(), key=lambda x: x['sort']),
#         "lead_totals": lt,
#         "contacts": sorted(bc_map.values(), key=lambda x: x['sort']),
#         "contact_totals": ct
#     }


import frappe
import datetime
from frappe.utils import flt, getdate, today, get_last_day

def parse_period_and_dates(period=None, fiscal_year=None, from_date=None, to_date=None, month=None):
    """Resolve period, fiscal_year, month, and custom dates to explicit from_date and to_date."""
    if from_date and to_date:
        return str(from_date), str(to_date)

    if period and str(period).startswith("fy:"):
        fiscal_year = str(period)[3:]
        period = None

    today_date = getdate(today())

    if month:
        try:
            m_str = str(month).strip()
            # Case 1: YYYY-MM or YYYY-M
            if '-' in m_str and len(m_str) >= 6:
                parts = m_str.split('-')
                yr = int(parts[0])
                m = int(parts[1])
                start_d = datetime.date(yr, m, 1)
                last_d = get_last_day(start_d)
                return start_d.strftime("%Y-%m-%d"), last_d.strftime("%Y-%m-%d")
            # Case 2: "Month Year" like "September 2025" or "Sep 2025"
            import re
            m_match = re.search(r'([A-Za-z]+)\s*(\d{4})', m_str)
            if m_match:
                m_name = m_match.group(1).lower()
                yr = int(m_match.group(2))
                months_map = {
                    'january':1, 'jan':1, 'february':2, 'feb':2, 'march':3, 'mar':3,
                    'april':4, 'apr':4, 'may':5, 'june':6, 'jun':6, 'july':7, 'jul':7,
                    'august':8, 'aug':8, 'september':9, 'sep':9, 'sept':9, 'october':10, 'oct':10,
                    'november':11, 'nov':11, 'december':12, 'dec':12
                }
                m = months_map.get(m_name, 1)
                start_d = datetime.date(yr, m, 1)
                last_d = get_last_day(start_d)
                return start_d.strftime("%Y-%m-%d"), last_d.strftime("%Y-%m-%d")
            # Case 3: integer month "9" or "09"
            if m_str.isdigit():
                m = int(m_str)
                fy_start_year = int(str(fiscal_year).split('-')[0]) if (fiscal_year and '-' in str(fiscal_year)) else today_date.year
                year = (fy_start_year + 1) if m in (1, 2, 3) else fy_start_year
                start_d = datetime.date(year, m, 1)
                last_d = get_last_day(start_d)
                return start_d.strftime("%Y-%m-%d"), last_d.strftime("%Y-%m-%d")
        except Exception:
            pass

    if not fiscal_year:
        if period == "this_year":
            fy_s = (today_date.year - 1) if today_date.month in (1, 2, 3) else today_date.year
            fiscal_year = f"{fy_s}-{fy_s + 1}"
        elif period == "last_year":
            fy_s = (today_date.year - 2) if today_date.month in (1, 2, 3) else (today_date.year - 1)
            fiscal_year = f"{fy_s}-{fy_s + 1}"

    if period:
        if period == "all":
            return None, None
        today_date = getdate(today())
        if period == "today":
            t_str = today_date.strftime("%Y-%m-%d")
            return t_str, t_str
        elif period == "this_week":
            start_w = today_date - datetime.timedelta(days=today_date.weekday())
            end_w = start_w + datetime.timedelta(days=6)
            return start_w.strftime("%Y-%m-%d"), end_w.strftime("%Y-%m-%d")
        elif period == "this_month":
            return today_date.replace(day=1).strftime("%Y-%m-%d"), get_last_day(today_date).strftime("%Y-%m-%d")
        elif period == "last_month":
            first_this = today_date.replace(day=1)
            last_prev = first_this - datetime.timedelta(days=1)
            return last_prev.replace(day=1).strftime("%Y-%m-%d"), last_prev.strftime("%Y-%m-%d")
        elif period == "this_quarter":
            m_curr = today_date.month
            q_start_month = 3 * ((m_curr - 1) // 3) + 1
            start_d = today_date.replace(month=q_start_month, day=1)
            q_end_month = q_start_month + 2
            last_d = get_last_day(today_date.replace(month=q_end_month, day=1))
            return start_d.strftime("%Y-%m-%d"), last_d.strftime("%Y-%m-%d")
        elif period == "last_quarter":
            m_curr = today_date.month
            q_start_month = 3 * ((m_curr - 1) // 3) + 1
            first_this_q = today_date.replace(month=q_start_month, day=1)
            last_prev_q = first_this_q - datetime.timedelta(days=1)
            prev_m_curr = last_prev_q.month
            prev_q_start_month = 3 * ((prev_m_curr - 1) // 3) + 1
            start_d = last_prev_q.replace(month=prev_q_start_month, day=1)
            return start_d.strftime("%Y-%m-%d"), last_prev_q.strftime("%Y-%m-%d")
        elif period == "this_year":
            fy_start_year = (today_date.year - 1) if today_date.month in (1, 2, 3) else today_date.year
            return f"{fy_start_year}-04-01", f"{fy_start_year + 1}-03-31"
        elif period == "last_year":
            fy_start_year = (today_date.year - 2) if today_date.month in (1, 2, 3) else (today_date.year - 1)
            return f"{fy_start_year}-04-01", f"{fy_start_year + 1}-03-31"

    if fiscal_year and frappe.db.exists("Fiscal Year", fiscal_year):
        fy_doc = frappe.db.get_value("Fiscal Year", fiscal_year, ["year_start_date", "year_end_date"], as_dict=1)
        if fy_doc:
            return str(fy_doc.year_start_date), str(fy_doc.year_end_date)

    return from_date, to_date


# When a Business Contact moved to 'Converted to Lead'. Used by BOTH the
# Contact -> Lead count and its drilldown; if the two ever use different
# expressions, a period can show a count whose drilldown comes back empty.
# Only records currently in 'Converted to Lead' are measured, so no
# 'Existing Customer' history is read here.
BC_CONVERSION_DATE_SQL = """
        DATE(COALESCE(
            (SELECT MIN(sh.updated_at) FROM `tabLead Status Change History` sh
             WHERE sh.parent = `tabBusiness Contacts`.name
               AND sh.parenttype = 'Business Contacts'
               AND sh.parentfield = 'contact_status_change_history'
               AND sh.new_status = 'Converted to Lead'),
            modified,
            creation
        ))
    """


# Contact -> Lead counts only contacts whose status is 'Converted to Lead'.
# 'Existing Customer' is deliberately out of scope for conversion: those records
# are not evaluated at all, and their history is not consulted. Conversion
# therefore concerns only the Open -> Converted to Lead move.
BC_CONVERTED_COND_SQL = """
        (`tabBusiness Contacts`.status = 'Converted to Lead')
    """


# Reopening a won lead means duplicating it (create_duplicate_lead, or the standard
# Duplicate button), and frappe.copy_doc carries the whole status-change child table
# across with the ORIGINAL's timestamps. The copy therefore inherits its parent's
# 'Order' rows and gets counted as a second conversion of the same real order, dated
# months before the copy existed. validate_lead now drops that inherited history on
# insert, but leads duplicated before this guard still carry it, so every read of the
# child table also ignores rows older than the lead itself. Nothing genuine is lost:
# validate_lead only ever appends to an already-saved lead, so a real row can never
# predate creation.
LEAD_OWN_HISTORY_SQL = " AND sh.updated_at >= l.creation "


# Which leads count as converted: the ones standing at Order now. A lead that won and
# was then reopened or lost is back in Pipeline / Lost and is no longer an order, so it
# drops out again. This keeps Lead -> Order equal to the Order column at all times.
LEAD_ORDER_COND_SQL = " l.custom_lead_category = 'Order' "


# When a Lead most recently transitioned to Order.
# A lead can cycle (Order -> Lost -> reopened -> Order again), so the child table
# holds several 'Order' rows for it; DAC-LEAD-0730 has five spanning three months.
# Taking MAX collapses those to one conversion event per lead, so no lead is counted
# twice, and it picks the move that reflects where the lead stands now.
# The COALESCE covers Order leads carrying no transition row of their own — four exist
# (DAC-LEAD-1013, 1129, 1144, 1147), duplicated leads whose move to Order was never
# written to the child table. Without it they would be orders that no conversion figure
# could see. This matches lead_action_date exactly, so a lead lands in the same period
# in the Order column and in Lead -> Order.
LEAD_ORDER_CONV_DATE_SQL = f"""
        COALESCE(
            (SELECT DATE(MAX(sh.updated_at)) FROM `tabLead Status Change History` sh
             WHERE sh.parent = l.name AND sh.parenttype = 'Lead'
               AND sh.parentfield = 'custom_lead_status_change_history'
               AND sh.new_status = 'Order'
               {LEAD_OWN_HISTORY_SQL}),
            DATE(l.creation)
        )
    """


def period_sql_for(date_expr, period_type):
    """(label, sort, group) SQL for bucketing date_expr by the chosen interval."""
    if period_type == "daily":
        return (f"DATE_FORMAT({date_expr}, '%%d-%%b-%%Y (%%a)')",
                f"DATE_FORMAT({date_expr}, '%%Y-%%m-%%d')",
                f"DATE_FORMAT({date_expr}, '%%Y-%%m-%%d')")
    if period_type == "weekly":
        return (f"CONCAT('Week ', WEEK({date_expr}, 1), ' (', "
                f"DATE_FORMAT(MIN(DATE_SUB(DATE({date_expr}), INTERVAL WEEKDAY({date_expr}) DAY)), '%%d-%%b-%%Y (%%a)'), "
                f"' to ', DATE_FORMAT(MIN(DATE_ADD(DATE_SUB(DATE({date_expr}), INTERVAL WEEKDAY({date_expr}) DAY), "
                f"INTERVAL 6 DAY)), '%%d-%%b-%%Y (%%a)'), ')')",
                f"CONCAT(YEAR({date_expr}), '-', LPAD(WEEK({date_expr}, 1), 2, '0'))",
                f"YEAR({date_expr}), WEEK({date_expr}, 1)")
    return (f"CONCAT(MONTHNAME({date_expr}), ' ', YEAR({date_expr}))",
            f"CONCAT(CASE WHEN MONTH({date_expr}) >= 4 THEN YEAR({date_expr}) ELSE YEAR({date_expr}) - 1 END, "
            f"'-', LPAD(CASE WHEN MONTH({date_expr}) >= 4 THEN MONTH({date_expr}) - 3 "
            f"ELSE MONTH({date_expr}) + 9 END, 2, '0'))",
            f"YEAR({date_expr}), MONTH({date_expr})")


def as_filter_list(value):
    """Normalise a multiselect dashboard filter into a list of values.

    Accepts a list, a JSON array string (frappe.call serialises arrays that way),
    or a comma-separated string. Returns [] when nothing is selected (= no filter).
    """
    if not value:
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(v).strip() for v in value if str(v).strip()]
    text = str(value).strip()
    if text.startswith("["):
        try:
            parsed = json.loads(text)
            if isinstance(parsed, list):
                return [str(v).strip() for v in parsed if str(v).strip()]
        except Exception:
            pass
    return [v.strip() for v in text.split(",") if v.strip()]


def build_in_clause(values, prefix, column, params):
    """Return ' AND <column> IN (%(prefix_0)s, ...) ' and load params. '' if no values."""
    if not values:
        return ""
    placeholders = ",".join(["%({}_{})s".format(prefix, i) for i in range(len(values))])
    for i, v in enumerate(values):
        params["{}_{}".format(prefix, i)] = v
    return " AND {} IN ({}) ".format(column, placeholders)


@frappe.whitelist()
def get_tabular_dashboard_data(from_date=None, to_date=None, user=None, industry=None, source=None, period_type="monthly", fiscal_year=None, period=None, month=None, group_by="period", **kwargs):
    from_date, to_date = parse_period_and_dates(period, fiscal_year, from_date, to_date, month=month or kwargs.get("month"))

    # Determine the User Filter based on permissions
    is_head = "DAC CRM Head" in frappe.get_roles()
    effective_user = user if (is_head and user) else (None if (is_head and not user) else frappe.session.user)

    # Multiselect filters (list, JSON array string, or comma-separated string)
    industry_list = as_filter_list(industry)
    source_list = as_filter_list(source)

    params = {"sd": from_date, "ed": to_date, "usr": effective_user}
    
    where_lead = " WHERE l.docstatus < 2 "
    where_cont = " WHERE docstatus < 2 "

    if effective_user:
        where_lead += " AND l.lead_owner = %(usr)s "
        where_cont += " AND assign_to = %(usr)s "
    where_lead += build_in_clause(industry_list, "ind", "l.industry", params)
    where_cont += build_in_clause(industry_list, "ind", "industry", params)
    where_lead += build_in_clause(source_list, "src", "l.source", params)
    where_cont += build_in_clause(source_list, "src", "source", params)

    lead_action_date = f"""
        CASE
            WHEN l.custom_lead_category = 'Enquiry' THEN DATE(l.creation)
            ELSE COALESCE(
                (SELECT DATE(MAX(sh.updated_at)) FROM `tabLead Status Change History` sh
                 WHERE sh.parent = l.name AND sh.parenttype = 'Lead'
                   AND sh.parentfield = 'custom_lead_status_change_history'
                   AND sh.new_status = l.custom_lead_category
                   {LEAD_OWN_HISTORY_SQL}),
                DATE(l.creation)
            )
        END
    """

    contact_action_date = """
        CASE
            WHEN status = 'Open' THEN DATE(creation)
            WHEN status = 'Converted to Lead' THEN DATE(COALESCE(
                (SELECT MAX(sh.updated_at) FROM `tabLead Status Change History` sh
                 WHERE sh.parent = `tabBusiness Contacts`.name AND sh.parenttype = 'Business Contacts' AND sh.parentfield = 'contact_status_change_history' AND sh.new_status = 'Converted to Lead'),
                modified,
                creation
            ))
            WHEN status = 'Existing Customer' THEN DATE(COALESCE(
                (SELECT MAX(sh.updated_at) FROM `tabLead Status Change History` sh
                 WHERE sh.parent = `tabBusiness Contacts`.name AND sh.parenttype = 'Business Contacts' AND sh.parentfield = 'contact_status_change_history' AND sh.new_status = 'Existing Customer'),
                modified,
                creation
            ))
            ELSE DATE(modified)
        END
    """

    if from_date and to_date:
        where_lead_creation = f" AND ({lead_action_date}) BETWEEN %(sd)s AND %(ed)s "
        where_cont += f" AND ({contact_action_date}) BETWEEN %(sd)s AND %(ed)s "
    else:
        where_lead_creation = ""

    # Grouping expressions
    lead_date_expr = f"({lead_action_date})"
    contact_date_expr = f"({contact_action_date})"
    if period_type == "daily":
        p_label_hist = "DATE_FORMAT(sh.updated_at, '%%d-%%b-%%Y (%%a)')"
        p_sort_hist = "DATE_FORMAT(sh.updated_at, '%%Y-%%m-%%d')"
        p_group_hist = "DATE_FORMAT(sh.updated_at, '%%Y-%%m-%%d')"

        p_label_lead = f"DATE_FORMAT({lead_date_expr}, '%%d-%%b-%%Y (%%a)')"
        p_sort_lead = f"DATE_FORMAT({lead_date_expr}, '%%Y-%%m-%%d')"
        p_group_lead = f"DATE_FORMAT({lead_date_expr}, '%%Y-%%m-%%d')"

        p_label_cont = f"DATE_FORMAT({contact_date_expr}, '%%d-%%b-%%Y (%%a)')"
        p_sort_cont = f"DATE_FORMAT({contact_date_expr}, '%%Y-%%m-%%d')"
        p_group_cont = f"DATE_FORMAT({contact_date_expr}, '%%Y-%%m-%%d')"
    elif period_type == "weekly":
        p_label_hist = "CONCAT('Week ', WEEK(sh.updated_at, 1), ' (', DATE_FORMAT(MIN(DATE_SUB(DATE(sh.updated_at), INTERVAL WEEKDAY(sh.updated_at) DAY)), '%%d-%%b-%%Y (%%a)'), ' to ', DATE_FORMAT(MIN(DATE_ADD(DATE_SUB(DATE(sh.updated_at), INTERVAL WEEKDAY(sh.updated_at) DAY), INTERVAL 6 DAY)), '%%d-%%b-%%Y (%%a)'), ')') "
        p_sort_hist = "CONCAT(YEAR(sh.updated_at), '-', LPAD(WEEK(sh.updated_at, 1), 2, '0'))"
        p_group_hist = "YEAR(sh.updated_at), WEEK(sh.updated_at, 1)"

        p_label_lead = f"CONCAT('Week ', WEEK({lead_date_expr}, 1), ' (', DATE_FORMAT(MIN(DATE_SUB(DATE({lead_date_expr}), INTERVAL WEEKDAY({lead_date_expr}) DAY)), '%%d-%%b-%%Y (%%a)'), ' to ', DATE_FORMAT(MIN(DATE_ADD(DATE_SUB(DATE({lead_date_expr}), INTERVAL WEEKDAY({lead_date_expr}) DAY), INTERVAL 6 DAY)), '%%d-%%b-%%Y (%%a)'), ')') "
        p_sort_lead = f"CONCAT(YEAR({lead_date_expr}), '-', LPAD(WEEK({lead_date_expr}, 1), 2, '0'))"
        p_group_lead = f"YEAR({lead_date_expr}), WEEK({lead_date_expr}, 1)"

        p_label_cont = f"CONCAT('Week ', WEEK({contact_date_expr}, 1), ' (', DATE_FORMAT(MIN(DATE_SUB(DATE({contact_date_expr}), INTERVAL WEEKDAY({contact_date_expr}) DAY)), '%%d-%%b-%%Y (%%a)'), ' to ', DATE_FORMAT(MIN(DATE_ADD(DATE_SUB(DATE({contact_date_expr}), INTERVAL WEEKDAY({contact_date_expr}) DAY), INTERVAL 6 DAY)), '%%d-%%b-%%Y (%%a)'), ')') "
        p_sort_cont = f"CONCAT(YEAR({contact_date_expr}), '-', LPAD(WEEK({contact_date_expr}, 1), 2, '0'))"
        p_group_cont = f"YEAR({contact_date_expr}), WEEK({contact_date_expr}, 1)"
    else: # monthly
        p_label_hist = "CONCAT(MONTHNAME(sh.updated_at), ' ', YEAR(sh.updated_at))"
        p_sort_hist = "CONCAT(CASE WHEN MONTH(sh.updated_at) >= 4 THEN YEAR(sh.updated_at) ELSE YEAR(sh.updated_at) - 1 END, '-', LPAD(CASE WHEN MONTH(sh.updated_at) >= 4 THEN MONTH(sh.updated_at) - 3 ELSE MONTH(sh.updated_at) + 9 END, 2, '0'))"
        p_group_hist = "YEAR(sh.updated_at), MONTH(sh.updated_at)"

        p_label_lead = f"CONCAT(MONTHNAME({lead_date_expr}), ' ', YEAR({lead_date_expr}))"
        p_sort_lead = f"CONCAT(CASE WHEN MONTH({lead_date_expr}) >= 4 THEN YEAR({lead_date_expr}) ELSE YEAR({lead_date_expr}) - 1 END, '-', LPAD(CASE WHEN MONTH({lead_date_expr}) >= 4 THEN MONTH({lead_date_expr}) - 3 ELSE MONTH({lead_date_expr}) + 9 END, 2, '0'))"
        p_group_lead = f"YEAR({lead_date_expr}), MONTH({lead_date_expr})"

        p_label_cont = f"CONCAT(MONTHNAME({contact_date_expr}), ' ', YEAR({contact_date_expr}))"
        p_sort_cont = f"CONCAT(CASE WHEN MONTH({contact_date_expr}) >= 4 THEN YEAR({contact_date_expr}) ELSE YEAR({contact_date_expr}) - 1 END, '-', LPAD(CASE WHEN MONTH({contact_date_expr}) >= 4 THEN MONTH({contact_date_expr}) - 3 ELSE MONTH({contact_date_expr}) + 9 END, 2, '0'))"
        p_group_cont = f"YEAR({contact_date_expr}), MONTH({contact_date_expr})"

    # Team-member mode: swap the period buckets for owner buckets. Everything
    # downstream (row assembly, totals, drilldown) is unchanged — only what a row
    # represents changes. Head-only; enforced below.
    is_user_grouping = str(group_by or "period").lower() in ("user", "owner", "team", "team_member")
    if is_user_grouping and not is_head:
        is_user_grouping = False

    p_user_lead = p_user_cont = "''"
    if is_user_grouping:
        owner_name = ("COALESCE(NULLIF((SELECT u.full_name FROM `tabUser` u WHERE u.name = {col}), ''), "
                      "{col}, 'Unassigned')")
        p_label_lead = owner_name.format(col="l.lead_owner")
        p_sort_lead = p_label_lead
        p_group_lead = "l.lead_owner"
        p_user_lead = "COALESCE(l.lead_owner, '')"

        p_label_cont = owner_name.format(col="assign_to")
        p_sort_cont = p_label_cont
        p_group_cont = "assign_to"
        p_user_cont = "COALESCE(assign_to, '')"

    # 1. Lead Data grouped by period
    lead_creation_date_where = where_lead_creation

    lead_created_data = frappe.db.sql(f"""
        SELECT 
            {p_label_lead} as label,
            {p_sort_lead} as f_sort,
            {p_user_lead} as grp_user,
            DATE_FORMAT(({lead_action_date}), '%%Y-%%m') as ym_num,
            MONTH({lead_action_date}) as m_num,
            MIN(DATE({lead_action_date})) as row_from_date,
            MAX(DATE({lead_action_date})) as row_to_date,
            SUM(CASE WHEN l.custom_lead_category = 'Enquiry' THEN 1 ELSE 0 END) as enq_c,
            SUM(CASE WHEN l.custom_lead_category = 'Enquiry' THEN COALESCE(l.custom_expected_revenue, 0) ELSE 0 END) as enq_v,
            SUM(CASE WHEN l.custom_lead_category = 'Pipeline' THEN 1 ELSE 0 END) as pipe_c,
            SUM(CASE WHEN l.custom_lead_category = 'Pipeline' THEN COALESCE(l.custom_expected_revenue, 0) ELSE 0 END) as pipe_v,
            SUM(CASE WHEN l.custom_lead_category = 'Order' THEN 1 ELSE 0 END) as ord_c,
            SUM(CASE WHEN l.custom_lead_category = 'Order' THEN COALESCE(l.custom_po_value, 0) ELSE 0 END) as ord_v,
            SUM(CASE WHEN l.custom_lead_category = 'Lost Enquiry' THEN 1 ELSE 0 END) as lenq_c,
            SUM(CASE WHEN l.custom_lead_category = 'Lost Enquiry' THEN COALESCE(l.custom_expected_revenue, 0) ELSE 0 END) as lenq_v,
            SUM(CASE WHEN l.custom_lead_category = 'Lost Pipeline' THEN 1 ELSE 0 END) as lpipe_c,
            SUM(CASE WHEN l.custom_lead_category = 'Lost Pipeline' THEN COALESCE(l.custom_expected_revenue, 0) ELSE 0 END) as lpipe_v
        FROM `tabLead` l
        {where_lead} {lead_creation_date_where}
        GROUP BY {p_group_lead} ORDER BY f_sort ASC
    """, params, as_dict=1)

    # Populate Lead Data
    lead_period_map = {}
    for r in lead_created_data:
        key = r.label
        if key not in lead_period_map:
            lead_period_map[key] = {
                "label": key, "f_sort": str(r.f_sort), "m_num": str(r.ym_num or r.label or ''),
                "grp_user": str(r.get("grp_user") or ''),
                "row_from_date": str(r.row_from_date) if r.row_from_date else '',
                "row_to_date": str(r.row_to_date) if r.row_to_date else '',
                "enq_c": 0, "enq_v": 0.0, "pipe_c": 0, "pipe_v": 0.0, "ord_c": 0, "ord_v": 0.0,
                "lenq_c": 0, "lenq_v": 0.0, "lpipe_c": 0, "lpipe_v": 0.0,
                "conv_bc_to_lead": 0, "conv_lead_to_order": 0, "convert_to": 0
            }
        p = lead_period_map[key]
        p["enq_c"] += int(r.enq_c or 0)
        p["enq_v"] += float(r.enq_v or 0.0)
        p["pipe_c"] += int(r.pipe_c or 0)
        p["pipe_v"] += float(r.pipe_v or 0.0)
        p["ord_c"] += int(r.ord_c or 0)
        p["ord_v"] += float(r.ord_v or 0.0)
        p["lenq_c"] += int(r.lenq_c or 0)
        p["lenq_v"] += float(r.lenq_v or 0.0)
        p["lpipe_c"] += int(r.lpipe_c or 0)
        p["lpipe_v"] += float(r.lpipe_v or 0.0)

    # Conversions for Lead Performance Table
    conv_params = {"sd": from_date, "ed": to_date, "usr": effective_user}
    bc_conv_filter = " WHERE docstatus < 2 AND " + BC_CONVERTED_COND_SQL
    lead_conv_filter = " WHERE docstatus < 2 "
    if effective_user:
        bc_conv_filter += " AND assign_to = %(usr)s "
        lead_conv_filter += " AND lead_owner = %(usr)s "
    ind_clause_conv = build_in_clause(industry_list, "ind", "industry", conv_params)
    src_clause_conv = build_in_clause(source_list, "src", "source", conv_params)
    bc_conv_filter += ind_clause_conv + src_clause_conv
    lead_conv_filter += ind_clause_conv + src_clause_conv
    # Shared with the drilldown so both bucket a contact into the same period
    bc_lead_conv_date = BC_CONVERSION_DATE_SQL

    # Dynamically build labels using bc_lead_conv_date
    if is_user_grouping:
        p_label_bc_conv = p_label_cont
        p_sort_bc_conv = p_sort_cont
        p_group_bc_conv = p_group_cont
    elif period_type == "daily":
        p_label_bc_conv = f"DATE_FORMAT({bc_lead_conv_date}, '%%d-%%b-%%Y (%%a)')"
        p_sort_bc_conv = f"DATE_FORMAT({bc_lead_conv_date}, '%%Y-%%m-%%d')"
        p_group_bc_conv = f"DATE_FORMAT({bc_lead_conv_date}, '%%Y-%%m-%%d')"
    elif period_type == "weekly":
        p_label_bc_conv = f"CONCAT('Week ', WEEK({bc_lead_conv_date}, 1), ' (', DATE_FORMAT(MIN(DATE_SUB(DATE({bc_lead_conv_date}), INTERVAL WEEKDAY({bc_lead_conv_date}) DAY)), '%%d-%%b-%%Y (%%a)'), ' to ', DATE_FORMAT(MIN(DATE_ADD(DATE_SUB(DATE({bc_lead_conv_date}), INTERVAL WEEKDAY({bc_lead_conv_date}) DAY), INTERVAL 6 DAY)), '%%d-%%b-%%Y (%%a)'), ')') "
        p_sort_bc_conv = f"CONCAT(YEAR({bc_lead_conv_date}), '-', LPAD(WEEK({bc_lead_conv_date}, 1), 2, '0'))"
        p_group_bc_conv = f"YEAR({bc_lead_conv_date}), WEEK({bc_lead_conv_date}, 1)"
    else: # monthly
        p_label_bc_conv = f"CONCAT(MONTHNAME({bc_lead_conv_date}), ' ', YEAR({bc_lead_conv_date}))"
        p_sort_bc_conv = f"CONCAT(CASE WHEN MONTH({bc_lead_conv_date}) >= 4 THEN YEAR({bc_lead_conv_date}) ELSE YEAR({bc_lead_conv_date}) - 1 END, '-', LPAD(CASE WHEN MONTH({bc_lead_conv_date}) >= 4 THEN MONTH({bc_lead_conv_date}) - 3 ELSE MONTH({bc_lead_conv_date}) + 9 END, 2, '0'))"
        p_group_bc_conv = f"YEAR({bc_lead_conv_date}), MONTH({bc_lead_conv_date})"

    # Lead -> Order is an event from the history table, one row per lead
    lead_order_conv_date = LEAD_ORDER_CONV_DATE_SQL
    if is_user_grouping:
        p_label_lo, p_sort_lo, p_group_lo = p_label_lead, p_sort_lead, p_group_lead
    else:
        p_label_lo, p_sort_lo, p_group_lo = period_sql_for(lead_order_conv_date, period_type)
    lead_conv_filter += f" AND {LEAD_ORDER_COND_SQL} "

    if from_date and to_date:
        bc_conv_filter += f" AND ({bc_lead_conv_date}) BETWEEN %(sd)s AND %(ed)s "
        lead_conv_filter += f" AND ({lead_order_conv_date}) BETWEEN %(sd)s AND %(ed)s "

    bc_conversions = frappe.db.sql(f"""
        SELECT {p_label_bc_conv} as label, {p_sort_bc_conv} as f_sort, DATE_FORMAT(({bc_lead_conv_date}), '%%Y-%%m') as ym_num, 
               MIN(DATE({bc_lead_conv_date})) as row_from_date, MAX(DATE({bc_lead_conv_date})) as row_to_date, COUNT(name) as cnt
        FROM `tabBusiness Contacts` {bc_conv_filter}
        GROUP BY {p_group_bc_conv}
    """, conv_params, as_dict=1)

    lead_conversions = frappe.db.sql(f"""
        SELECT {p_label_lo} as label, {p_sort_lo} as f_sort,
               DATE_FORMAT(({lead_order_conv_date}), '%%Y-%%m') as ym_num,
               MIN({lead_order_conv_date}) as row_from_date,
               MAX({lead_order_conv_date}) as row_to_date,
               COUNT(DISTINCT l.name) as cnt
        FROM `tabLead` l
        {lead_conv_filter}
        GROUP BY {p_group_lo}
    """, conv_params, as_dict=1)

    def widen_row_span(row, frm, to):
        """Row click drilldowns filter on row_from_date..row_to_date.

        That span is seeded by whichever query created the period row (usually the
        lead query), so a conversion dated outside it would be counted but then
        missing from the drilldown. Widen the span to cover every contributor.
        """
        frm = str(frm) if frm else ""
        to = str(to) if to else ""
        if frm and (not row.get("row_from_date") or frm < row["row_from_date"]):
            row["row_from_date"] = frm
        if to and (not row.get("row_to_date") or to > row["row_to_date"]):
            row["row_to_date"] = to

    for c in bc_conversions:
        key = c.label
        if key not in lead_period_map:
            lead_period_map[key] = {
                "label": key, "f_sort": str(c.f_sort), "m_num": str(c.ym_num or c.label or ''),
                "row_from_date": str(c.row_from_date) if c.row_from_date else '',
                "row_to_date": str(c.row_to_date) if c.row_to_date else '',
                "enq_c": 0, "enq_v": 0.0, "pipe_c": 0, "pipe_v": 0.0, "ord_c": 0, "ord_v": 0.0,
                "lenq_c": 0, "lenq_v": 0.0, "lpipe_c": 0, "lpipe_v": 0.0,
                "conv_bc_to_lead": 0, "conv_lead_to_order": 0, "convert_to": 0
            }
        widen_row_span(lead_period_map[key], c.row_from_date, c.row_to_date)
        lead_period_map[key]["conv_bc_to_lead"] += int(c.cnt or 0)
        lead_period_map[key]["convert_to"] += int(c.cnt or 0)

    for c in lead_conversions:
        key = c.label
        if key not in lead_period_map:
            lead_period_map[key] = {
                "label": key, "f_sort": str(c.f_sort), "m_num": str(c.ym_num or c.label or ''),
                "row_from_date": str(c.row_from_date) if c.row_from_date else '',
                "row_to_date": str(c.row_to_date) if c.row_to_date else '',
                "enq_c": 0, "enq_v": 0.0, "pipe_c": 0, "pipe_v": 0.0, "ord_c": 0, "ord_v": 0.0,
                "lenq_c": 0, "lenq_v": 0.0, "lpipe_c": 0, "lpipe_v": 0.0,
                "conv_bc_to_lead": 0, "conv_lead_to_order": 0, "convert_to": 0
            }
        widen_row_span(lead_period_map[key], c.row_from_date, c.row_to_date)
        lead_period_map[key]["conv_lead_to_order"] += int(c.cnt or 0)
        lead_period_map[key]["convert_to"] += int(c.cnt or 0)

    sorted_leads = sorted(lead_period_map.values(), key=lambda x: x["f_sort"])

    # 3. Contact Breakup Data
    contacts_data = frappe.db.sql(f"""
        SELECT 
            {p_label_cont} as label,
            {p_sort_cont} as f_sort,
            {p_user_cont} as grp_user,
            DATE_FORMAT(({contact_action_date}), '%%Y-%%m') as ym_num,
            MONTH({contact_action_date}) as m_num,
            MIN(DATE({contact_action_date})) as row_from_date,
            MAX(DATE({contact_action_date})) as row_to_date,
            SUM(CASE WHEN status = 'Open' THEN 1 ELSE 0 END) as o,
            SUM(CASE WHEN status = 'Converted to Lead' THEN 1 ELSE 0 END) as c,
            SUM(CASE WHEN status = 'Existing Customer' THEN 1 ELSE 0 END) as e
        FROM `tabBusiness Contacts`
        {where_cont}
        GROUP BY {p_group_cont} ORDER BY f_sort ASC
    """, params, as_dict=1)

    sorted_contacts = []
    for r in contacts_data:
        sorted_contacts.append({
            "label": r.label, "f_sort": str(r.f_sort), "m_num": str(r.ym_num or r.label or ''),
            "grp_user": str(r.get("grp_user") or ''),
            "row_from_date": str(r.row_from_date) if r.row_from_date else '',
            "row_to_date": str(r.row_to_date) if r.row_to_date else '',
            "o": int(r.o or 0), "c": int(r.c or 0), "e": int(r.e or 0)
        })

    # Calculate Totals
    lt = {"enq_c": 0, "enq_v": 0.0, "pipe_c": 0, "pipe_v": 0.0, "ord_c": 0, "ord_v": 0.0, "lenq_c": 0, "lenq_v": 0.0, "lpipe_c": 0, "lpipe_v": 0.0, "conv_bc_to_lead": 0, "conv_lead_to_order": 0, "convert_to": 0}
    for l in sorted_leads:
        lt["enq_c"] += l["enq_c"]
        lt["enq_v"] += l["enq_v"]
        lt["pipe_c"] += l["pipe_c"]
        lt["pipe_v"] += l["pipe_v"]
        lt["ord_c"] += l["ord_c"]
        lt["ord_v"] += l["ord_v"]
        lt["lenq_c"] += l["lenq_c"]
        lt["lenq_v"] += l["lenq_v"]
        lt["lpipe_c"] += l["lpipe_c"]
        lt["lpipe_v"] += l["lpipe_v"]
        lt["conv_bc_to_lead"] += l["conv_bc_to_lead"]
        lt["conv_lead_to_order"] += l["conv_lead_to_order"]
        lt["convert_to"] += l["convert_to"]

    # Calculate active lead category totals directly from tabLead
    where_lead_creation = f" AND ({lead_action_date}) BETWEEN %(sd)s AND %(ed)s " if (from_date and to_date) else ""
    active_lead_counts = frappe.db.sql(f"""
        SELECT 
            SUM(CASE WHEN l.custom_lead_category = 'Enquiry' THEN 1 ELSE 0 END) as enq_c,
            SUM(CASE WHEN l.custom_lead_category = 'Enquiry' THEN COALESCE(l.custom_expected_revenue, 0) ELSE 0 END) as enq_v,
            SUM(CASE WHEN l.custom_lead_category = 'Pipeline' THEN 1 ELSE 0 END) as pipe_c,
            SUM(CASE WHEN l.custom_lead_category = 'Pipeline' THEN COALESCE(l.custom_expected_revenue, 0) ELSE 0 END) as pipe_v,
            SUM(CASE WHEN l.custom_lead_category = 'Order' THEN 1 ELSE 0 END) as ord_c,
            SUM(CASE WHEN l.custom_lead_category = 'Order' THEN COALESCE(l.custom_po_value, 0) ELSE 0 END) as ord_v,
            SUM(CASE WHEN l.custom_lead_category = 'Lost Enquiry' THEN 1 ELSE 0 END) as lenq_c,
            SUM(CASE WHEN l.custom_lead_category = 'Lost Enquiry' THEN COALESCE(l.custom_expected_revenue, 0) ELSE 0 END) as lenq_v,
            SUM(CASE WHEN l.custom_lead_category = 'Lost Pipeline' THEN 1 ELSE 0 END) as lpipe_c,
            SUM(CASE WHEN l.custom_lead_category = 'Lost Pipeline' THEN COALESCE(l.custom_expected_revenue, 0) ELSE 0 END) as lpipe_v
        FROM `tabLead` l
        {where_lead}{where_lead_creation}
    """, params, as_dict=1)

    if active_lead_counts:
        alc = active_lead_counts[0]
        lt["enq_c"] = int(alc.get("enq_c") or 0)
        lt["enq_v"] = float(alc.get("enq_v") or 0.0)
        lt["pipe_c"] = int(alc.get("pipe_c") or 0)
        lt["pipe_v"] = float(alc.get("pipe_v") or 0.0)
        lt["ord_c"] = int(alc.get("ord_c") or 0)
        lt["ord_v"] = float(alc.get("ord_v") or 0.0)
        lt["lenq_c"] = int(alc.get("lenq_c") or 0)
        lt["lenq_v"] = float(alc.get("lenq_v") or 0.0)
        lt["lpipe_c"] = int(alc.get("lpipe_c") or 0)
        lt["lpipe_v"] = float(alc.get("lpipe_v") or 0.0)

    ct = {"o": 0, "c": 0, "e": 0}
    for c in sorted_contacts:
        ct["o"] += c["o"]
        ct["c"] += c["c"]
        ct["e"] += c["e"]

    return {
        "group_by": "user" if is_user_grouping else "period",
        "leads": sorted_leads,
        "lead_totals": lt,
        "contacts": sorted_contacts,
        "contact_totals": ct
    }

@frappe.whitelist()
def get_event_activity_breakup_data(from_date=None, to_date=None, user=None, period_type="monthly", reference_type="All", activity_basis="completed", fiscal_year=None, period=None, month=None, group_by="period", **kwargs):
    from_date, to_date = parse_period_and_dates(period, fiscal_year, from_date, to_date, month=month or kwargs.get("month"))

    is_head = "DAC CRM Head" in frappe.get_roles()
    effective_user = user if (is_head and user) else (None if (is_head and not user) else frappe.session.user)

    params = {"sd": from_date, "ed": to_date, "usr": effective_user}
    
    # Fetch all distinct non-empty categories from tabEvent Activity
    cat_docs = frappe.db.sql("""
        SELECT DISTINCT category 
        FROM `tabEvent Activity` 
        WHERE docstatus < 2 AND category IS NOT NULL AND category != ''
        ORDER BY category ASC
    """, as_dict=1)
    categories = [c.category for c in cat_docs]

    # Filter String building
    base_filter = " WHERE docstatus < 2 "
    if effective_user:
        base_filter += " AND assigned_to = %(usr)s "
    if reference_type and reference_type != "All":
        base_filter += " AND reference_type = %(ref_type)s "
        params["ref_type"] = reference_type

    cp_filter = base_filter + " AND status = 'Completed' "
    if from_date and to_date:
        cp_filter += " AND DATE(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)) BETWEEN %(sd)s AND %(ed)s "

    if period_type == "daily":
        cp_group = "DATE(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified))"
        cp_label = "DATE_FORMAT(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified), '%%d-%%b-%%Y (%%a)')"
        cp_sort = "DATE(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified))"
    elif period_type == "weekly":
        cp_group = "YEAR(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)), WEEK(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified), 1)"
        cp_label = "CONCAT('Week ', WEEK(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified), 1), ' (', DATE_FORMAT(MIN(DATE_SUB(DATE(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)), INTERVAL WEEKDAY(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)) DAY)), '%%d-%%b-%%Y (%%a)'), ' to ', DATE_FORMAT(MIN(DATE_ADD(DATE_SUB(DATE(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)), INTERVAL WEEKDAY(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)) DAY), INTERVAL 6 DAY)), '%%d-%%b-%%Y (%%a)'), ')')"
        cp_sort = "CONCAT(YEAR(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)), '-', LPAD(WEEK(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)), 2, '0'))"
    else:
        cp_group = "YEAR(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)), MONTH(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified))"
        cp_label = "CONCAT(MONTHNAME(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)), ' ', YEAR(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)))"
        cp_sort = "CONCAT(YEAR(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)), '-', LPAD(MONTH(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified)), 2, '0'))"

    # Team-member mode: bucket by the person the activity is assigned to instead
    # of by date. Head-only, same as the lead/contact tables.
    ea_user_grouping = str(group_by or "period").lower() in ("user", "owner", "team", "team_member")
    if ea_user_grouping and not is_head:
        ea_user_grouping = False
    cp_user = "''"
    if ea_user_grouping:
        cp_label = ("COALESCE(NULLIF((SELECT u.full_name FROM `tabUser` u WHERE u.name = assigned_to), ''), "
                    "assigned_to, 'Unassigned')")
        cp_sort = cp_label
        cp_group = "assigned_to"
        cp_user = "COALESCE(assigned_to, '')"

    cat_case_parts = []
    for idx, cat in enumerate(categories):
        params[f"cat_{idx}"] = cat
        cat_case_parts.append(f"SUM(CASE WHEN category = %(cat_{idx})s THEN 1 ELSE 0 END) as cat_{idx}_total")
        cat_case_parts.append(f"SUM(CASE WHEN category = %(cat_{idx})s AND reference_type = 'Lead' THEN 1 ELSE 0 END) as cat_{idx}_lead")
        cat_case_parts.append(f"SUM(CASE WHEN category = %(cat_{idx})s AND reference_type = 'Business Contacts' THEN 1 ELSE 0 END) as cat_{idx}_cont")

    cat_case_parts.append("SUM(CASE WHEN reference_type = 'Lead' THEN 1 ELSE 0 END) as lead_cnt")
    cat_case_parts.append("SUM(CASE WHEN reference_type = 'Business Contacts' THEN 1 ELSE 0 END) as contact_cnt")
    cat_case_parts.append("COUNT(name) as total_activities")
    cat_case_parts.append("MIN(DATE(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified))) as row_from_date")
    cat_case_parts.append("MAX(DATE(COALESCE(ends_on, actual_checked_out_at, actual_visit_at, modified))) as row_to_date")

    cat_case_sql = ",\n".join(cat_case_parts)

    cp_activities = frappe.db.sql(f"""
        SELECT {cp_label} as period_label, {cp_sort} as sort_key, {cp_user} as grp_user, {cat_case_sql}
        FROM `tabEvent Activity` {cp_filter}
        GROUP BY {cp_group} ORDER BY sort_key ASC
    """, params, as_dict=1)

    activities = []
    category_totals = {c: {"total": 0, "lead": 0, "cont": 0} for c in categories}
    overall_total = 0
    overall_lead = 0
    overall_cont = 0

    for row in cp_activities:
        act_row = {
            "period_label": row.period_label,
            "sort_key": str(row.sort_key),
            "grp_user": str(row.get("grp_user") or ""),
            "row_from_date": str(row.row_from_date) if row.row_from_date else "",
            "row_to_date": str(row.row_to_date) if row.row_to_date else "",
            "lead_cnt": int(row.lead_cnt or 0),
            "contact_cnt": int(row.contact_cnt or 0),
            "total_completed": int(row.total_activities or 0),
            "categories": {}
        }
        overall_total += act_row["total_completed"]
        overall_lead += act_row["lead_cnt"]
        overall_cont += act_row["contact_cnt"]

        for idx, cat in enumerate(categories):
            tot = int(row.get(f"cat_{idx}_total") or 0)
            ld = int(row.get(f"cat_{idx}_lead") or 0)
            ct = int(row.get(f"cat_{idx}_cont") or 0)

            act_row["categories"][cat] = {
                "total": tot,
                "lead": ld,
                "cont": ct
            }
            category_totals[cat]["total"] += tot
            category_totals[cat]["lead"] += ld
            category_totals[cat]["cont"] += ct

        activities.append(act_row)

    # 5 Specific Event Activity Number Card calculations
    today_str = frappe.utils.today()
    user_cond_ea = " AND assigned_to = %(usr)s " if effective_user else ""
    user_cond_lead = " AND l.lead_owner = %(usr)s " if effective_user else ""
    user_cond_cont = " AND assign_to = %(usr)s " if effective_user else ""
    
    date_cond_ea = " AND DATE(COALESCE(ends_on, modified)) BETWEEN %(sd)s AND %(ed)s " if (from_date and to_date) else ""
    date_cond_lead = " AND DATE(l.creation) BETWEEN %(sd)s AND %(ed)s " if (from_date and to_date) else ""
    date_cond_cont = " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s " if (from_date and to_date) else ""
    
    ref_cond_ea = ""
    if reference_type == "Lead":
        ref_cond_ea = " AND reference_type = 'Lead' "
    elif reference_type == "Business Contacts":
        ref_cond_ea = " AND reference_type = 'Business Contacts' "

    overdue_count = frappe.db.sql(f"""
        SELECT COUNT(name) FROM `tabEvent Activity`
        WHERE docstatus < 2 AND status = 'Open'
          AND (
            (ends_on IS NOT NULL AND DATE(ends_on) < '{today_str}')
            OR (ends_on IS NULL AND starts_on IS NOT NULL AND DATE(starts_on) < '{today_str}')
          ) {user_cond_ea}{ref_cond_ea}
    """, params)[0][0] or 0

    today_count = frappe.db.sql(f"""
        SELECT COUNT(name) FROM `tabEvent Activity`
        WHERE docstatus < 2 AND status = 'Open'
          AND (
            (ends_on IS NOT NULL AND DATE(ends_on) >= '{today_str}')
            OR (ends_on IS NULL AND starts_on IS NOT NULL AND DATE(starts_on) >= '{today_str}')
          ) {user_cond_ea}{date_cond_ea}{ref_cond_ea}
    """, params)[0][0] or 0

    open_total_count = frappe.db.sql(f"""
        SELECT COUNT(name) FROM `tabEvent Activity`
        WHERE docstatus < 2 AND status = 'Open' {user_cond_ea}{date_cond_ea}{ref_cond_ea}
    """, params)[0][0] or 0

    lead_noact_count = 0
    if reference_type in ("All", "Lead"):
        lead_noact_count = frappe.db.sql(f"""
            SELECT COUNT(l.name) FROM `tabLead` l
            WHERE l.docstatus < 2
            AND l.name NOT IN (SELECT DISTINCT reference_name FROM `tabEvent Activity` WHERE reference_type = 'Lead' AND docstatus < 2)
            {user_cond_lead} {date_cond_lead}
        """, params)[0][0] or 0

    cont_noact_count = 0
    if reference_type in ("All", "Business Contacts"):
        cont_noact_count = frappe.db.sql(f"""
            SELECT COUNT(name) FROM `tabBusiness Contacts`
            WHERE docstatus < 2
            AND name NOT IN (SELECT DISTINCT reference_name FROM `tabEvent Activity` WHERE reference_type = 'Business Contacts' AND docstatus < 2)
            {user_cond_cont} {date_cond_cont}
        """, params)[0][0] or 0

    return {
        "group_by": "user" if ea_user_grouping else "period",
        "categories": categories,
        "activities": activities,
        "totals": {
            "category_totals": category_totals,
            "total_completed": overall_total,
            "cp_lead_cnt": overall_lead,
            "cp_contact_cnt": overall_cont,
            "overdue_count": overdue_count,
            "today_count": today_count,
            "open_total_count": open_total_count,
            "lead_noact_count": lead_noact_count,
            "cont_noact_count": cont_noact_count
        }
    }


@frappe.whitelist()
def get_card_detail_records(card_type, from_date=None, to_date=None, user=None, industry=None, source=None, fiscal_year=None, period=None, month=None, **kwargs):
    """Return list of records for a dashboard number-card drilldown modal with 100% exact count parity and rich details."""
    from_date, to_date = parse_period_and_dates(period, fiscal_year, from_date, to_date, month=month or kwargs.get("month"))

    is_head = "DAC CRM Head" in frappe.get_roles()
    effective_user = user if (is_head and user) else (None if (is_head and not user) else frappe.session.user)

    # Multiselect filters — must mirror get_tabular_dashboard_data so drilldown counts match the cards
    industry_list = as_filter_list(industry)
    source_list = as_filter_list(source)

    params = {"sd": from_date, "ed": to_date, "usr": effective_user}
    contact_action_date = """
        CASE
            WHEN status = 'Open' THEN DATE(creation)
            WHEN status = 'Converted to Lead' THEN DATE(COALESCE(
                (SELECT MAX(sh.updated_at) FROM `tabLead Status Change History` sh
                 WHERE sh.parent = `tabBusiness Contacts`.name AND sh.parenttype = 'Business Contacts' AND sh.parentfield = 'contact_status_change_history' AND sh.new_status = 'Converted to Lead'),
                modified,
                creation
            ))
            WHEN status = 'Existing Customer' THEN DATE(COALESCE(
                (SELECT MAX(sh.updated_at) FROM `tabLead Status Change History` sh
                 WHERE sh.parent = `tabBusiness Contacts`.name AND sh.parenttype = 'Business Contacts' AND sh.parentfield = 'contact_status_change_history' AND sh.new_status = 'Existing Customer'),
                modified,
                creation
            ))
            ELSE DATE(modified)
        END
    """
    date_cond_cont = f" AND ({contact_action_date}) BETWEEN %(sd)s AND %(ed)s " if (from_date and to_date) else ""
    # conv_bc is a *conversion* view, not a status view, so it buckets on the
    # conversion date exactly as get_tabular_dashboard_data counts it.
    date_cond_conv = (f" AND ({BC_CONVERSION_DATE_SQL}) BETWEEN %(sd)s AND %(ed)s "
                      if (from_date and to_date) else "")
    user_cond_lead = " AND l.lead_owner = %(usr)s " if effective_user else ""
    user_cond_cont = " AND assign_to = %(usr)s " if effective_user else ""
    scope_cond_lead = (build_in_clause(industry_list, "ind", "l.industry", params)
                       + build_in_clause(source_list, "src", "l.source", params))
    scope_cond_cont = (build_in_clause(industry_list, "ind", "industry", params)
                       + build_in_clause(source_list, "src", "source", params))

    records = []

    # ---- CONTACT cards ----
    if card_type in ("c_open", "c_conv", "c_exist", "conv_bc"):
        status_cond = " status = 'Open' " if card_type == "c_open" else (
            " status = 'Converted to Lead' " if card_type == "c_conv" else (
                " status = 'Existing Customer' " if card_type == "c_exist" else (
                    BC_CONVERTED_COND_SQL
                )
            )
        )
        cont_date_cond = date_cond_conv if card_type == "conv_bc" else date_cond_cont
        records = frappe.db.sql(f"""
            SELECT name, contact_name, COALESCE(organization_name, contact_name) as company,
                   mobile_number as mobile_no, email_id, industry, status, assign_to as owner,
                   lead_id,
                   (SELECT l2.lead_name FROM `tabLead` l2
                     WHERE l2.name = `tabBusiness Contacts`.lead_id) as lead_title,
                   city, country, DATE_FORMAT(creation,'%%d-%%b-%%Y') as created_on,
                   DATEDIFF(NOW(), creation) as age_days,
                   DATE_FORMAT(next_follow_up_date,'%%d-%%b-%%Y') as next_followup,
                   last_completion_notes,
                   (SELECT COUNT(q.name) FROM `tabQuotation` q
                    WHERE q.docstatus < 2 AND (q.party_name = lead_id OR q.custom_lead_id = lead_id)) as quotation_count
            FROM `tabBusiness Contacts`
            WHERE docstatus < 2 AND {status_cond}
            {user_cond_cont}{scope_cond_cont}{cont_date_cond}
            ORDER BY creation DESC LIMIT 1000
        """, params, as_dict=1)
        for r in records:
            r["doctype"] = "Business Contacts"

    # ---- LEAD cards ----
    elif card_type in ("l_enq", "l_pipe", "l_ord", "l_lenq", "l_lpipe", "conv_lead"):
        cat_map = {
            "l_enq": "Enquiry",
            "l_pipe": "Pipeline",
            "l_ord": "Order",
            "l_lenq": "Lost Enquiry",
            "l_lpipe": "Lost Pipeline",
            "conv_lead": "Order"
        }
        cat = cat_map[card_type]
        params["cat"] = cat

        lead_action_date = f"""
            CASE
                WHEN l.custom_lead_category = 'Enquiry' THEN DATE(l.creation)
                ELSE COALESCE(
                    (SELECT DATE(MAX(sh.updated_at)) FROM `tabLead Status Change History` sh
                     WHERE sh.parent = l.name AND sh.parenttype = 'Lead'
                   AND sh.parentfield = 'custom_lead_status_change_history'
                   AND sh.new_status = l.custom_lead_category
                   {LEAD_OWN_HISTORY_SQL}),
                    DATE(l.creation)
                )
            END
        """
        if card_type == "conv_lead":
            # Conversion view: leads whose FIRST Order transition lands in range,
            # matching how get_tabular_dashboard_data counts them.
            cat_cond = f" AND {LEAD_ORDER_COND_SQL} "
            lead_date_cond = (f" AND ({LEAD_ORDER_CONV_DATE_SQL}) BETWEEN %(sd)s AND %(ed)s "
                              if (from_date and to_date) else "")
        else:
            cat_cond = " AND l.custom_lead_category = %(cat)s "
            lead_date_cond = (f" AND ({lead_action_date}) BETWEEN %(sd)s AND %(ed)s "
                              if (from_date and to_date) else "")

        records = frappe.db.sql(f"""
            SELECT l.name, l.lead_name, l.company_name as company,
                   l.mobile_no, l.email_id, l.industry,
                   l.custom_lead_category as category, l.lead_owner as owner,
                   CASE WHEN l.custom_lead_category = 'Order' THEN COALESCE(l.custom_po_value, 0)
                        ELSE COALESCE(l.custom_expected_revenue, 0) END as po_value,
                   COALESCE(l.custom_expected_revenue, 0) as expected_revenue,
                   l.territory, l.city, l.state, l.country, l.source,
                   l.custom_lead_type, DATE_FORMAT(l.custom_expected_closure_date,'%%d-%%b-%%Y') as expected_close,
                   DATE_FORMAT(l.creation,'%%d-%%b-%%Y') as created_on,
                   DATEDIFF(NOW(), l.creation) as age_days,
                   DATE_FORMAT(l.custom_next_followup_date,'%%d-%%b-%%Y') as next_followup,
                   (SELECT notes FROM `tabEvent Activity`
                    WHERE reference_type = 'Lead' AND reference_name = l.name AND status = 'Completed' AND docstatus < 2
                    ORDER BY ends_on DESC LIMIT 1) as last_completed_notes,
                   (SELECT COUNT(q.name) FROM `tabQuotation` q
                    WHERE q.docstatus < 2 AND (q.party_name = l.name OR q.custom_lead_id = l.name)) as quotation_count,
                   l.custom_expected_closure_date as closure_date
            FROM `tabLead` l
            WHERE l.docstatus < 2
              {cat_cond}
              {user_cond_lead}{scope_cond_lead}{lead_date_cond}
            ORDER BY l.creation DESC LIMIT 1000
        """, params, as_dict=1)
        for r in records:
            r["doctype"] = "Lead"

    # ---- ACTIVITY cards ----
    elif card_type in ("ea_overdue", "ea_today", "ea_lead_noact", "ea_cont_noact", "ea_open"):
        today_str = frappe.utils.today()
        date_cond_lead = " AND DATE(l.creation) BETWEEN %(sd)s AND %(ed)s " if (from_date and to_date) else ""
        date_cond_cont = " AND DATE(bc.creation) BETWEEN %(sd)s AND %(ed)s " if (from_date and to_date) else ""

        reference_type = kwargs.get("reference_type", "All")
        ref_cond_ea = ""
        if reference_type == "Lead":
            ref_cond_ea = " AND ea.reference_type = 'Lead' "
        elif reference_type == "Business Contacts":
            ref_cond_ea = " AND ea.reference_type = 'Business Contacts' "

        if card_type in ("ea_overdue", "ea_today", "ea_open"):
            if card_type == "ea_overdue":
                date_cond_ea = ""
                date_cond = f""" AND ea.status = 'Open' AND (
                    (ea.ends_on IS NOT NULL AND DATE(ea.ends_on) < '{today_str}')
                    OR (ea.ends_on IS NULL AND ea.starts_on IS NOT NULL AND DATE(ea.starts_on) < '{today_str}')
                ) """
            elif card_type == "ea_today":
                date_cond_ea = " AND DATE(COALESCE(ea.ends_on, ea.modified)) BETWEEN %(sd)s AND %(ed)s " if (from_date and to_date) else ""
                date_cond = f""" AND ea.status = 'Open' AND (
                    (ea.ends_on IS NOT NULL AND DATE(ea.ends_on) >= '{today_str}')
                    OR (ea.ends_on IS NULL AND ea.starts_on IS NOT NULL AND DATE(ea.starts_on) >= '{today_str}')
                ) """
            else:
                date_cond_ea = " AND DATE(COALESCE(ea.ends_on, ea.modified)) BETWEEN %(sd)s AND %(ed)s " if (from_date and to_date) else ""
                date_cond = " AND ea.status = 'Open' "

            user_cond_ea = " AND ea.assigned_to = %(usr)s " if effective_user else ""
            records = frappe.db.sql(f"""
                SELECT ea.name, ea.subject, ea.category, ea.reference_type, ea.reference_name,
                       ea.assigned_to as owner, ea.status, ea.description, ea.notes,
                       ea.starts_on,
                       DATE_FORMAT(ea.starts_on,'%%d-%%b-%%Y %%h:%%i %%p') as start_date,
                       DATE_FORMAT(ea.ends_on,'%%d-%%b-%%Y %%h:%%i %%p') as end_date,
                       CASE
                           WHEN ea.reference_type = 'Lead' THEN l.lead_name
                           WHEN ea.reference_type = 'Business Contacts' THEN COALESCE(bc.organization_name, bc.contact_name)
                           ELSE COALESCE(ea.reference_doc_name, ea.reference_name)
                       END as ref_display_name,
                       CASE
                           WHEN ea.reference_type = 'Lead' THEN l.company_name
                           WHEN ea.reference_type = 'Business Contacts' THEN bc.organization_name
                           ELSE ''
                       END as ref_extra,
                       CASE
                           WHEN ea.reference_type = 'Lead' THEN l.mobile_no
                           WHEN ea.reference_type = 'Business Contacts' THEN bc.mobile_number
                           ELSE ''
                       END as ref_mobile,
                       CASE
                           WHEN ea.reference_type = 'Lead' THEN l.email_id
                           WHEN ea.reference_type = 'Business Contacts' THEN bc.email_id
                           ELSE ''
                       END as ref_email
                FROM `tabEvent Activity` ea
                LEFT JOIN `tabLead` l ON l.name = ea.reference_name AND ea.reference_type = 'Lead'
                LEFT JOIN `tabBusiness Contacts` bc ON bc.name = ea.reference_name AND ea.reference_type = 'Business Contacts'
                WHERE ea.docstatus < 2 {user_cond_ea}{date_cond}{date_cond_ea}{ref_cond_ea}
                ORDER BY 
                    CASE WHEN DATE(COALESCE(ea.ends_on, ea.starts_on)) = '{today_str}' THEN 0 ELSE 1 END ASC,
                    COALESCE(ea.ends_on, ea.starts_on) ASC LIMIT 1000
            """, params, as_dict=1)
            for r in records:
                r["doctype"] = "Event Activity"

        elif card_type == "ea_lead_noact":
            if reference_type in ("All", "Lead"):
                records = frappe.db.sql(f"""
                    SELECT l.name, l.lead_name, l.company_name as company, l.mobile_no, l.email_id, l.industry,
                           l.custom_lead_category as category, l.lead_owner as owner,
                           COALESCE(l.custom_expected_revenue, 0) as expected_revenue,
                           l.source, l.custom_lead_type, DATE_FORMAT(l.custom_expected_closure_date,'%%d-%%b-%%Y') as expected_close,
                           DATE_FORMAT(l.creation,'%%d-%%b-%%Y') as created_on,
                           DATEDIFF(NOW(), l.creation) as age_days,
                           DATE_FORMAT(l.custom_next_followup_date,'%%d-%%b-%%Y') as next_followup,
                           (SELECT notes FROM `tabEvent Activity`
                            WHERE reference_type = 'Lead' AND reference_name = l.name AND status = 'Completed' AND docstatus < 2
                            ORDER BY ends_on DESC LIMIT 1) as last_completed_notes,
                           (SELECT COUNT(q.name) FROM `tabQuotation` q
                            WHERE q.docstatus < 2 AND (q.party_name = l.name OR q.custom_lead_id = l.name)) as quotation_count,
                           l.custom_expected_closure_date as closure_date
                    FROM `tabLead` l
                    WHERE l.docstatus < 2
                      AND l.name NOT IN (SELECT DISTINCT reference_name FROM `tabEvent Activity` WHERE reference_type = 'Lead' AND docstatus < 2)
                    {user_cond_lead}{date_cond_lead}
                    ORDER BY l.creation DESC LIMIT 1000
                """, params, as_dict=1)
                for r in records:
                    r["doctype"] = "Lead"
            else:
                records = []

        elif card_type == "ea_cont_noact":
            if reference_type in ("All", "Business Contacts"):
                records = frappe.db.sql(f"""
                    SELECT bc.name, bc.contact_name, COALESCE(bc.organization_name, bc.contact_name) as company,
                           bc.mobile_number as mobile_no, bc.email_id, bc.industry, bc.status, bc.assign_to as owner,
                           bc.city, bc.country, DATE_FORMAT(bc.creation,'%%d-%%b-%%Y') as created_on,
                           DATEDIFF(NOW(), bc.creation) as age_days,
                           DATE_FORMAT(bc.next_follow_up_date,'%%d-%%b-%%Y') as next_followup,
                           bc.last_completion_notes
                    FROM `tabBusiness Contacts` bc
                    WHERE bc.docstatus < 2
                      AND bc.name NOT IN (SELECT DISTINCT reference_name FROM `tabEvent Activity` WHERE reference_type = 'Business Contacts' AND docstatus < 2)
                    {user_cond_cont}{date_cond_cont}
                    ORDER BY bc.creation DESC LIMIT 1000
                """, params, as_dict=1)
                for r in records:
                    r["doctype"] = "Business Contacts"
            else:
                records = []

    return {"records": records, "card_type": card_type}


@frappe.whitelist()
def get_activity_detail_records(category=None, category_group=None, from_date=None, to_date=None, user=None, reference_type="All", fiscal_year=None, period=None, month=None, **kwargs):
    """Return Event Activity records for a specific category drilldown modal with exact date matching and rich details."""
    from_date, to_date = parse_period_and_dates(period, fiscal_year, from_date, to_date, month=month or kwargs.get("month"))

    is_head = "DAC CRM Head" in frappe.get_roles()
    effective_user = user if (is_head and user) else (None if (is_head and not user) else frappe.session.user)

    cat_name = category or category_group

    params = {"sd": from_date, "ed": to_date, "usr": effective_user}

    base_filter = " WHERE ea.docstatus < 2 AND ea.status = 'Completed' "
    if effective_user:
        base_filter += " AND ea.assigned_to = %(usr)s "
    if reference_type and reference_type != "All":
        params["ref_type"] = reference_type
        base_filter += " AND ea.reference_type = %(ref_type)s "

    if from_date and to_date:
        base_filter += " AND DATE(COALESCE(ea.ends_on, ea.actual_checked_out_at, ea.actual_visit_at, ea.modified)) BETWEEN %(sd)s AND %(ed)s "

    if cat_name and cat_name != "All":
        params["cat"] = cat_name
        base_filter += " AND ea.category = %(cat)s "

    records = frappe.db.sql(f"""
        SELECT ea.name, ea.subject, ea.category, ea.reference_type, ea.reference_name,
               ea.assigned_to as owner, ea.status, ea.description, ea.notes,
               ea.starts_on,
               DATE_FORMAT(ea.starts_on,'%%d-%%b-%%Y %%h:%%i %%p') as start_date,
               DATE_FORMAT(COALESCE(ea.ends_on, ea.modified),'%%d-%%b-%%Y %%h:%%i %%p') as end_date,
               CASE
                   WHEN ea.reference_type = 'Lead' THEN l.lead_name
                   WHEN ea.reference_type = 'Business Contacts' THEN COALESCE(bc.organization_name, bc.contact_name)
                   ELSE COALESCE(ea.reference_doc_name, ea.reference_name)
               END as ref_display_name,
               CASE
                   WHEN ea.reference_type = 'Lead' THEN l.company_name
                   WHEN ea.reference_type = 'Business Contacts' THEN bc.organization_name
                   ELSE ''
               END as ref_extra,
               CASE
                   WHEN ea.reference_type = 'Lead' THEN l.mobile_no
                   WHEN ea.reference_type = 'Business Contacts' THEN bc.mobile_number
                   ELSE ''
               END as ref_mobile,
               CASE
                   WHEN ea.reference_type = 'Lead' THEN l.email_id
                   WHEN ea.reference_type = 'Business Contacts' THEN bc.email_id
                   ELSE ''
               END as ref_email
        FROM `tabEvent Activity` ea
        LEFT JOIN `tabLead` l ON l.name = ea.reference_name AND ea.reference_type = 'Lead'
        LEFT JOIN `tabBusiness Contacts` bc ON bc.name = ea.reference_name AND ea.reference_type = 'Business Contacts'
        {base_filter}
        ORDER BY ea.ends_on DESC LIMIT 1000
    """, params, as_dict=1)

    for r in records:
        r["doctype"] = "Event Activity"

    return {"records": records, "category": cat_name}


def require_crm_head():
    """Admin Review analytics are head-only. Hiding the UI is not a control."""
    if "DAC CRM Head" not in frappe.get_roles(frappe.session.user):
        frappe.throw("Only DAC CRM Head can view CRM analytics.", frappe.PermissionError)


def _analytics_lead_where(restrict_user, ind_list, src_list, from_date, to_date, params):
    wh = "WHERE l.docstatus < 2"
    if restrict_user:
        wh += " AND l.lead_owner = %(usr)s"
        params["usr"] = restrict_user
    wh += build_in_clause(ind_list, "ind", "l.industry", params)
    wh += build_in_clause(src_list, "src", "l.source", params)
    if from_date and to_date:
        wh += " AND DATE(l.creation) BETWEEN %(sd)s AND %(ed)s"
        params["sd"] = from_date
        params["ed"] = to_date
    return wh


@frappe.whitelist()
def get_crm_analytics_breakdowns(from_date=None, to_date=None, user=None, industry=None,
                                 source=None, fiscal_year=None, period=None, month=None,
                                 limit=12, **kwargs):
    """Dimension breakdowns for the Admin Review charts.

    Returns leads / orders / order value / lost per industry, source, territory and
    product category, plus lost-reason and activity-category splits. One grouped
    query per dimension, all honouring the same filter set as the rest of the
    dashboard.
    """
    require_crm_head()
    from_date, to_date = parse_period_and_dates(period, fiscal_year, from_date, to_date,
                                                month=month or kwargs.get("month"))
    restrict_user = user or None
    ind_list = as_filter_list(industry)
    src_list = as_filter_list(source)
    try:
        limit = max(3, min(int(limit or 12), 30))
    except Exception:
        limit = 12

    def lead_group(column):
        p = {}
        wh = _analytics_lead_where(restrict_user, ind_list, src_list, from_date, to_date, p)
        rows = frappe.db.sql(
            "SELECT COALESCE(NULLIF({col}, ''), 'Not Set') label,"
            " COUNT(*) leads,"
            " SUM(CASE WHEN l.custom_lead_category='Order' THEN 1 ELSE 0 END) orders,"
            " SUM(CASE WHEN l.custom_lead_category='Order'"
            "     THEN COALESCE(l.custom_po_value, 0) ELSE 0 END) order_value,"
            " SUM(CASE WHEN l.custom_lead_category IN ('Lost Enquiry','Lost Pipeline')"
            "     THEN 1 ELSE 0 END) lost"
            " FROM `tabLead` l {wh} GROUP BY label ORDER BY leads DESC LIMIT {lim}".format(
                col=column, wh=wh, lim=limit),
            p, as_dict=1)
        out = []
        for r in rows:
            leads = int(r.get("leads") or 0)
            orders = int(r.get("orders") or 0)
            out.append({
                "label": r.get("label") or "Not Set",
                "leads": leads,
                "orders": orders,
                "lost": int(r.get("lost") or 0),
                "order_value": flt(r.get("order_value") or 0),
                "conversion": round((orders * 100.0 / leads), 1) if leads else 0.0,
            })
        return out

    # Leads by category — the share view, mirroring Contact Status Split
    p_cat = {}
    wh_cat = _analytics_lead_where(restrict_user, ind_list, src_list, from_date, to_date, p_cat)
    cat_rows = frappe.db.sql(
        "SELECT COALESCE(NULLIF(l.custom_lead_category, ''), 'Not Set') label, COUNT(*) total,"
        " SUM(CASE WHEN l.custom_lead_category = 'Order' THEN COALESCE(l.custom_po_value, 0)"
        "          ELSE COALESCE(l.custom_expected_revenue, 0) END) value"
        " FROM `tabLead` l {wh} GROUP BY label".format(wh=wh_cat), p_cat, as_dict=1)
    CAT_ORDER = ["Enquiry", "Pipeline", "Order", "Lost Enquiry", "Lost Pipeline"]
    cat_map = {r.get("label"): r for r in cat_rows}
    lead_category = []
    for name in CAT_ORDER:
        r = cat_map.pop(name, None)
        lead_category.append({"label": name,
                              "total": int((r or {}).get("total") or 0),
                              "value": flt((r or {}).get("value") or 0)})
    for name, r in cat_map.items():          # anything unexpected still shows
        lead_category.append({"label": name or "Not Set",
                              "total": int(r.get("total") or 0),
                              "value": flt(r.get("value") or 0)})

    # Lost is two distinct things and each stores its reason in its own column,
    # so they are reported separately rather than merged.
    def lost_reasons(category, column):
        p = {}
        wh = _analytics_lead_where(restrict_user, ind_list, src_list, from_date, to_date, p)
        wh += " AND l.custom_lead_category = %(cat)s"
        p["cat"] = category
        rows = frappe.db.sql(
            "SELECT COALESCE(NULLIF(TRIM({col}), ''), 'Not Specified') label, COUNT(*) total"
            " FROM `tabLead` l {wh} GROUP BY label ORDER BY total DESC LIMIT {lim}".format(
                col=column, wh=wh, lim=limit),
            p, as_dict=1)
        return [{"label": r.get("label") or "Not Specified",
                 "total": int(r.get("total") or 0)} for r in rows]

    lost_enquiry_reason = lost_reasons("Lost Enquiry", "l.custom_lost_enquiry_reason")
    lost_pipeline_reason = lost_reasons("Lost Pipeline", "l.custom_lost_pipeline_reason")
    lost_split = [
        {"label": "Lost Enquiry", "total": sum(r["total"] for r in lost_enquiry_reason)},
        {"label": "Lost Pipeline", "total": sum(r["total"] for r in lost_pipeline_reason)},
    ]

    # Overall conversion funnel. Definitions match the Lead Performance table:
    # Contact -> Lead counts contacts that reached Converted to Lead or Existing
    # Customer; Lead -> Order counts leads in the Order category.
    p_cv = {}
    wh_cv = _analytics_lead_where(restrict_user, ind_list, src_list, from_date, to_date, p_cv)
    lead_cv = frappe.db.sql(
        "SELECT COUNT(*) leads,"
        " SUM(CASE WHEN l.custom_lead_category='Order' THEN 1 ELSE 0 END) orders"
        " FROM `tabLead` l {wh}".format(wh=wh_cv), p_cv, as_dict=1)
    # Contacts CREATED in the period — the acquisition base
    p_cc = {}
    wh_cc = "WHERE docstatus < 2"
    if restrict_user:
        wh_cc += " AND assign_to = %(usr)s"
        p_cc["usr"] = restrict_user
    wh_cc += build_in_clause(ind_list, "ind", "industry", p_cc)
    wh_cc += build_in_clause(src_list, "src", "source", p_cc)
    if from_date and to_date:
        wh_cc += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
        p_cc["sd"] = from_date
        p_cc["ed"] = to_date
    cont_cv = frappe.db.sql(
        "SELECT COUNT(*) contacts FROM `tabBusiness Contacts` {wh}".format(wh=wh_cc),
        p_cc, as_dict=1)

    # Contacts CONVERTED in the period — dated by when the conversion happened, not
    # when the contact was created, so this agrees with the Lead Performance table.
    p_cvd = {}
    wh_cvd = "WHERE docstatus < 2 AND " + BC_CONVERTED_COND_SQL
    if restrict_user:
        wh_cvd += " AND assign_to = %(usr)s"
        p_cvd["usr"] = restrict_user
    wh_cvd += build_in_clause(ind_list, "ind", "industry", p_cvd)
    wh_cvd += build_in_clause(src_list, "src", "source", p_cvd)
    if from_date and to_date:
        wh_cvd += " AND ({d}) BETWEEN %(sd)s AND %(ed)s".format(d=BC_CONVERSION_DATE_SQL)
        p_cvd["sd"] = from_date
        p_cvd["ed"] = to_date
    conv_cnt = frappe.db.sql(
        "SELECT COUNT(*) converted FROM `tabBusiness Contacts` {wh}".format(wh=wh_cvd),
        p_cvd, as_dict=1)

    # Leads that reached Order in the period, first transition only — same rule as
    # the Lead -> Order column, so the chart and the table cannot disagree.
    p_ord = {}
    wh_ord = "WHERE l.docstatus < 2"
    if restrict_user:
        wh_ord += " AND l.lead_owner = %(usr)s"
        p_ord["usr"] = restrict_user
    wh_ord += build_in_clause(ind_list, "ind", "l.industry", p_ord)
    wh_ord += build_in_clause(src_list, "src", "l.source", p_ord)
    wh_ord += " AND " + LEAD_ORDER_COND_SQL
    if from_date and to_date:
        wh_ord += " AND ({d}) BETWEEN %(sd)s AND %(ed)s".format(d=LEAD_ORDER_CONV_DATE_SQL)
        p_ord["sd"] = from_date
        p_ord["ed"] = to_date
    ord_cnt = frappe.db.sql(
        "SELECT COUNT(DISTINCT l.name) orders FROM `tabLead` l {wh}".format(wh=wh_ord),
        p_ord, as_dict=1)

    _lv = (lead_cv[0] if lead_cv else {}) or {}
    _cv = (cont_cv[0] if cont_cv else {}) or {}
    contacts_total = int(_cv.get("contacts") or 0)
    contacts_converted = int((conv_cnt[0] if conv_cnt else {}).get("converted") or 0)
    leads_total = int(_lv.get("leads") or 0)
    orders_total = int((ord_cnt[0] if ord_cnt else {}).get("orders") or 0)
    conversion_overall = {
        # counts are dated by their own event: created-in-period for the bases,
        # converted-in-period for the conversions
        "contacts": contacts_total,
        "contacts_converted": contacts_converted,
        "leads": leads_total,
        "orders": orders_total,
        "bc_to_lead_rate": round(contacts_converted * 100.0 / contacts_total, 1) if contacts_total else 0.0,
        "lead_to_order_rate": round(orders_total * 100.0 / leads_total, 1) if leads_total else 0.0,
    }

    # Activity mix
    p_ac = {}
    wh_ac = "WHERE docstatus < 2"
    if restrict_user:
        wh_ac += " AND assigned_to = %(usr)s"
        p_ac["usr"] = restrict_user
    if from_date and to_date:
        wh_ac += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
        p_ac["sd"] = from_date
        p_ac["ed"] = to_date
    act_rows = frappe.db.sql(
        "SELECT COALESCE(NULLIF(category, ''), 'Uncategorised') label, COUNT(*) total,"
        " SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END) done,"
        " SUM(CASE WHEN status!='Completed' THEN 1 ELSE 0 END) pending"
        " FROM `tabEvent Activity` {wh} GROUP BY label ORDER BY total DESC LIMIT {lim}".format(
            wh=wh_ac, lim=limit),
        p_ac, as_dict=1)

    # Contacts per industry / source, so the acquisition mix is visible too
    def contact_group(column):
        p = {}
        wh = "WHERE docstatus < 2"
        if restrict_user:
            wh += " AND assign_to = %(usr)s"
            p["usr"] = restrict_user
        wh += build_in_clause(ind_list, "ind", "industry", p)
        wh += build_in_clause(src_list, "src", "source", p)
        if from_date and to_date:
            wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
            p["sd"] = from_date
            p["ed"] = to_date
        sql = (
            "SELECT COALESCE(NULLIF({col}, ''), 'Not Set') label, COUNT(*) total,"
            " SUM(CASE WHEN {conv} THEN 1 ELSE 0 END) converted"
            " FROM `tabBusiness Contacts` {wh} GROUP BY label ORDER BY total DESC LIMIT {lim}"
        ).format(col=column, conv=BC_CONVERTED_COND_SQL, wh=wh, lim=limit)
        rows = frappe.db.sql(sql, p, as_dict=1)
        return [{"label": r.get("label") or "Not Set",
                 "total": int(r.get("total") or 0),
                 "converted": int(r.get("converted") or 0)} for r in rows]

    return {
        "industry": lead_group("l.industry"),
        "source": lead_group("l.source"),
        "territory": lead_group("l.territory"),
        "product_category": lead_group("l.custom_product_category"),
        "lead_category": lead_category,
        "lost_split": lost_split,
        "lost_enquiry_reason": lost_enquiry_reason,
        "lost_pipeline_reason": lost_pipeline_reason,
        "conversion_overall": conversion_overall,
        "activity_category": [{"label": r.get("label") or "Uncategorised",
                               "total": int(r.get("total") or 0),
                               "done": int(r.get("done") or 0),
                               "pending": int(r.get("pending") or 0)} for r in act_rows],
        "contact_industry": contact_group("industry"),
        "contact_source": contact_group("source"),
    }


@frappe.whitelist()
def get_crm_team_comparison(from_date=None, to_date=None, user=None, industry=None, source=None,
                            fiscal_year=None, period=None, month=None, **kwargs):
    """Per-user aggregates for the admin comparison chart.

    One grouped query per entity rather than N per-user dashboard calls, so the
    chart stays cheap as the team grows.
    """
    from_date, to_date = parse_period_and_dates(period, fiscal_year, from_date, to_date,
                                                month=month or kwargs.get("month"))

    require_crm_head()
    restrict_user = user or None

    ind_list = as_filter_list(industry)
    src_list = as_filter_list(source)

    # Leads per owner
    lp = {}
    wl = "WHERE l.docstatus < 2"
    if restrict_user:
        wl += " AND l.lead_owner = %(usr)s"
        lp["usr"] = restrict_user
    wl += build_in_clause(ind_list, "ind", "l.industry", lp)
    wl += build_in_clause(src_list, "src", "l.source", lp)
    if from_date and to_date:
        wl += " AND DATE(l.creation) BETWEEN %(sd)s AND %(ed)s"
        lp["sd"] = from_date
        lp["ed"] = to_date
    leads = frappe.db.sql(
        "SELECT l.lead_owner uid, COUNT(*) total,"
        " SUM(CASE WHEN l.custom_lead_category='Order' THEN 1 ELSE 0 END) orders,"
        " SUM(CASE WHEN l.custom_lead_category='Order'"
        "     THEN COALESCE(l.custom_po_value, 0) ELSE 0 END) order_value"
        " FROM `tabLead` l " + wl + " GROUP BY l.lead_owner", lp, as_dict=1)

    # Activities per owner
    ap = {}
    wa = "WHERE docstatus < 2"
    if restrict_user:
        wa += " AND assigned_to = %(usr)s"
        ap["usr"] = restrict_user
    if from_date and to_date:
        wa += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
        ap["sd"] = from_date
        ap["ed"] = to_date
    acts = frappe.db.sql(
        "SELECT assigned_to uid, COUNT(*) total,"
        " SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END) done"
        " FROM `tabEvent Activity` " + wa + " GROUP BY assigned_to", ap, as_dict=1)

    # Contacts per owner
    cp = {}
    wc = "WHERE docstatus < 2"
    if restrict_user:
        wc += " AND assign_to = %(usr)s"
        cp["usr"] = restrict_user
    wc += build_in_clause(ind_list, "ind", "industry", cp)
    wc += build_in_clause(src_list, "src", "source", cp)
    if from_date and to_date:
        wc += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
        cp["sd"] = from_date
        cp["ed"] = to_date
    conts = frappe.db.sql(
        "SELECT assign_to uid, COUNT(*) total,"
        " SUM(CASE WHEN " + BC_CONVERTED_COND_SQL + " THEN 1 ELSE 0 END) converted"
        " FROM `tabBusiness Contacts` " + wc + " GROUP BY assign_to", cp, as_dict=1)

    rows = {}

    def slot(uid):
        if not uid:
            return None
        return rows.setdefault(uid, {
            "user": uid, "full_name": uid, "leads": 0, "orders": 0,
            "order_value": 0.0, "activities": 0, "activities_done": 0,
            "contacts": 0, "contacts_converted": 0,
        })

    for r in leads:
        s = slot(r.get("uid"))
        if s:
            s["leads"] = int(r.get("total") or 0)
            s["orders"] = int(r.get("orders") or 0)
            s["order_value"] = flt(r.get("order_value") or 0)
    for r in acts:
        s = slot(r.get("uid"))
        if s:
            s["activities"] = int(r.get("total") or 0)
            s["activities_done"] = int(r.get("done") or 0)
    for r in conts:
        s = slot(r.get("uid"))
        if s:
            s["contacts"] = int(r.get("total") or 0)
            s["contacts_converted"] = int(r.get("converted") or 0)

    names = dict(frappe.db.sql(
        "SELECT name, COALESCE(NULLIF(full_name, ''), name) FROM `tabUser`") or [])
    out = []
    for uid, v in rows.items():
        v["full_name"] = names.get(uid) or uid
        out.append(v)
    out.sort(key=lambda x: (-x["leads"], x["full_name"]))
    return {"rows": out}


@frappe.whitelist()
def get_crm_report_emails(report_type="daily"):
    """Return email IDs stored in Admin Settings for the given report type."""
    field_map = {
        "daily":   "crm_daily_report_emails",
        "weekly":  "crm_weekly_report_emails",
        "monthly": "crm_monthly_report_emails",
    }
    field = field_map.get(report_type, "crm_daily_report_emails")
    try:
        val = frappe.db.get_single_value("Admin Settings", field) or ""
    except Exception:
        val = ""
    emails = [e.strip() for e in val.split(",") if e.strip()]
    return {"emails": emails}


@frappe.whitelist()
def export_crm_dashboard_excel(
    c_user=None, c_fiscal_year=None, c_industry=None, c_source=None, c_period_type="monthly",
    c_from_date=None, c_to_date=None, c_month=None, c_period=None,
    l_user=None, l_fiscal_year=None, l_industry=None, l_source=None, l_period_type="monthly",
    l_from_date=None, l_to_date=None, l_month=None, l_period=None,
    t_fiscal_year=None,
    ea_user=None, ea_fiscal_year=None, ea_period_type="monthly", ea_entity="All",
    ea_from_date=None, ea_to_date=None, ea_month=None, ea_period=None,
    export_type="applied", send_email=0, email_address=None
):
    """Export CRM Analytics: Summary + per-owner tabs + detail tabs. Optionally email."""
    import base64
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from frappe.utils import flt

    # ── Access guard ──────────────────────────────────────────────
    if frappe.session.user != "Administrator" and not frappe.flags.ignore_permissions:
        if "DAC CRM Head" not in frappe.get_roles(frappe.session.user):
            frappe.throw("Only DAC CRM Head can export the CRM report.", frappe.PermissionError)

    # ── Period override: compute concrete date range for daily/weekly/monthly ─
    import datetime as _dt
    _today = frappe.utils.getdate(frappe.utils.today())

    summary_period_type = c_period_type
    period_label_str = export_type.title()
    if export_type == "daily":
        summary_period_type = "daily"
        c_period_type = l_period_type = ea_period_type = "daily"
        _d_str = _today.strftime("%Y-%m-%d")
        # Override: only today's records
        c_from_date = l_from_date = ea_from_date = _d_str
        c_to_date   = l_to_date   = ea_to_date   = _d_str
        c_period    = l_period    = ea_period     = None  # use explicit from/to
        period_label_str = "Daily ({})".format(_d_str)
    elif export_type == "weekly":
        summary_period_type = "weekly"
        c_period_type = l_period_type = ea_period_type = "weekly"
        _wday = _today.weekday()  # Monday=0
        _week_start = _today - _dt.timedelta(days=_wday)
        _week_end   = _week_start + _dt.timedelta(days=6)
        c_from_date = l_from_date = ea_from_date = _week_start.strftime("%Y-%m-%d")
        c_to_date   = l_to_date   = ea_to_date   = _week_end.strftime("%Y-%m-%d")
        c_period    = l_period    = ea_period     = None
        period_label_str = "Weekly ({} to {})".format(
            _week_start.strftime("%d-%b-%Y"), _week_end.strftime("%d-%b-%Y"))
    elif export_type == "monthly":
        summary_period_type = "monthly"
        c_period_type = l_period_type = ea_period_type = "monthly"
        _month_start = _today.replace(day=1)
        import calendar as _cal
        _last_day = _cal.monthrange(_today.year, _today.month)[1]
        _month_end = _today.replace(day=_last_day)
        c_from_date = l_from_date = ea_from_date = _month_start.strftime("%Y-%m-%d")
        c_to_date   = l_to_date   = ea_to_date   = _month_end.strftime("%Y-%m-%d")
        c_period    = l_period    = ea_period     = None
        period_label_str = "Monthly ({} {})".format(
            _today.strftime("%B"), _today.year)

    # ── Fiscal year ────────────────────────────────────────────────
    fiscal_year = c_fiscal_year or l_fiscal_year or t_fiscal_year or ea_fiscal_year
    if not fiscal_year:
        fy_doc = frappe.db.get_value("Fiscal Year", {"disabled": 0}, "name",
                                     order_by="year_start_date desc")
        fiscal_year = fy_doc or "2025-2026"

    c_from, c_to = parse_period_and_dates(c_period, c_fiscal_year or fiscal_year,
                                          c_from_date, c_to_date, month=c_month)
    l_from, l_to = parse_period_and_dates(l_period, l_fiscal_year or fiscal_year,
                                          l_from_date, l_to_date, month=l_month)
    ea_from, ea_to = parse_period_and_dates(ea_period, ea_fiscal_year or fiscal_year,
                                            ea_from_date, ea_to_date, month=ea_month)

    # ── Effective users (head sees all by default) ─────────────────
    eff_c_user  = c_user  or None
    eff_l_user  = l_user  or None
    eff_ea_user = ea_user or None

    # ── Multiselect scope filters (Contacts and Leads are filtered independently) ──
    c_ind_list = as_filter_list(c_industry)
    c_src_list = as_filter_list(c_source)
    l_ind_list = as_filter_list(l_industry)
    l_src_list = as_filter_list(l_source)

    def scope_label(ind_list, src_list):
        parts = []
        parts.append(", ".join(ind_list) if ind_list else "All Industries")
        parts.append(", ".join(src_list) if src_list else "All Sources")
        return "  |  ".join(parts)

    # ── Style helpers ─────────────────────────────────────────────
    SLATE  = "0F172A"; BLUE   = "2563EB"; GREEN  = "16A34A"
    ORANGE = "EA580C"; PURPLE = "7C3AED"; WHITE  = "FFFFFF"
    SILVER = "F1F5F9"; ALT    = "EFF6FF"

    def _font(bold=False, size=9, color=SLATE, name="Calibri"):
        return Font(name=name, size=size, bold=bold, color=color)
    def _fill(h):
        return PatternFill(start_color=h, end_color=h, fill_type="solid")
    def _border(c="CBD5E1"):
        s = Side(style="thin", color=c)
        return Border(left=s, right=s, top=s, bottom=s)
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    TITLE_FONT  = _font(bold=True, size=14, color=WHITE)
    SEC_FONT    = _font(bold=True, size=11, color=BLUE)
    HDR_FONT    = _font(bold=True, size=9,  color=WHITE)
    DATA_FONT   = _font(size=9)
    BOLD_FONT   = _font(bold=True, size=9)
    LINK_FONT   = Font(name="Calibri", size=9, color=BLUE, underline="single")
    TOT_FONT    = _font(bold=True, size=9)
    LBL_FONT    = _font(bold=True, size=9, color="475569")
    VAL_FONT    = _font(size=9, color=SLATE)
    THIN        = _border()
    SILVER_FILL = _fill(SILVER)
    ALT_FILL    = _fill(ALT)
    FILTER_FILL = _fill("F8FAFC")

    def apply_row(ws, row, cs, ce, font=None, fill=None):
        for c in range(cs, ce + 1):
            cell = ws.cell(row=row, column=c)
            if font: cell.font = font
            if fill: cell.fill = fill
            cell.border = THIN

    def auto_width(ws, mn=10, mx=42):
        for col in ws.columns:
            w = max((len(str(cl.value or "")) for cl in col), default=mn)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, mn), mx)

    def title_banner(ws, text, span, fill_color, row=1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        c = ws.cell(row=row, column=1, value=text)
        c.font = TITLE_FONT; c.fill = _fill(fill_color)
        c.alignment = _align("left", "center")
        ws.row_dimensions[row].height = 32

    def money(v):
        return "Rs.{:,.0f}".format(flt(v))

    # user id -> full name. User ids are login/email addresses; the sheets should
    # never show those where a person's name is meant. Defined up front because the
    # team list below needs it too.
    _user_names = dict(frappe.db.sql(
        "SELECT name, COALESCE(NULLIF(full_name, ''), name) FROM `tabUser`") or [])

    def user_name(uid):
        if not uid:
            return "-"
        return _user_names.get(uid) or uid

    # ── Fetch active sales persons ────────────────────────────────
    # Pull all CRM users (not just Sales Persons) for a complete team view
    crm_users = frappe.db.sql("""
        SELECT u.name as user_id, u.full_name as sp_name
        FROM `tabUser` u
        JOIN `tabHas Role` hr ON hr.parent = u.name
        WHERE u.enabled = 1 AND u.user_type = 'System User'
          AND hr.role IN ('DAC CRM Head', 'DAC CRM')
        GROUP BY u.name
        ORDER BY u.full_name ASC
    """, as_dict=True)
    # Also try Sales Person linkage for backward compat
    sales_persons = frappe.get_all("Sales Person",
        fields=["name", "sales_person_name", "employee"],
        filters={"enabled": 1, "is_group": 0})
    sp_user_ids = set()
    sp_users = []
    for sp in sales_persons:
        uid = frappe.db.get_value("Employee", sp.employee, "user_id") if sp.employee else None
        if uid:
            sp_users.append({"sp_name": sp.sales_person_name, "user_id": uid})
            sp_user_ids.add(uid)
    # Add CRM users not already in sp_users
    for cu in crm_users:
        if cu["user_id"] not in sp_user_ids:
            sp_users.append({"sp_name": cu["sp_name"], "user_id": cu["user_id"]})
            sp_user_ids.add(cu["user_id"])

    # Anyone who actually owns a Quotation, Lead or Contact but holds no CRM role
    # and no Sales Person record. Without this the per-user sheets can never add up
    # to the real totals — their records were simply dropped, which is what made the
    # summary disagree with the document count.
    extra_owners = frappe.db.sql("""
        SELECT DISTINCT u FROM (
            SELECT COALESCE(NULLIF(custom_lead_owner, ''), owner) AS u
              FROM `tabQuotation` WHERE docstatus < 2
            UNION SELECT lead_owner FROM `tabLead` WHERE docstatus < 2
            UNION SELECT assign_to  FROM `tabBusiness Contacts` WHERE docstatus < 2
        ) x
        WHERE u IS NOT NULL AND u != ''
    """)
    for (uid,) in (extra_owners or []):
        if uid not in sp_user_ids:
            sp_users.append({"sp_name": user_name(uid), "user_id": uid})
            sp_user_ids.add(uid)

    sp_users = sorted(sp_users, key=lambda x: x["sp_name"] or "")

    # When the workbook is built for ONE person (the individual report mailed to a
    # team member), the team-comparison sheets must show that person only. The
    # record sheets were already filtered by c_user/l_user/ea_user, but the
    # comparison sheets loop over sp_users — so without this the attachment leaked
    # every colleague's numbers into a personal report.
    _only_user = eff_l_user or eff_c_user or eff_ea_user
    if _only_user:
        sp_users = [u for u in sp_users if u["user_id"] == _only_user] or [
            {"sp_name": user_name(_only_user), "user_id": _only_user}
        ]

    # Fetch all distinct categories dynamically that have activities in the period
    cat_query = "SELECT DISTINCT category FROM `tabEvent Activity` WHERE category IS NOT NULL AND category != '' AND docstatus < 2"
    cat_params = {}
    if ea_from and ea_to:
        cat_query += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
        cat_params["sd"] = ea_from
        cat_params["ed"] = ea_to
    db_cats = frappe.db.sql(cat_query, cat_params)
    dynamic_categories = sorted([c[0] for c in db_cats if c[0]])

    def agg_user(uid, from_d=None, to_d=None, use_period=None, all_time=False):
        """Aggregate contact, lead, and activity counts for a user.

        Contacts and Leads are queried separately: each dashboard section carries its
        own date range and its own industry/source multiselect, so a single combined
        call would apply the Lead filters to the Contact figures.

        Pass all_time=True for the unfiltered overview tab — passing from_d/to_d as
        None means "use the section default", not "no date filter".
        """
        if all_time:
            c_fd = c_td = l_fd = l_td = None
            fy_arg = period_arg = c_month_arg = l_month_arg = None
        else:
            c_fd = from_d if from_d is not None else c_from
            c_td = to_d   if to_d   is not None else c_to
            l_fd = from_d if from_d is not None else l_from
            l_td = to_d   if to_d   is not None else l_to
            fy_arg = fiscal_year
            period_arg = use_period
            c_month_arg, l_month_arg = c_month, l_month

        cd = get_tabular_dashboard_data(
            from_date=c_fd, to_date=c_td, user=uid,
            industry=c_ind_list, source=c_src_list,
            period_type=summary_period_type, fiscal_year=fy_arg,
            period=period_arg, month=c_month_arg)
        ld = get_tabular_dashboard_data(
            from_date=l_fd, to_date=l_td, user=uid,
            industry=l_ind_list, source=l_src_list,
            period_type=summary_period_type, fiscal_year=fy_arg,
            period=period_arg, month=l_month_arg)

        # Activity aggregation below uses the lead window, matching the summary tab
        fd, td = l_fd, l_td

        d = dict(o=0,c=0,e=0,enq_c=0,enq_v=0.0,pipe_c=0,pipe_v=0.0,
                 ord_c=0,ord_v=0.0,lenq_c=0,lenq_v=0.0,lpipe_c=0,lpipe_v=0.0,
                 c_to_l=0,l_to_o=0,conv=0,
                 act_total=0,act_done=0,act_open=0)
        # Contact totals
        ct = cd.get("contact_totals", {})
        d["o"] = ct.get("o", 0)
        d["c"] = ct.get("c", 0)
        d["e"] = ct.get("e", 0)
        # Lead totals
        lt = ld.get("lead_totals", {})
        d["enq_c"]  = lt.get("enq_c", 0);   d["enq_v"]  = flt(lt.get("enq_v", 0))
        d["pipe_c"] = lt.get("pipe_c", 0);   d["pipe_v"] = flt(lt.get("pipe_v", 0))
        d["ord_c"]  = lt.get("ord_c", 0);    d["ord_v"]  = flt(lt.get("ord_v", 0))
        d["lenq_c"] = lt.get("lenq_c", 0);   d["lenq_v"] = flt(lt.get("lenq_v", 0))
        d["lpipe_c"]= lt.get("lpipe_c", 0);  d["lpipe_v"]= flt(lt.get("lpipe_v", 0))
        d["c_to_l"] = lt.get("conv_bc_to_lead", 0)
        d["l_to_o"] = lt.get("conv_lead_to_order", 0)
        # Activity counts for this user within date range
        ea_p = {"usr": uid}
        ea_wh = "WHERE docstatus < 2 AND assigned_to = %(usr)s"
        if fd and td:
            ea_wh += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
            ea_p["sd"] = fd; ea_p["ed"] = td
        ea_agg = frappe.db.sql(
            f"SELECT COUNT(*) total, "
            f"SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END) done, "
            f"SUM(CASE WHEN status!='Completed' THEN 1 ELSE 0 END) pending "
            f"FROM `tabEvent Activity` {ea_wh}",
            ea_p, as_dict=1)
        if ea_agg:
            d["act_total"] = int(ea_agg[0].get("total") or 0)
            d["act_done"]  = int(ea_agg[0].get("done") or 0)
            d["act_open"]  = int(ea_agg[0].get("pending") or 0)

        # ── ADDITIONAL PERFORMANCE METRICS ──
        # Leads created in the period:
        l_p_c = {"usr": uid}
        l_p_wh = "WHERE docstatus < 2 AND lead_owner = %(usr)s"
        if l_fd and l_td:
            l_p_wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
            l_p_c["sd"] = l_fd; l_p_c["ed"] = l_td
        d["leads_created"] = frappe.db.sql(f"SELECT COUNT(*) FROM `tabLead` {l_p_wh}", l_p_c)[0][0] or 0

        # Contacts created in the period:
        c_p_c = {"usr": uid}
        c_p_wh = "WHERE docstatus < 2 AND assign_to = %(usr)s"
        if c_fd and c_td:
            c_p_wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
            c_p_c["sd"] = c_fd; c_p_c["ed"] = c_td
        d["contacts_created"] = frappe.db.sql(f"SELECT COUNT(*) FROM `tabBusiness Contacts` {c_p_wh}", c_p_c)[0][0] or 0

        # Quotations created in the period:
        q_p_c = {"usr": uid}
        q_p_wh = "WHERE docstatus < 2 AND COALESCE(NULLIF(custom_lead_owner, ''), owner) = %(usr)s"
        if l_fd and l_td:
            q_p_wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
            q_p_c["sd"] = l_fd; q_p_c["ed"] = l_td
        d["quotations_created"] = frappe.db.sql(f"SELECT COUNT(*) FROM `tabQuotation` {q_p_wh}", q_p_c)[0][0] or 0

        # One count per status, from a single grouped query, so the breakdown is
        # guaranteed to sum to quotations_created above.
        for _st in QUOTATION_STATUSES:
            d[_q_status_key(_st)] = 0
        for _row in frappe.db.sql(
                f"SELECT status, COUNT(*) FROM `tabQuotation` {q_p_wh} GROUP BY status", q_p_c) or []:
            _key = _q_status_key(_row[0] or "")
            if _key in d:
                d[_key] = _row[1] or 0
            else:
                # An unexpected status still has to be counted somewhere.
                d.setdefault("q_st_other", 0)
                d["q_st_other"] += _row[1] or 0
        # Kept for callers that still read the old names.
        d["q_created_draft"] = d.get("q_st_draft", 0)
        d["q_created_open"] = d.get("q_st_open", 0)
        d["q_created_ordered"] = d.get("q_st_ordered", 0)

        # Event activities performed (Completed) breakdown of contact and lead:
        # Lead activities performed:
        al_p = {"usr": uid}
        al_wh = "WHERE docstatus < 2 AND assigned_to = %(usr)s AND status = 'Completed' AND reference_type = 'Lead'"
        if l_fd and l_td:
            al_wh += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
            al_p["sd"] = l_fd; al_p["ed"] = l_td
        d["act_lead"] = frappe.db.sql(f"SELECT COUNT(*) FROM `tabEvent Activity` {al_wh}", al_p)[0][0] or 0

        # Contact activities performed:
        ac_p = {"usr": uid}
        ac_wh = "WHERE docstatus < 2 AND assigned_to = %(usr)s AND status = 'Completed' AND reference_type = 'Business Contacts'"
        if c_fd and c_td:
            ac_wh += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
            ac_p["sd"] = c_fd; ac_p["ed"] = c_td
        d["act_contact"] = frappe.db.sql(f"SELECT COUNT(*) FROM `tabEvent Activity` {ac_wh}", ac_p)[0][0] or 0

        # Activities created in the period:
        ea_c_p = {"usr": uid}
        ea_c_wh = "WHERE docstatus < 2 AND owner = %(usr)s"
        if fd and td:
            ea_c_wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
            ea_c_p["sd"] = fd; ea_c_p["ed"] = td
        # "Activities created" is deliberately not reported — only completed work is.

        # Categories breakdown query dynamically:
        act_cat_vals = {}
        for cat in dynamic_categories:
            act_cat_vals[f"act_{cat}_lead"] = 0
            act_cat_vals[f"act_{cat}_contact"] = 0
            
        act_p_all = {"usr": uid}
        act_wh_all = "WHERE docstatus < 2 AND assigned_to = %(usr)s AND status = 'Completed'"
        if fd and td:
            act_wh_all += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
            act_p_all["sd"] = fd; act_p_all["ed"] = td
            
        activities_done = frappe.db.sql(
            "SELECT category, reference_type, COUNT(*) as cnt "
            f"FROM `tabEvent Activity` {act_wh_all} "
            "GROUP BY category, reference_type", act_p_all, as_dict=True)
            
        for a in activities_done:
            cat = a.get("category")
            ref = a.get("reference_type")
            cnt = a.get("cnt") or 0
            ref_key = "lead" if ref == "Lead" else "contact"
            
            if cat in dynamic_categories:
                act_cat_vals[f"act_{cat}_{ref_key}"] = cnt

        d.update(act_cat_vals)

        # User-wise all-time draft and open quotations:
        qo_p = {"usr": uid}
        d["draft_quotations"] = frappe.db.sql("SELECT COUNT(*) FROM `tabQuotation` WHERE docstatus = 0 AND COALESCE(NULLIF(custom_lead_owner, ''), owner) = %(usr)s", qo_p)[0][0] or 0
        d["open_quotations"] = frappe.db.sql("SELECT COUNT(*) FROM `tabQuotation` WHERE docstatus = 1 AND status = 'Open' AND COALESCE(NULLIF(custom_lead_owner, ''), owner) = %(usr)s", qo_p)[0][0] or 0

        return d

    wb = openpyxl.Workbook()

    # ==========================================================
    # TAB 1 -- SUMMARY
    # ==========================================================
    ws = wb.active
    ws.title = "Summary Overview"
    title_banner(ws,
        "DAC CRM Executive Dashboard Report  |  {}  |  {} Report".format(
            frappe.utils.today(), period_label_str),
        15, SLATE)

    r = 3
    ws.cell(r, 1, "Report Filter Details").font = SEC_FONT; r += 1
    for lbl, val in [
        ("Fiscal Year",           fiscal_year),
        ("Report Period Mode",    summary_period_type.title()),
        ("Contacts Date Range",   "{} to {}".format(c_from or "All Time", c_to or "All Time")),
        ("Leads Date Range",      "{} to {}".format(l_from or "All Time", l_to or "All Time")),
        ("Activities Date Range", "{} to {}".format(ea_from or "All Time", ea_to or "All Time")),
        ("Contacts Filter",       scope_label(c_ind_list, c_src_list)),
        ("Leads Filter",          scope_label(l_ind_list, l_src_list)),
        ("Generated By",          frappe.session.user_fullname or frappe.session.user),
        ("Generated On",          frappe.utils.today()),
    ]:
        lc = ws.cell(r, 1, lbl); lc.font = LBL_FONT; lc.fill = FILTER_FILL; lc.border = THIN
        vc = ws.cell(r, 2, val); vc.font = VAL_FONT;  vc.fill = FILTER_FILL; vc.border = THIN
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1
    r += 1

    # Overall Totals
    ws.cell(r, 1, "Overall Totals (All Team)").font = SEC_FONT; r += 1
    for ci, h in enumerate(["Metric", "Count / Value"], 1):
        c = ws.cell(r, ci, h); c.font = HDR_FONT; c.fill = _fill(BLUE); c.border = THIN
    r += 1

    p_ct = {}; p_ct_wh_extras = []
    wh_ct = "WHERE docstatus < 2"
    if eff_c_user: wh_ct += " AND assign_to = %(cu)s"; p_ct["cu"] = eff_c_user
    wh_ct += build_in_clause(c_ind_list, "cind", "industry", p_ct)
    wh_ct += build_in_clause(c_src_list, "csrc", "source", p_ct)
    if c_from and c_to:
        wh_ct += " AND DATE(creation) BETWEEN %(csd)s AND %(ced)s"
        p_ct["csd"] = c_from; p_ct["ced"] = c_to
    c_tot = frappe.db.sql(
        "SELECT SUM(CASE WHEN status='Open' THEN 1 ELSE 0 END) oc,"
        "SUM(CASE WHEN status='Converted to Lead' THEN 1 ELSE 0 END) cc,"
        "SUM(CASE WHEN status='Existing Customer' THEN 1 ELSE 0 END) ec,"
        "COUNT(*) tc FROM `tabBusiness Contacts` " + wh_ct,
        p_ct, as_dict=1)[0] or {}

    p_lt = {}; wh_lt = "WHERE docstatus < 2"
    if eff_l_user: wh_lt += " AND lead_owner = %(lu)s"; p_lt["lu"] = eff_l_user
    wh_lt += build_in_clause(l_ind_list, "lind", "industry", p_lt)
    wh_lt += build_in_clause(l_src_list, "lsrc", "source", p_lt)
    if l_from and l_to:
        wh_lt += " AND DATE(creation) BETWEEN %(lsd)s AND %(led)s"
        p_lt["lsd"] = l_from; p_lt["led"] = l_to
    l_tot = frappe.db.sql(
        "SELECT SUM(CASE WHEN custom_lead_category='Enquiry'  THEN 1 ELSE 0 END) enq_c,"
        "SUM(CASE WHEN custom_lead_category='Enquiry' THEN COALESCE(custom_expected_revenue,0) ELSE 0 END) enq_v,"
        "SUM(CASE WHEN custom_lead_category='Pipeline' THEN 1 ELSE 0 END) pipe_c,"
        "SUM(CASE WHEN custom_lead_category='Pipeline' THEN COALESCE(custom_expected_revenue,0) ELSE 0 END) pipe_v,"
        "SUM(CASE WHEN custom_lead_category='Order'    THEN 1 ELSE 0 END) ord_c,"
        "SUM(CASE WHEN custom_lead_category='Order' THEN COALESCE(custom_po_value,0) ELSE 0 END) ord_v,"
        "SUM(CASE WHEN custom_lead_category IN ('Lost Enquiry','Lost Pipeline') THEN 1 ELSE 0 END) lost_c,"
        "COUNT(*) tl FROM `tabLead` " + wh_lt,
        p_lt, as_dict=1)[0] or {}

    p_ea = {}; wh_ea = "WHERE docstatus<2"
    if eff_ea_user: wh_ea += " AND assigned_to = %(eu)s"; p_ea["eu"] = eff_ea_user
    if ea_from and ea_to:
        wh_ea += " AND DATE(COALESCE(ends_on,starts_on)) BETWEEN %(esd)s AND %(eed)s"
        p_ea["esd"] = ea_from; p_ea["eed"] = ea_to
    ea_tot = frappe.db.sql(
        "SELECT COUNT(*) total,SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END) done "
        "FROM `tabEvent Activity` " + wh_ea, p_ea, as_dict=1)[0] or {}

    for i, (lbl, val) in enumerate([
        ("Total Contacts",       c_tot.get("tc", 0)),
        ("  Open",               c_tot.get("oc", 0)),
        ("  Converted to Lead",  c_tot.get("cc", 0)),
        ("  Existing Customers", c_tot.get("ec", 0)),
        ("Total Leads",          l_tot.get("tl", 0)),
        ("  Enquiry",            "{} ({})".format(l_tot.get("enq_c",0), money(l_tot.get("enq_v",0)))),
        ("  Pipeline",           "{} ({})".format(l_tot.get("pipe_c",0), money(l_tot.get("pipe_v",0)))),
        ("  Order",              "{} ({})".format(l_tot.get("ord_c",0), money(l_tot.get("ord_v",0)))),
        ("  Lost",               l_tot.get("lost_c", 0)),
        ("Total Activities",     ea_tot.get("total", 0)),
        ("  Completed",          ea_tot.get("done", 0)),
        ("Conversion Rates (%)", ""),
        ("  Contact to Lead",    "{}%".format(round(
            (flt(c_tot.get("cc", 0)) / flt(c_tot.get("tc", 0)) * 100) if flt(c_tot.get("tc", 0)) else 0, 1))),
        ("  Lead to Order",      "{}%".format(round(
            (flt(l_tot.get("ord_c", 0)) / flt(l_tot.get("tl", 0)) * 100) if flt(l_tot.get("tl", 0)) else 0, 1))),
    ]):
        fill = ALT_FILL if i % 2 == 0 else None
        lc = ws.cell(r, 1, lbl); lc.font = BOLD_FONT if not lbl.startswith("  ") else DATA_FONT
        vc = ws.cell(r, 2, val); vc.font = DATA_FONT
        apply_row(ws, r, 1, 2, fill=fill); r += 1
    r += 1

    # Team Member Wise Comparison (Simple - only high-level totals)
    ws.cell(r, 1, "Team Member Performance Summary  –  {} Report".format(period_label_str)).font = SEC_FONT; r += 1
    simple_hdrs = ["Team Member", "Leads Created", "Contacts Created", "Quotations Created",
                   "Activities Completed", "BC to Lead (%)", "Lead to Order (%)"]
    for ci, h in enumerate(simple_hdrs, 1):
        c = ws.cell(r, ci, h); c.font = HDR_FONT; c.fill = _fill(BLUE)
        c.border = THIN; c.alignment = _align("center", "center", wrap=True)
    ws.row_dimensions[r].height = 28; r += 1
    
    # c_to_l/l_to_o are percentages, so the TOTAL row shows their average, not a sum.
    simple_totals = {k: 0 for k in ["leads", "contacts", "quotations", "act_completed",
                                    "c_to_l_sum", "l_to_o_sum", "rows"]}
    row_idx = 0
    for sp in sp_users:
        d = agg_user(sp["user_id"])
        if (d["leads_created"] == 0 and 
            d["contacts_created"] == 0 and 
            d["quotations_created"] == 0 and
            d["act_done"] == 0):
            continue
        fill = ALT_FILL if row_idx % 2 == 0 else None
        row_vals = [
            sp["sp_name"], d["leads_created"], d["contacts_created"], d["quotations_created"],
            d["act_done"], d["c_to_l"], d["l_to_o"]
        ]
        for ci, v in enumerate(row_vals, 1):
            ws.cell(r, ci, v).font = DATA_FONT
        apply_row(ws, r, 1, len(simple_hdrs), fill=fill); r += 1
        simple_totals["leads"] += d["leads_created"]
        simple_totals["contacts"] += d["contacts_created"]
        simple_totals["quotations"] += d["quotations_created"]
        simple_totals["act_completed"] += d["act_done"]
        simple_totals["c_to_l_sum"] += flt(d.get("c_to_l", 0))
        simple_totals["l_to_o_sum"] += flt(d.get("l_to_o", 0))
        simple_totals["rows"] += 1
        row_idx += 1
        
    tot_row_vals_simple = [
        "TOTAL", simple_totals["leads"], simple_totals["contacts"], simple_totals["quotations"],
        simple_totals["act_completed"],
        round(simple_totals["c_to_l_sum"] / simple_totals["rows"], 1) if simple_totals["rows"] else 0,
        round(simple_totals["l_to_o_sum"] / simple_totals["rows"], 1) if simple_totals["rows"] else 0
    ]
    for ci, v in enumerate(tot_row_vals_simple, 1):
        ws.cell(r, ci, v).font = TOT_FONT
    apply_row(ws, r, 1, len(simple_hdrs), font=TOT_FONT, fill=SILVER_FILL); r += 1

    # Create a dedicated sheet for Team Performance comparison table with all the splits
    ws_tp = wb.create_sheet(title="Detailed Period Splits")
    title_banner(ws_tp,
        "DAC CRM  |  Detailed Period Performance Splits  –  {} Report".format(period_label_str),
        23, "2563EB")
    r_tp = 3
    
    # "Quotations Created" is the total raised in the period; the columns after it
    # split that same total by status, so Total = sum of the status columns. The old
    # Draft/Open/Ordered trio left Expired and Lost out, which made the total look
    # wrong. Activities Created is gone — only Completed is reported.
    tm_hdrs = ["Team Member", "Leads Created", "Contacts Created",
               "Quotations Created (Total)"]
    tm_hdrs += ["→ {}".format(st) for st in QUOTATION_STATUSES]
    tm_hdrs += ["Activities Completed"]
    for cat in dynamic_categories:
        tm_hdrs.append(f"{cat} (Lead)")
        tm_hdrs.append(f"{cat} (Contact)")
    tm_hdrs.extend([
        "Contacts (Open)", "Contacts (Conv)", "Contacts (Cust)",
        "Enquiry #", "Enquiry Val", "Pipeline #", "Pipeline Val",
        "Order #", "Order Val", "Lost #", "BC to Lead", "Lead to Order",
        "Activity (Total)", "Activity (Done)", "Activity (Open)"
    ])

    groups = [
        ("Sales Person", 1, "1E3A8A"),
        ("Creation Counts (This Period)", 2, "2563EB"),
        ("Quotations Created (This Period)  –  Total = sum of statuses",
         1 + len(QUOTATION_STATUSES), "0284C7"),
        ("Activities Completed (This Period)", 1, "0D9488"),
    ]
    if dynamic_categories:
        groups.append(("Completed Activities by Category", len(dynamic_categories) * 2, "059669"))
    groups.extend([
        # Pending Quotations is an ALL-TIME figure and does not belong on a sheet of
        # period splits — it is reported on its own in "Overall Summary (All-Time)".
        ("Contact Status", 3, "475569"),
        ("Lead Pipelines", 7, "7C3AED"),
        ("Conversion Rates (%)", 2, "DB2777"),
        ("Activity Status", 3, "374151")
    ])

    # Write Group Headers
    c_idx = 1
    for g_name, num_cols, color in groups:
        cell = ws_tp.cell(r_tp, c_idx, g_name)
        cell.font = openpyxl.styles.Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = _fill(color)
        cell.alignment = _align("center", "center", wrap=True)
        cell.border = THIN
        if num_cols > 1:
            ws_tp.merge_cells(start_row=r_tp, start_column=c_idx, end_row=r_tp, end_column=c_idx + num_cols - 1)
            for col in range(c_idx, c_idx + num_cols):
                ws_tp.cell(r_tp, col).border = THIN
                ws_tp.cell(r_tp, col).fill = _fill(color)
        c_idx += num_cols
    ws_tp.row_dimensions[r_tp].height = 24; r_tp += 1

    # Write Sub-Headers
    for ci, h in enumerate(tm_hdrs, 1):
        cell = ws_tp.cell(r_tp, ci, h)
        cell.font = openpyxl.styles.Font(name="Segoe UI", size=9, bold=True, color="333333")
        cell.fill = _fill("F1F5F9")
        cell.border = THIN
        cell.alignment = _align("center", "center", wrap=True)
    ws_tp.row_dimensions[r_tp].height = 28; r_tp += 1
    team_totals = {k: 0 for k in ["o", "c", "e", "enq_c", "enq_v", "pipe_c", "pipe_v",
                                  "ord_c", "ord_v", "lenq_c", "lpipe_c",
                                  "c_to_l", "l_to_o", "act_total", "act_done", "act_open",
                                  "leads_created", "contacts_created", "quotations_created",
                                  "draft_quotations", "open_quotations"]}
    for _st in QUOTATION_STATUSES:
        team_totals[_q_status_key(_st)] = 0
    for cat in dynamic_categories:
        team_totals[f"act_{cat}_lead"] = 0
        team_totals[f"act_{cat}_contact"] = 0

    row_idx_tp = 0
    for sp in sp_users:
        d = agg_user(sp["user_id"])
        if (d["leads_created"] == 0 and
            d["contacts_created"] == 0 and
            d["quotations_created"] == 0 and
            d["act_done"] == 0):
            continue
        fill = ALT_FILL if row_idx_tp % 2 == 0 else None
        row_vals = [
            sp["sp_name"], d["leads_created"], d["contacts_created"], d["quotations_created"],
        ]
        row_vals += [d.get(_q_status_key(st), 0) for st in QUOTATION_STATUSES]
        row_vals += [d["act_done"]]
        for cat in dynamic_categories:
            row_vals.append(d.get(f"act_{cat}_lead", 0))
            row_vals.append(d.get(f"act_{cat}_contact", 0))
        row_vals.extend([
            d["o"], d["c"], d["e"],
            d["enq_c"],  money(d["enq_v"]),
            d["pipe_c"], money(d["pipe_v"]),
            d["ord_c"],  money(d["ord_v"]),
            d["lenq_c"]+d["lpipe_c"],
            d["c_to_l"], d["l_to_o"],
            d["act_total"], d["act_done"], d["act_open"],
        ])
        for ci, v in enumerate(row_vals, 1):
            ws_tp.cell(r_tp, ci, v).font = DATA_FONT
        apply_row(ws_tp, r_tp, 1, len(tm_hdrs), fill=fill); r_tp += 1
        row_idx_tp += 1
        totals_keys = ["o", "c", "e", "enq_c", "pipe_c", "ord_c", "c_to_l", "l_to_o", "act_total", "act_done", "act_open",
                       "leads_created", "contacts_created", "quotations_created",
                       "draft_quotations", "open_quotations"]
        totals_keys += [_q_status_key(st) for st in QUOTATION_STATUSES]
        for cat in dynamic_categories:
            totals_keys.append(f"act_{cat}_lead")
            totals_keys.append(f"act_{cat}_contact")
        for k in totals_keys:
            team_totals[k] += d.get(k, 0)
        team_totals["lenq_c"] += d.get("lenq_c", 0) + d.get("lpipe_c", 0)
        team_totals["enq_v"]  += d.get("enq_v", 0)
        team_totals["pipe_v"] += d.get("pipe_v", 0)
        team_totals["ord_v"]  += d.get("ord_v", 0)
    # Total row
    tot_row_vals = [
        "TOTAL", team_totals["leads_created"], team_totals["contacts_created"], team_totals["quotations_created"],
    ]
    tot_row_vals += [team_totals[_q_status_key(st)] for st in QUOTATION_STATUSES]
    tot_row_vals += [team_totals["act_done"]]
    for cat in dynamic_categories:
        tot_row_vals.append(team_totals[f"act_{cat}_lead"])
        tot_row_vals.append(team_totals[f"act_{cat}_contact"])
    tot_row_vals.extend([
        team_totals["o"], team_totals["c"], team_totals["e"],
        team_totals["enq_c"], money(team_totals["enq_v"]),
        team_totals["pipe_c"], money(team_totals["pipe_v"]),
        team_totals["ord_c"],  money(team_totals["ord_v"]),
        team_totals["lenq_c"],
        team_totals["c_to_l"], team_totals["l_to_o"],
        team_totals["act_total"], team_totals["act_done"], team_totals["act_open"],
    ])
    for ci, v in enumerate(tot_row_vals, 1):
        ws_tp.cell(r_tp, ci, v).font = TOT_FONT
    apply_row(ws_tp, r_tp, 1, len(tm_hdrs), font=TOT_FONT, fill=SILVER_FILL); r_tp += 1

    auto_width(ws); auto_width(ws_tp)
    ws.column_dimensions["A"].width = 24; ws_tp.column_dimensions["A"].width = 24
    ws_tp.freeze_panes = "A5"

    # Individual per-owner tabs are omitted to simplify the workbook layout.

    # ==========================================================
    # TAB: ALL CONTACTS
    # ==========================================================
    ws_c = wb.create_sheet(title="All Contacts")
    p_cd = {"usr": eff_c_user, "sd": c_from, "ed": c_to}
    wh_cd = "WHERE docstatus < 2"
    if eff_c_user: wh_cd += " AND assign_to = %(usr)s"
    wh_cd += build_in_clause(c_ind_list, "cind", "industry", p_cd)
    wh_cd += build_in_clause(c_src_list, "csrc", "source", p_cd)
    if c_from and c_to: wh_cd += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
    contacts_list = frappe.db.sql(
        "SELECT name,contact_name,organization_name,mobile_number,email_id,"
        "industry,assign_to,status,DATE_FORMAT(creation,'%%d-%%b-%%Y') created_on"
        " FROM `tabBusiness Contacts` " + wh_cd + " ORDER BY creation DESC",
        p_cd, as_dict=1)
    for _rec in contacts_list:
        _rec["assign_to"] = user_name(_rec.get("assign_to"))

    title_banner(ws_c, "Business Contacts Detail  |  {} Records".format(len(contacts_list)), 10, GREEN)
    for ci, h in enumerate(["#","Contact","Organisation","Mobile","Email","Industry","Owner","Status","Created"], 1):
        cc = ws_c.cell(3, ci, h); cc.font = HDR_FONT; cc.fill = _fill("22C55E"); cc.border = THIN
    ws_c.row_dimensions[3].height = 22
    for idx, rec in enumerate(contacts_list, 1):
        rw = idx + 3; fill = ALT_FILL if idx % 2 == 0 else None
        ws_c.cell(rw, 1, idx).font = DATA_FONT
        # Link field: shows the person's name, points at the record.
        lk = ws_c.cell(rw, 2, rec.get("contact_name") or rec["name"]); lk.font = LINK_FONT
        lk.hyperlink = "{}/app/business-contacts/{}".format(frappe.utils.get_url(), rec["name"])
        for ci, fld in enumerate(["organization_name","mobile_number",
                                   "email_id","industry","assign_to","status","created_on"], 3):
            ws_c.cell(rw, ci, rec.get(fld) or "-").font = DATA_FONT
        apply_row(ws_c, rw, 1, 9, fill=fill)
    tr = len(contacts_list) + 4
    ws_c.cell(tr, 1, "Total:").font = TOT_FONT; ws_c.cell(tr, 2, len(contacts_list)).font = TOT_FONT
    apply_row(ws_c, tr, 1, 9, font=TOT_FONT, fill=SILVER_FILL)
    auto_width(ws_c)

    # ==========================================================
    # TAB: ALL LEADS
    # ==========================================================
    ws_l = wb.create_sheet(title="All Leads")
    p_ld = {"usr": eff_l_user, "sd": l_from, "ed": l_to}
    wh_ld = "WHERE l.docstatus < 2"
    if eff_l_user: wh_ld += " AND l.lead_owner = %(usr)s"
    wh_ld += build_in_clause(l_ind_list, "lind", "l.industry", p_ld)
    wh_ld += build_in_clause(l_src_list, "lsrc", "l.source", p_ld)
    if l_from and l_to: wh_ld += " AND DATE(l.creation) BETWEEN %(sd)s AND %(ed)s"
    leads_list = frappe.db.sql(
        "SELECT l.name, l.lead_name, l.company_name, l.custom_lead_category, l.status, "
        "l.source, l.custom_sub_source, l.industry, l.custom_business_contacts, "
        "l.mobile_no, l.email_id, l.lead_owner, "
        "COALESCE(l.custom_expected_revenue, 0) as est_revenue, "
        "COALESCE(l.custom_po_value, 0) as po_value, "
        "DATE_FORMAT(l.custom_expected_closure_date, '%%d-%%b-%%Y') as expected_closure_date, "
        "COALESCE(l.custom_lost_enquiry_reason, l.custom_lost_pipeline_reason) as lost_reason, "
        "l.custom_lead_description, "
        "DATE_FORMAT(l.creation, '%%d-%%b-%%Y') created_on, "
        "DATEDIFF(NOW(), l.creation) age_days "
        " FROM `tabLead` l " + wh_ld + " ORDER BY l.creation DESC",
        p_ld, as_dict=1)
    for _rec in leads_list:
        _rec["lead_owner"] = user_name(_rec.get("lead_owner"))

    headers = [
        "#", "Lead ID", "Lead Name", "Company", "Category", "Status", "Source", "Sub Source",
        "Industry", "Converted From Contact", "Mobile", "Email", "Owner", "Est Revenue",
        "PO Value", "Expected Closure Date", "Lost Reason", "Lead Description", "Created", "Age (days)"
    ]

    title_banner(ws_l, "Leads Detail  |  {} Records".format(len(leads_list)), len(headers), ORANGE)
    for ci, h in enumerate(headers, 1):
        lc2 = ws_l.cell(3, ci, h); lc2.font = HDR_FONT; lc2.fill = _fill("F97316"); lc2.border = THIN
    ws_l.row_dimensions[3].height = 22
    total_rev = 0.0
    total_po_val = 0.0
    for idx, rec in enumerate(leads_list, 1):
        rw = idx + 3; fill = ALT_FILL if idx % 2 == 0 else None
        ws_l.cell(rw, 1, idx).font = DATA_FONT
        
        # Lead ID (Hyperlinked)
        lk_id = ws_l.cell(rw, 2, rec["name"]); lk_id.font = LINK_FONT
        lk_id.hyperlink = "{}/app/lead/{}".format(frappe.utils.get_url(), rec["name"])
        
        # Other simple fields
        ws_l.cell(rw, 3, rec.get("lead_name") or "-").font = DATA_FONT
        ws_l.cell(rw, 4, rec.get("company_name") or "-").font = DATA_FONT
        ws_l.cell(rw, 5, rec.get("custom_lead_category") or "-").font = DATA_FONT
        ws_l.cell(rw, 6, rec.get("status") or "-").font = DATA_FONT
        ws_l.cell(rw, 7, rec.get("source") or "-").font = DATA_FONT
        ws_l.cell(rw, 8, rec.get("custom_sub_source") or "-").font = DATA_FONT
        ws_l.cell(rw, 9, rec.get("industry") or "-").font = DATA_FONT
        ws_l.cell(rw, 10, rec.get("custom_business_contacts") or "-").font = DATA_FONT
        ws_l.cell(rw, 11, rec.get("mobile_no") or "-").font = DATA_FONT
        ws_l.cell(rw, 12, rec.get("email_id") or "-").font = DATA_FONT
        ws_l.cell(rw, 13, rec.get("lead_owner") or "-").font = DATA_FONT
        
        # Estimated Revenue
        rev = flt(rec.get("est_revenue", 0)); total_rev += rev
        rv = ws_l.cell(rw, 14, rev); rv.font = DATA_FONT; rv.number_format = '"Rs."#,##0.00'
        
        # PO Value
        po_val = flt(rec.get("po_value", 0)); total_po_val += po_val
        pv = ws_l.cell(rw, 15, po_val); pv.font = DATA_FONT; pv.number_format = '"Rs."#,##0.00'
        
        # Expected Closure Date, Lost Reason, Description
        ws_l.cell(rw, 16, rec.get("expected_closure_date") or "-").font = DATA_FONT
        ws_l.cell(rw, 17, rec.get("lost_reason") or "-").font = DATA_FONT
        ws_l.cell(rw, 18, rec.get("custom_lead_description") or "-").font = DATA_FONT
        
        # Created On and Age
        ws_l.cell(rw, 19, rec.get("created_on") or "-").font = DATA_FONT
        ws_l.cell(rw, 20, rec.get("age_days") or 0).font = DATA_FONT
        
        apply_row(ws_l, rw, 1, len(headers), fill=fill)
        
    tr2 = len(leads_list) + 4
    ws_l.cell(tr2, 1, "Total:").font = TOT_FONT; ws_l.cell(tr2, 2, len(leads_list)).font = TOT_FONT
    
    # Write summary totals for Est Revenue and PO Value
    trv = ws_l.cell(tr2, 14, total_rev); trv.font = TOT_FONT; trv.number_format = '"Rs."#,##0.00'
    tpv = ws_l.cell(tr2, 15, total_po_val); tpv.font = TOT_FONT; tpv.number_format = '"Rs."#,##0.00'
    
    apply_row(ws_l, tr2, 1, len(headers), font=TOT_FONT, fill=SILVER_FILL)
    auto_width(ws_l)

    # ==========================================================
    # TAB: ALL ACTIVITIES
    # ==========================================================
    ws_a = wb.create_sheet(title="All Activities")
    p_ad = {"usr": eff_ea_user, "sd": ea_from, "ed": ea_to}
    wh_ad = "WHERE ea.docstatus < 2"
    if eff_ea_user: wh_ad += " AND ea.assigned_to = %(usr)s"
    if ea_entity != "All": wh_ad += " AND ea.reference_type = %(ent)s"; p_ad["ent"] = ea_entity
    if ea_from and ea_to: wh_ad += " AND DATE(COALESCE(ea.ends_on,ea.starts_on)) BETWEEN %(sd)s AND %(ed)s"
    activities_list = frappe.db.sql(
        "SELECT ea.name,ea.subject,ea.category,ea.reference_type,ea.reference_name,"
        "ea.assigned_to,ea.status,"
        "DATE_FORMAT(ea.starts_on,'%%d-%%b-%%Y %%h:%%i %%p') starts_on,"
        "DATE_FORMAT(ea.ends_on,'%%d-%%b-%%Y %%h:%%i %%p') ends_on,ea.notes"
        " FROM `tabEvent Activity` ea " + wh_ad + " ORDER BY ea.creation DESC",
        p_ad, as_dict=1)
    for _rec in activities_list:
        _rec["assigned_to"] = user_name(_rec.get("assigned_to"))

    title_banner(ws_a, "Activities Detail  |  {} Records".format(len(activities_list)), 11, PURPLE)
    for ci, h in enumerate(["#","Activity","Category","Ref Type","Ref Name","Assigned To","Status","Start","End","Notes"], 1):
        ac2 = ws_a.cell(3, ci, h); ac2.font = HDR_FONT; ac2.fill = _fill("9333EA"); ac2.border = THIN
    ws_a.row_dimensions[3].height = 22
    for idx, rec in enumerate(activities_list, 1):
        rw = idx + 3; fill = ALT_FILL if idx % 2 == 0 else None
        ws_a.cell(rw, 1, idx).font = DATA_FONT
        lk3 = ws_a.cell(rw, 2, rec.get("subject") or rec["name"]); lk3.font = LINK_FONT
        lk3.hyperlink = "{}/app/event-activity/{}".format(frappe.utils.get_url(), rec["name"])
        for ci, fld in enumerate(["category","reference_type","reference_name",
                                   "assigned_to","status","starts_on","ends_on","notes"], 3):
            ws_a.cell(rw, ci, rec.get(fld) or "-").font = DATA_FONT
        apply_row(ws_a, rw, 1, 10, fill=fill)
    tr3 = len(activities_list) + 4
    ws_a.cell(tr3, 1, "Total:").font = TOT_FONT; ws_a.cell(tr3, 2, len(activities_list)).font = TOT_FONT
    apply_row(ws_a, tr3, 1, 10, font=TOT_FONT, fill=SILVER_FILL)
    auto_width(ws_a)

    # ==========================================================
    # TAB: OVERALL SUMMARY (ALL-TIME)
    # ==========================================================
    ws_ov = wb.create_sheet(title="Overall Summary (All-Time)")
    title_banner(ws_ov,
        "DAC CRM  |  Overall All-Time Team Performance Summary  |  Generated {}".format(frappe.utils.today()),
        23, "1E3A8A")
    rov = 3
    ws_ov.cell(rov, 1, "All-Time Team Member Performance (No Date Filter)").font = SEC_FONT; rov += 1
    ov_hdrs = ["Team Member", "Leads Created", "Contacts Created",
               "Quotations Created (Total)"]
    ov_hdrs += ["→ {}".format(st) for st in QUOTATION_STATUSES]
    ov_hdrs += ["Activities Completed"]
    for cat in dynamic_categories:
        ov_hdrs.append(f"{cat} (Lead)")
        ov_hdrs.append(f"{cat} (Contact)")
    ov_hdrs.extend([
        "Draft Quotations (All-Time)", "Open Quotations (All-Time)",
        "Contacts (Open)", "Contacts (Conv)", "Contacts (Cust)",
        "Enquiry #", "Enquiry Val", "Pipeline #", "Pipeline Val",
        "Order #", "Order Val", "Lost #", "BC to Lead", "Lead to Order",
        "Activity (Total)", "Activity (Done)", "Activity (Open)"
    ])

    groups = [
        ("Sales Person", 1, "1E3A8A"),
        ("Creation Counts (All-Time)", 2, "2563EB"),
        ("Quotations Created (All-Time)  –  Total = sum of statuses",
         1 + len(QUOTATION_STATUSES), "0284C7"),
        ("Activities Completed (All-Time)", 1, "0D9488"),
    ]
    if dynamic_categories:
        groups.append(("Completed Activities by Category", len(dynamic_categories) * 2, "059669"))
    groups.extend([
        ("Pending Quotations (All-Time)", 2, "B45309"),
        ("Contact Status", 3, "475569"),
        ("Lead Pipelines", 7, "7C3AED"),
        ("Conversion Rates (%)", 2, "DB2777"),
        ("Activity Status", 3, "374151")
    ])

    # Write Group Headers
    c_idx = 1
    for g_name, num_cols, color in groups:
        cell = ws_ov.cell(rov, c_idx, g_name)
        cell.font = openpyxl.styles.Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = _fill(color)
        cell.alignment = _align("center", "center", wrap=True)
        cell.border = THIN
        if num_cols > 1:
            ws_ov.merge_cells(start_row=rov, start_column=c_idx, end_row=rov, end_column=c_idx + num_cols - 1)
            for col in range(c_idx, c_idx + num_cols):
                ws_ov.cell(rov, col).border = THIN
                ws_ov.cell(rov, col).fill = _fill(color)
        c_idx += num_cols
    ws_ov.row_dimensions[rov].height = 24; rov += 1

    # Write Sub-Headers
    for ci, h in enumerate(ov_hdrs, 1):
        cell = ws_ov.cell(rov, ci, h)
        cell.font = openpyxl.styles.Font(name="Segoe UI", size=9, bold=True, color="333333")
        cell.fill = _fill("F1F5F9")
        cell.border = THIN
        cell.alignment = _align("center", "center", wrap=True)
    ws_ov.row_dimensions[rov].height = 28; rov += 1
    ov_totals = {k: 0 for k in ["o", "c", "e", "enq_c", "enq_v", "pipe_c", "pipe_v",
                                "ord_c", "ord_v", "lenq_c", "c_to_l", "l_to_o",
                                "act_total", "act_done", "act_open",
                                "leads_created", "contacts_created", "quotations_created",
                                "draft_quotations", "open_quotations"]}
    for _st in QUOTATION_STATUSES:
        ov_totals[_q_status_key(_st)] = 0
    for cat in dynamic_categories:
        ov_totals[f"act_{cat}_lead"] = 0
        ov_totals[f"act_{cat}_contact"] = 0

    row_idx_ov = 0
    for sp in sp_users:
        d_ov = agg_user(sp["user_id"], all_time=True)  # all-time, no date/fiscal-year filter
        if (d_ov["leads_created"] == 0 and
            d_ov["contacts_created"] == 0 and
            d_ov["quotations_created"] == 0 and
            d_ov["act_done"] == 0):
            continue
        fill = ALT_FILL if row_idx_ov % 2 == 0 else None
        row_vals = [
            sp["sp_name"], d_ov["leads_created"], d_ov["contacts_created"], d_ov["quotations_created"],
        ]
        row_vals += [d_ov.get(_q_status_key(st), 0) for st in QUOTATION_STATUSES]
        row_vals += [d_ov["act_done"]]
        for cat in dynamic_categories:
            row_vals.append(d_ov.get(f"act_{cat}_lead", 0))
            row_vals.append(d_ov.get(f"act_{cat}_contact", 0))
        row_vals.extend([
            d_ov["draft_quotations"], d_ov["open_quotations"],
            d_ov["o"], d_ov["c"], d_ov["e"],
            d_ov["enq_c"],  money(d_ov["enq_v"]),
            d_ov["pipe_c"], money(d_ov["pipe_v"]),
            d_ov["ord_c"],  money(d_ov["ord_v"]),
            d_ov["lenq_c"]+d_ov["lpipe_c"],
            d_ov["c_to_l"], d_ov["l_to_o"],
            d_ov["act_total"], d_ov["act_done"], d_ov["act_open"],
        ])
        for ci, v in enumerate(row_vals, 1):
            ws_ov.cell(rov, ci, v).font = DATA_FONT
        apply_row(ws_ov, rov, 1, len(ov_hdrs), fill=fill); rov += 1
        row_idx_ov += 1
        totals_keys = ["o", "c", "e", "enq_c", "pipe_c", "ord_c", "c_to_l", "l_to_o", "act_total", "act_done", "act_open",
                       "leads_created", "contacts_created", "quotations_created",
                       "draft_quotations", "open_quotations"]
        totals_keys += [_q_status_key(st) for st in QUOTATION_STATUSES]
        for cat in dynamic_categories:
            totals_keys.append(f"act_{cat}_lead")
            totals_keys.append(f"act_{cat}_contact")
        for k in totals_keys:
            ov_totals[k] += d_ov.get(k, 0)
        ov_totals["lenq_c"] += d_ov.get("lenq_c", 0) + d_ov.get("lpipe_c", 0)
        ov_totals["enq_v"]  += d_ov.get("enq_v", 0)
        ov_totals["pipe_v"] += d_ov.get("pipe_v", 0)
        ov_totals["ord_v"]  += d_ov.get("ord_v", 0)
    # Overall total row
    tot_row_vals = [
        "TOTAL", ov_totals["leads_created"], ov_totals["contacts_created"], ov_totals["quotations_created"],
    ]
    tot_row_vals += [ov_totals[_q_status_key(st)] for st in QUOTATION_STATUSES]
    tot_row_vals += [ov_totals["act_done"]]
    for cat in dynamic_categories:
        tot_row_vals.append(ov_totals[f"act_{cat}_lead"])
        tot_row_vals.append(ov_totals[f"act_{cat}_contact"])
    tot_row_vals.extend([
        ov_totals["draft_quotations"], ov_totals["open_quotations"],
        ov_totals["o"], ov_totals["c"], ov_totals["e"],
        ov_totals["enq_c"], money(ov_totals["enq_v"]),
        ov_totals["pipe_c"], money(ov_totals["pipe_v"]),
        ov_totals["ord_c"],  money(ov_totals["ord_v"]),
        ov_totals["lenq_c"],
        ov_totals["c_to_l"], ov_totals["l_to_o"],
        ov_totals["act_total"], ov_totals["act_done"], ov_totals["act_open"],
    ])
    for ci, v in enumerate(tot_row_vals, 1):
        ws_ov.cell(rov, ci, v).font = TOT_FONT
    apply_row(ws_ov, rov, 1, len(ov_hdrs), font=TOT_FONT, fill=SILVER_FILL)
    auto_width(ws_ov); ws_ov.column_dimensions["A"].width = 24; ws_ov.freeze_panes = "A2"

    # ==========================================================
    # TAB: PENDING QUOTATIONS (Draft & Open)
    # ==========================================================
    ws_pq = wb.create_sheet(title="Pending Quotations (All-Time)")
    p_pq = {"usr": eff_l_user or eff_c_user}
    # Every live status (docstatus < 2). Restricting this to Draft + Open hid
    # Expired and Lost quotations, which are exactly the ones needing attention.
    wh_pq = "WHERE q.docstatus < 2"
    if eff_l_user or eff_c_user:
        wh_pq += " AND COALESCE(NULLIF(q.custom_lead_owner, ''), q.owner) = %(usr)s"

    pending_list = frappe.db.sql(
        "SELECT q.name, q.customer_name, q.quotation_to, q.custom_lead_owner, q.owner, q.status, q.docstatus, q.grand_total, "
        "DATE_FORMAT(q.creation,'%%d-%%b-%%Y') created_on"
        " FROM `tabQuotation` q " + wh_pq + " ORDER BY q.creation DESC",
        p_pq, as_dict=1)
    for _rec in pending_list:
        _rec["custom_lead_owner"] = user_name(_rec.get("custom_lead_owner") or _rec.get("owner"))
        _rec["owner"] = user_name(_rec.get("owner"))

    title_banner(ws_pq, "Pending Quotations (All Live Statuses)  |  {} Records".format(len(pending_list)), 9, "e63946")
    for ci, h in enumerate(["#", "Customer", "Quotation To", "Lead Owner", "Creator", "Status", "Docstatus", "Grand Total", "Created"], 1):
        qc2 = ws_pq.cell(3, ci, h); qc2.font = HDR_FONT; qc2.fill = _fill("e63946"); qc2.border = THIN
    ws_pq.row_dimensions[3].height = 22
    total_pq_amt = 0.0
    for idx, rec in enumerate(pending_list, 1):
        rw = idx + 3; fill = ALT_FILL if idx % 2 == 0 else None
        ws_pq.cell(rw, 1, idx).font = DATA_FONT
        # Link field: shows the customer, points at the quotation.
        lkq = ws_pq.cell(rw, 2, rec.get("customer_name") or rec["name"]); lkq.font = LINK_FONT
        lkq.hyperlink = "{}/app/quotation/{}".format(frappe.utils.get_url(), rec["name"])
        docstat_str = "Draft" if rec["docstatus"] == 0 else ("Submitted" if rec["docstatus"] == 1 else "Cancelled")
        for ci, val in enumerate([
            rec.get("quotation_to") or "-",
            rec.get("custom_lead_owner") or "-", rec.get("owner") or "-",
            rec.get("status") or "-", docstat_str
        ], 3):
            ws_pq.cell(rw, ci, val).font = DATA_FONT
        
        amt = flt(rec.get("grand_total", 0)); total_pq_amt += amt
        rvq = ws_pq.cell(rw, 8, amt); rvq.font = DATA_FONT; rvq.number_format = '"Rs."#,##0.00'
        ws_pq.cell(rw, 9, rec.get("created_on") or "-").font = DATA_FONT
        apply_row(ws_pq, rw, 1, 9, fill=fill)
        
    trpq = len(pending_list) + 4
    ws_pq.cell(trpq, 1, "Total:").font = TOT_FONT; ws_pq.cell(trpq, 2, len(pending_list)).font = TOT_FONT
    trvpq = ws_pq.cell(trpq, 8, total_pq_amt); trvpq.font = TOT_FONT; trvpq.number_format = '"Rs."#,##0.00'
    apply_row(ws_pq, trpq, 1, 9, font=TOT_FONT, fill=SILVER_FILL)
    auto_width(ws_pq)

    # ==========================================================
    # TAB: ALL QUOTATIONS
    # ==========================================================
    ws_q = wb.create_sheet(title="Quotations (This Period)")
    p_qd = {"usr": eff_l_user or eff_c_user, "sd": l_from, "ed": l_to}
    wh_qd = "WHERE q.docstatus < 2"
    if eff_l_user or eff_c_user:
        wh_qd += " AND COALESCE(NULLIF(q.custom_lead_owner, ''), q.owner) = %(usr)s"
    if l_from and l_to:
        wh_qd += " AND DATE(q.creation) BETWEEN %(sd)s AND %(ed)s"

    quotations_list = frappe.db.sql(
        "SELECT q.name, q.customer_name, q.quotation_to, q.custom_lead_owner, q.owner, q.status, q.docstatus, q.grand_total, "
        "DATE_FORMAT(q.creation,'%%d-%%b-%%Y') created_on"
        " FROM `tabQuotation` q " + wh_qd + " ORDER BY q.creation DESC",
        p_qd, as_dict=1)
    for _rec in quotations_list:
        _rec["custom_lead_owner"] = user_name(_rec.get("custom_lead_owner") or _rec.get("owner"))
        _rec["owner"] = user_name(_rec.get("owner"))

    title_banner(ws_q, "Quotations Created in {}  |  {} Records".format(period_label_str, len(quotations_list)), 9, "4361ee")
    for ci, h in enumerate(["#", "Customer", "Quotation To", "Lead Owner", "Creator", "Status", "Docstatus", "Grand Total", "Created"], 1):
        qc2 = ws_q.cell(3, ci, h); qc2.font = HDR_FONT; qc2.fill = _fill("4361ee"); qc2.border = THIN
    ws_q.row_dimensions[3].height = 22
    total_q_amt = 0.0
    for idx, rec in enumerate(quotations_list, 1):
        rw = idx + 3; fill = ALT_FILL if idx % 2 == 0 else None
        ws_q.cell(rw, 1, idx).font = DATA_FONT
        # Link field: shows the customer, points at the quotation.
        lkq = ws_q.cell(rw, 2, rec.get("customer_name") or rec["name"]); lkq.font = LINK_FONT
        lkq.hyperlink = "{}/app/quotation/{}".format(frappe.utils.get_url(), rec["name"])
        docstat_str = "Draft" if rec["docstatus"] == 0 else ("Submitted" if rec["docstatus"] == 1 else "Cancelled")
        for ci, val in enumerate([
            rec.get("quotation_to") or "-",
            rec.get("custom_lead_owner") or "-", rec.get("owner") or "-",
            rec.get("status") or "-", docstat_str
        ], 3):
            ws_q.cell(rw, ci, val).font = DATA_FONT
        
        amt = flt(rec.get("grand_total", 0)); total_q_amt += amt
        rvq = ws_q.cell(rw, 8, amt); rvq.font = DATA_FONT; rvq.number_format = '"Rs."#,##0.00'
        ws_q.cell(rw, 9, rec.get("created_on") or "-").font = DATA_FONT
        apply_row(ws_q, rw, 1, 9, fill=fill)
        
    trq = len(quotations_list) + 4
    ws_q.cell(trq, 1, "Total:").font = TOT_FONT; ws_q.cell(trq, 2, len(quotations_list)).font = TOT_FONT
    trvq = ws_q.cell(trq, 8, total_q_amt); trvq.font = TOT_FONT; trvq.number_format = '"Rs."#,##0.00'
    apply_row(ws_q, trq, 1, 9, font=TOT_FONT, fill=SILVER_FILL)
    auto_width(ws_q)

    # ==========================================================
    # TAB: QUOTATIONS BY USER
    # Draft and Open quotations grouped under the person who owns them,
    # every ID clickable straight into ERP. A quotation belongs to its
    # custom_lead_owner when set, otherwise to whoever created it — the same
    # single rule the counts use, so this sheet always adds up to them.
    # ==========================================================
    ws_qu = wb.create_sheet(title="Quotations by User")
    base_url = frappe.utils.get_url()

    p_qu = {}
    wh_qu = ("WHERE q.docstatus < 2 "
             "AND (q.docstatus = 0 OR (q.docstatus = 1 AND q.status IN ('Open', 'Ordered')))")
    if eff_l_user or eff_c_user:
        wh_qu += " AND COALESCE(NULLIF(q.custom_lead_owner, ''), q.owner) = %(usr)s"
        p_qu["usr"] = eff_l_user or eff_c_user
    if l_from and l_to:
        wh_qu += " AND DATE(q.creation) BETWEEN %(sd)s AND %(ed)s"
        p_qu["sd"] = l_from; p_qu["ed"] = l_to

    qu_rows = frappe.db.sql(
        "SELECT COALESCE(NULLIF(q.custom_lead_owner, ''), q.owner) AS resp_user, "
        "q.name, q.customer_name, q.quotation_to, q.owner, q.status, q.docstatus, q.grand_total, "
        "q.valid_till, DATE_FORMAT(q.creation,'%%d-%%b-%%Y') created_on"
        " FROM `tabQuotation` q " + wh_qu +
        " ORDER BY resp_user ASC, q.docstatus ASC, q.creation DESC",
        p_qu, as_dict=1)

    grouped = {}
    for rec in qu_rows:
        grouped.setdefault(rec["resp_user"], []).append(rec)

    n_draft_all = sum(1 for r in qu_rows if r["docstatus"] == 0)
    n_open_all = sum(1 for r in qu_rows if r["docstatus"] == 1 and r.get("status") == "Open")

    title_banner(
        ws_qu,
        "Quotations by User  |  {} users  |  {} total  ({} Draft, {} Open)".format(
            len(grouped), len(qu_rows), n_draft_all, n_open_all),
        8, "7c3aed")

    # A link back to the CRM dashboard, so the sheet is a way in, not a dead end.
    dash = ws_qu.cell(2, 1, "Open CRM Dashboard")
    dash.hyperlink = "{}/app/crm-dashboard".format(base_url)
    dash.font = LINK_FONT

    qu_hdrs = ["#", "Customer", "Quotation To", "Stage",
               "Grand Total", "Valid Till", "Created"]
    r_qu = 4
    for ci, h in enumerate(qu_hdrs, 1):
        hc = ws_qu.cell(r_qu, ci, h)
        hc.font = HDR_FONT; hc.fill = _fill("7c3aed"); hc.border = THIN
        hc.alignment = _align("center", "center", wrap=True)
    ws_qu.row_dimensions[r_qu].height = 22
    r_qu += 1

    grand_total_all = 0.0
    for resp_user in sorted(grouped.keys(), key=lambda u: (user_name(u) or "").lower()):
        recs = grouped[resp_user]
        u_draft = sum(1 for r in recs if r["docstatus"] == 0)
        u_open = sum(1 for r in recs if r["docstatus"] == 1 and r.get("status") == "Open")
        u_total = sum(flt(r.get("grand_total", 0)) for r in recs)

        # One banded header per person
        uc = ws_qu.cell(r_qu, 1, "{}   —   {} quotation(s):  {} Draft, {} Open".format(
            user_name(resp_user), len(recs), u_draft, u_open))
        ws_qu.merge_cells(start_row=r_qu, start_column=1, end_row=r_qu, end_column=len(qu_hdrs))
        uc.font = openpyxl.styles.Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        uc.fill = _fill("4c1d95")
        uc.alignment = _align("left", "center")
        ws_qu.row_dimensions[r_qu].height = 20
        r_qu += 1

        for i, rec in enumerate(recs, 1):
            fill = ALT_FILL if i % 2 == 0 else None
            ws_qu.cell(r_qu, 1, i).font = DATA_FONT

            lk = ws_qu.cell(r_qu, 2, rec.get("customer_name") or rec["name"])
            lk.font = LINK_FONT
            lk.hyperlink = "{}/app/quotation/{}".format(base_url, rec["name"])

            stage = "Draft" if rec["docstatus"] == 0 else (rec.get("status") or "Submitted")
            ws_qu.cell(r_qu, 3, rec.get("quotation_to") or "-").font = DATA_FONT

            sc = ws_qu.cell(r_qu, 4, stage)
            sc.font = openpyxl.styles.Font(
                name="Segoe UI", size=9, bold=True,
                color="b45309" if stage == "Draft" else ("1d4ed8" if stage == "Open" else "15803d"))

            amt = flt(rec.get("grand_total", 0)); grand_total_all += amt
            ac = ws_qu.cell(r_qu, 5, amt); ac.font = DATA_FONT
            ac.number_format = '"Rs."#,##0.00'

            ws_qu.cell(r_qu, 6, rec.get("valid_till") or "-").font = DATA_FONT
            ws_qu.cell(r_qu, 7, rec.get("created_on") or "-").font = DATA_FONT
            apply_row(ws_qu, r_qu, 1, len(qu_hdrs), fill=fill)
            r_qu += 1

        st = ws_qu.cell(r_qu, 1, "Subtotal — {}".format(user_name(resp_user)))
        ws_qu.merge_cells(start_row=r_qu, start_column=1, end_row=r_qu, end_column=4)
        st.font = TOT_FONT
        sv = ws_qu.cell(r_qu, 5, u_total); sv.font = TOT_FONT
        sv.number_format = '"Rs."#,##0.00'
        apply_row(ws_qu, r_qu, 1, len(qu_hdrs), font=TOT_FONT, fill=SILVER_FILL)
        r_qu += 2

    if not qu_rows:
        ws_qu.cell(r_qu, 1, "No draft or open quotations for this period.").font = DATA_FONT
    else:
        gt = ws_qu.cell(r_qu, 1, "GRAND TOTAL — {} quotation(s)".format(len(qu_rows)))
        ws_qu.merge_cells(start_row=r_qu, start_column=1, end_row=r_qu, end_column=4)
        gt.font = TOT_FONT
        gv = ws_qu.cell(r_qu, 5, grand_total_all); gv.font = TOT_FONT
        gv.number_format = '"Rs."#,##0.00'
        apply_row(ws_qu, r_qu, 1, len(qu_hdrs), font=TOT_FONT, fill=_fill("ddd6fe"))

    auto_width(ws_qu)
    ws_qu.column_dimensions["A"].width = 30
    ws_qu.freeze_panes = "A5"

    # ==========================================================
    # TAB: OPEN QUOTATIONS BY USER (ALL-TIME)
    # Everything still in play, whatever its age: Draft, Open, Replied,
    # Partially Ordered, Lost and Expired. Ordered is excluded — that work is
    # won and has moved on to a Sales Order. No date filter, so a quotation
    # raised months ago and never closed cannot quietly drop off the report.
    # ==========================================================
    ws_qa = wb.create_sheet(title="Open Quotations (All-Time)")

    p_qa = {}
    wh_qa = ("WHERE q.docstatus < 2 "
             "AND IFNULL(q.status, '') != 'Ordered'")
    if eff_l_user or eff_c_user:
        wh_qa += " AND COALESCE(NULLIF(q.custom_lead_owner, ''), q.owner) = %(usr)s"
        p_qa["usr"] = eff_l_user or eff_c_user

    qa_rows = frappe.db.sql(
        "SELECT COALESCE(NULLIF(q.custom_lead_owner, ''), q.owner) AS resp_user, "
        "q.name, q.customer_name, q.quotation_to, q.status, q.docstatus, q.grand_total, "
        "q.valid_till, DATE_FORMAT(q.creation,'%%d-%%b-%%Y') created_on, "
        "DATEDIFF(CURDATE(), DATE(q.creation)) AS age_days"
        " FROM `tabQuotation` q " + wh_qa +
        " ORDER BY resp_user ASC, q.creation DESC",
        p_qa, as_dict=1)

    qa_grouped = {}
    for rec in qa_rows:
        qa_grouped.setdefault(rec["resp_user"], []).append(rec)

    qa_by_status = {}
    for rec in qa_rows:
        st = "Draft" if rec["docstatus"] == 0 else (rec.get("status") or "Submitted")
        qa_by_status[st] = qa_by_status.get(st, 0) + 1
    status_summary = "  |  ".join("{}: {}".format(k, qa_by_status[k]) for k in sorted(qa_by_status))

    title_banner(
        ws_qa,
        "Open Quotations — All-Time (everything except Ordered)  |  {} users  |  {} quotations".format(
            len(qa_grouped), len(qa_rows)),
        8, "0f766e")
    sc_cell = ws_qa.cell(2, 1, status_summary or "Nothing outstanding")
    sc_cell.font = openpyxl.styles.Font(name="Segoe UI", size=9, bold=True, color="0f766e")

    qa_hdrs = ["#", "Customer", "Quotation To", "Stage", "Grand Total",
               "Valid Till", "Created", "Age (days)"]
    r_qa = 4
    for ci, h in enumerate(qa_hdrs, 1):
        hc = ws_qa.cell(r_qa, ci, h)
        hc.font = HDR_FONT; hc.fill = _fill("0f766e"); hc.border = THIN
        hc.alignment = _align("center", "center", wrap=True)
    ws_qa.row_dimensions[r_qa].height = 22
    r_qa += 1

    qa_grand = 0.0
    for resp_user in sorted(qa_grouped.keys(), key=lambda u: (user_name(u) or "").lower()):
        recs = qa_grouped[resp_user]
        u_total = sum(flt(r.get("grand_total", 0)) for r in recs)
        u_status = {}
        for r in recs:
            st = "Draft" if r["docstatus"] == 0 else (r.get("status") or "Submitted")
            u_status[st] = u_status.get(st, 0) + 1

        uc = ws_qa.cell(r_qa, 1, "{}   —   {} open:  {}".format(
            user_name(resp_user), len(recs),
            ", ".join("{} {}".format(u_status[k], k) for k in sorted(u_status))))
        ws_qa.merge_cells(start_row=r_qa, start_column=1, end_row=r_qa, end_column=len(qa_hdrs))
        uc.font = openpyxl.styles.Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        uc.fill = _fill("115e59")
        uc.alignment = _align("left", "center")
        ws_qa.row_dimensions[r_qa].height = 20
        r_qa += 1

        for i, rec in enumerate(recs, 1):
            fill = ALT_FILL if i % 2 == 0 else None
            ws_qa.cell(r_qa, 1, i).font = DATA_FONT

            lk = ws_qa.cell(r_qa, 2, rec.get("customer_name") or rec["name"])
            lk.font = LINK_FONT
            lk.hyperlink = "{}/app/quotation/{}".format(base_url, rec["name"])

            stage = "Draft" if rec["docstatus"] == 0 else (rec.get("status") or "Submitted")
            ws_qa.cell(r_qa, 3, rec.get("quotation_to") or "-").font = DATA_FONT

            stage_colour = {"Draft": "b45309", "Open": "1d4ed8", "Expired": "b91c1c",
                            "Lost": "6b7280", "Replied": "7c3aed"}.get(stage, "334155")
            sc = ws_qa.cell(r_qa, 4, stage)
            sc.font = openpyxl.styles.Font(name="Segoe UI", size=9, bold=True, color=stage_colour)

            amt = flt(rec.get("grand_total", 0)); qa_grand += amt
            ac = ws_qa.cell(r_qa, 5, amt); ac.font = DATA_FONT
            ac.number_format = '"Rs."#,##0.00'

            ws_qa.cell(r_qa, 6, rec.get("valid_till") or "-").font = DATA_FONT
            ws_qa.cell(r_qa, 7, rec.get("created_on") or "-").font = DATA_FONT
            ws_qa.cell(r_qa, 8, rec.get("age_days") or 0).font = DATA_FONT
            apply_row(ws_qa, r_qa, 1, len(qa_hdrs), fill=fill)
            r_qa += 1

        st_cell = ws_qa.cell(r_qa, 1, "Subtotal — {}".format(user_name(resp_user)))
        ws_qa.merge_cells(start_row=r_qa, start_column=1, end_row=r_qa, end_column=4)
        st_cell.font = TOT_FONT
        sv = ws_qa.cell(r_qa, 5, u_total); sv.font = TOT_FONT
        sv.number_format = '"Rs."#,##0.00'
        apply_row(ws_qa, r_qa, 1, len(qa_hdrs), font=TOT_FONT, fill=SILVER_FILL)
        r_qa += 2

    if not qa_rows:
        ws_qa.cell(r_qa, 1, "No open quotations — everything is Ordered or closed.").font = DATA_FONT
    else:
        gt = ws_qa.cell(r_qa, 1, "GRAND TOTAL — {} open quotation(s)".format(len(qa_rows)))
        ws_qa.merge_cells(start_row=r_qa, start_column=1, end_row=r_qa, end_column=4)
        gt.font = TOT_FONT
        gv = ws_qa.cell(r_qa, 5, qa_grand); gv.font = TOT_FONT
        gv.number_format = '"Rs."#,##0.00'
        apply_row(ws_qa, r_qa, 1, len(qa_hdrs), font=TOT_FONT, fill=_fill("ccfbf1"))

    auto_width(ws_qa)
    ws_qa.column_dimensions["A"].width = 32
    ws_qa.freeze_panes = "A5"

    # ── Save ────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output); output.seek(0)
    raw_bytes = output.getvalue()
    fname = "DAC_CRM_{}_Report_{}.xlsx".format(export_type.title(), frappe.utils.today())
    file_base64 = base64.b64encode(raw_bytes).decode("utf-8")

    # ── Email ────────────────────────────────────────────────────
    email_sent = False
    if send_email and email_address:
        recipients = [e.strip() for e in str(email_address).split(",") if e.strip()]
        if recipients:
            plabel = {"daily":"Daily","weekly":"Weekly","monthly":"Monthly","applied":"Custom Filter"}.get(export_type, export_type.title())
            html_body = """<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
body{{font-family:Arial,sans-serif;background:#f8fafc;margin:0;padding:0;}}
.wrap{{max-width:640px;margin:32px auto;background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 4px 16px rgba(0,0,0,.10);}}
.hdr{{background:linear-gradient(135deg,#0f172a 0%,#1e3a8a 100%);padding:32px 40px;}}
.hdr h1{{color:#fff;margin:0;font-size:22px;font-weight:700;}}
.hdr p{{color:#93c5fd;margin:6px 0 0;font-size:14px;}}
.body{{padding:32px 40px;}}
.body p{{color:#334155;font-size:14px;line-height:1.7;margin:0 0 12px;}}
table.info{{width:100%;border-collapse:collapse;margin:20px 0;}}
table.info td{{padding:9px 12px;font-size:13px;border-bottom:1px solid #e2e8f0;color:#334155;}}
table.info td:first-child{{font-weight:600;color:#1e293b;width:42%;background:#f8fafc;}}
.badge{{display:inline-block;background:#eff6ff;color:#1d4ed8;padding:3px 10px;border-radius:20px;font-size:12px;font-weight:700;}}
ul{{color:#334155;font-size:14px;line-height:1.9;}}
.footer{{background:#f1f5f9;padding:20px 40px;text-align:center;color:#94a3b8;font-size:12px;border-top:1px solid #e2e8f0;}}
</style></head>
<body><div class="wrap">
<div class="hdr">
  <h1>&#128202; DAC CRM Executive Report</h1>
  <p>{plabel} Report &nbsp;&#183;&nbsp; Generated on {today}</p>
</div>
<div class="body">
  <p>Please find attached the <strong>{plabel} CRM Executive Dashboard Report</strong> generated on <strong>{today}</strong>.</p>
  <table class="info">
    <tr><td>Report Type</td>     <td><span class="badge">{plabel}</span></td></tr>
    <tr><td>Fiscal Year</td>     <td>{fy}</td></tr>
    <tr><td>Contacts Range</td>  <td>{c_rng}</td></tr>
    <tr><td>Leads Range</td>     <td>{l_rng}</td></tr>
    <tr><td>Total Contacts</td>  <td><strong>{tc}</strong></td></tr>
    <tr><td>Open Contacts</td>   <td><strong>{oc}</strong></td></tr>
    <tr><td>Total Leads</td>     <td><strong>{tl}</strong></td></tr>
    <tr><td>Order Count</td>     <td><strong>{ord_c}</strong></td></tr>
    <tr><td>Order Value</td>     <td><strong>{ord_v}</strong></td></tr>
    <tr><td>Total Activities</td><td><strong>{act}</strong></td></tr>
  </table>
  <p>The attached Excel workbook contains:</p>
  <ul>
    <li><strong>Summary Tab</strong> &mdash; Overall totals &amp; team-member comparison</li>
    <li><strong>Owner Tabs</strong> &mdash; Individual performance per sales executive</li>
    <li><strong>Contacts / Leads / Activities Detail</strong> &mdash; Full filtered records</li>
  </ul>
  <p>Best regards,<br><strong style="color:#1e3a8a;">DAC CRM System</strong></p>
<div class="footer">Automated report &mdash; Please do not reply to this email</div>
</div></body></html>""".format(
                plabel=plabel, today=frappe.utils.today(),
                fy=fiscal_year,
                c_rng="{} to {}".format(c_from or "All Time", c_to or "All Time"),
                l_rng="{} to {}".format(l_from or "All Time", l_to or "All Time"),
                tc=c_tot.get("tc",0), oc=c_tot.get("oc",0),
                tl=l_tot.get("tl",0), ord_c=l_tot.get("ord_c",0),
                ord_v=money(l_tot.get("ord_v",0)),
                act=ea_tot.get("total",0))

            frappe.sendmail(
                recipients=recipients,
                subject=report_subject(report_type),
                message=html_body,
                attachments=[{"fname": fname, "fcontent": raw_bytes}]
            )
            email_sent = True

    return {"filename": fname, "filecontent": file_base64, "email_sent": email_sent}


def get_user_stats(uid, report_type="daily"):
    import datetime as _dt
    import calendar as _cal
    _today = frappe.utils.getdate(frappe.utils.today())
    
    if report_type == "daily":
        _d_str = _today.strftime("%Y-%m-%d")
        c_from = c_to = l_from = l_to = ea_from = ea_to = _d_str
    elif report_type == "weekly":
        _wday = _today.weekday()
        _week_start = _today - _dt.timedelta(days=_wday)
        _week_end   = _week_start + _dt.timedelta(days=6)
        c_from = l_from = ea_from = _week_start.strftime("%Y-%m-%d")
        c_to   = l_to   = ea_to   = _week_end.strftime("%Y-%m-%d")
    elif report_type == "monthly":
        _month_start = _today.replace(day=1)
        _last_day = _cal.monthrange(_today.year, _today.month)[1]
        _month_end = _today.replace(day=_last_day)
        c_from = l_from = ea_from = _month_start.strftime("%Y-%m-%d")
        c_to   = l_to   = ea_to   = _month_end.strftime("%Y-%m-%d")
    else:
        c_from = c_to = l_from = l_to = ea_from = ea_to = None

    # Leads created in the period:
    l_p_c = {"usr": uid}
    l_p_wh = "WHERE docstatus < 2 AND lead_owner = %(usr)s"
    if l_from and l_to:
        l_p_wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
        l_p_c["sd"] = l_from; l_p_c["ed"] = l_to
    leads_created = frappe.db.sql(f"SELECT COUNT(*) FROM `tabLead` {l_p_wh}", l_p_c)[0][0] or 0
    newly_created_leads = frappe.db.sql(f"SELECT COUNT(*) FROM `tabLead` {l_p_wh} AND (custom_business_contacts IS NULL OR custom_business_contacts = '')", l_p_c)[0][0] or 0
    converted_leads = frappe.db.sql(f"SELECT COUNT(*) FROM `tabLead` {l_p_wh} AND (custom_business_contacts IS NOT NULL AND custom_business_contacts != '')", l_p_c)[0][0] or 0

    # Contacts created in the period:
    c_p_c = {"usr": uid}
    c_p_wh = "WHERE docstatus < 2 AND assign_to = %(usr)s"
    if c_from and c_to:
        c_p_wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
        c_p_c["sd"] = c_from; c_p_c["ed"] = c_to
    contacts_created = frappe.db.sql(f"SELECT COUNT(*) FROM `tabBusiness Contacts` {c_p_wh}", c_p_c)[0][0] or 0

    # Quotations created in the period:
    q_p_c = {"usr": uid}
    q_p_wh = "WHERE docstatus < 2 AND COALESCE(NULLIF(custom_lead_owner, ''), owner) = %(usr)s"
    if l_from and l_to:
        q_p_wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
        q_p_c["sd"] = l_from; q_p_c["ed"] = l_to
    quotations_created = frappe.db.sql(f"SELECT COUNT(*) FROM `tabQuotation` {q_p_wh}", q_p_c)[0][0] or 0
    q_status_counts = _add_quotation_status_counts({}, q_p_wh, q_p_c)
    q_created_draft = q_status_counts.get("q_st_draft", 0)
    q_created_open = q_status_counts.get("q_st_open", 0)
    q_created_ordered = q_status_counts.get("q_st_ordered", 0)

    # Event activities performed (Completed) breakdown of contact and lead:
    al_p = {"usr": uid}
    al_wh = "WHERE docstatus < 2 AND assigned_to = %(usr)s AND status = 'Completed' AND reference_type = 'Lead'"
    if l_from and l_to:
        al_wh += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
        al_p["sd"] = l_from; al_p["ed"] = l_to
    act_lead = frappe.db.sql(f"SELECT COUNT(*) FROM `tabEvent Activity` {al_wh}", al_p)[0][0] or 0

    ac_p = {"usr": uid}
    ac_wh = "WHERE docstatus < 2 AND assigned_to = %(usr)s AND status = 'Completed' AND reference_type = 'Business Contacts'"
    if c_from and c_to:
        ac_wh += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
        ac_p["sd"] = c_from; ac_p["ed"] = c_to
    act_contact = frappe.db.sql(f"SELECT COUNT(*) FROM `tabEvent Activity` {ac_wh}", ac_p)[0][0] or 0

    # User-wise all-time draft and open quotations:
    qo_p = {"usr": uid}
    draft_quotations = frappe.db.sql("SELECT COUNT(*) FROM `tabQuotation` WHERE docstatus = 0 AND COALESCE(NULLIF(custom_lead_owner, ''), owner) = %(usr)s", qo_p)[0][0] or 0
    open_quotations = frappe.db.sql("SELECT COUNT(*) FROM `tabQuotation` WHERE docstatus = 1 AND status = 'Open' AND COALESCE(NULLIF(custom_lead_owner, ''), owner) = %(usr)s", qo_p)[0][0] or 0
    
    # Calculate conversion metrics for the period
    p_cvd = {"usr": uid}
    wh_cvd = "WHERE docstatus < 2 AND assign_to = %(usr)s AND " + BC_CONVERTED_COND_SQL
    if c_from and c_to:
        wh_cvd += " AND ({d}) BETWEEN %(sd)s AND %(ed)s".format(d=BC_CONVERSION_DATE_SQL)
        p_cvd["sd"] = c_from; p_cvd["ed"] = c_to
    contacts_converted = frappe.db.sql(f"SELECT COUNT(*) FROM `tabBusiness Contacts` {wh_cvd}", p_cvd)[0][0] or 0

    p_ord = {"usr": uid}
    wh_ord = "WHERE l.docstatus < 2 AND l.lead_owner = %(usr)s AND " + LEAD_ORDER_COND_SQL
    if l_from and l_to:
        wh_ord += " AND ({d}) BETWEEN %(sd)s AND %(ed)s".format(d=LEAD_ORDER_CONV_DATE_SQL)
        p_ord["sd"] = l_from; p_ord["ed"] = l_to
    orders_total = frappe.db.sql(f"SELECT COUNT(DISTINCT l.name) FROM `tabLead` l {wh_ord}", p_ord)[0][0] or 0

    c_to_l = round(contacts_converted * 100.0 / contacts_created, 1) if contacts_created else 0.0
    l_to_o = round(orders_total * 100.0 / leads_created, 1) if leads_created else 0.0

    res = {
        "leads_created": leads_created,
        "newly_created_leads": newly_created_leads,
        "converted_leads": converted_leads,
        "contacts_created": contacts_created,
        "contacts_converted": contacts_converted,
        "orders_converted": orders_total,
        "quotations_created": quotations_created,
        "q_created_draft": q_created_draft,
        "q_created_open": q_created_open,
        "q_created_ordered": q_created_ordered,
        "act_lead": act_lead,
        "act_contact": act_contact,
        "draft_quotations": draft_quotations,
        "open_quotations": open_quotations,
        "c_to_l": c_to_l,
        "l_to_o": l_to_o,
        **q_status_counts
    }
    return res


@frappe.whitelist()
def trigger_crm_report(report_type="daily", send_to_admin=0, send_to_team=0, team_member=None):
    """
    Manually trigger / test sending daily, weekly, or monthly report.
    """
    import base64
    frappe.flags.ignore_permissions = True
    settings = frappe.get_single("Admin Settings")
    
    plabel = report_type.title()
    
    # Get all active CRM users
    crm_users = frappe.db.sql("""
        SELECT u.name as user_id, u.full_name as sp_name
        FROM `tabUser` u
        JOIN `tabHas Role` hr ON hr.parent = u.name
        WHERE u.enabled = 1 AND u.user_type = 'System User'
          AND hr.role IN ('DAC CRM Head', 'DAC CRM')
        GROUP BY u.name
        ORDER BY u.full_name ASC
    """, as_dict=True)
    
    admin_field = f"send_{report_type}_crm_report"
    team_field = f"send_{report_type}_crm_report_to_team"

    admin_enabled = bool(int(getattr(settings, admin_field, 0) or 0))
    team_enabled = bool(int(getattr(settings, team_field, 0) or 0))

    messages = []

    # 1. Send to Admin if selected
    if int(send_to_admin):
        if not admin_enabled:
            messages.append(f"Admin report for {report_type} is disabled in Admin Settings.")
        else:
            email_field = f"crm_{report_type}_report_emails"
            admin_emails = getattr(settings, email_field, "")
            if admin_emails:
                recipients = [e.strip() for e in admin_emails.split(",") if e.strip()]
                if recipients:
                    # Gather stats for all users for admin table
                    stats_list = []
                    for user in crm_users:
                        stats = get_user_stats(user["user_id"], report_type)
                        stats["sp_name"] = user["sp_name"]
                        stats_list.append(stats)
                    
                    overall_res = export_crm_dashboard_excel(export_type=report_type, send_email=0)
                    fname = overall_res["filename"]
                    raw_bytes = base64.b64decode(overall_res["filecontent"])
                    html_body = get_crm_report_html_template(report_type, is_individual=False, stats_data=stats_list)
                    
                    frappe.sendmail(
                        recipients=recipients,
                        subject=report_subject(report_type),
                        message=html_body,
                        attachments=[{"fname": fname, "fcontent": raw_bytes}]
                    )
                    messages.append(f"Admin report for {report_type} sent successfully.")
            else:
                messages.append(f"No Admin email recipients configured for {report_type}.")
                
    # 2. Send to specific team member if selected
    if int(send_to_team) and team_member:
        if not team_enabled:
            messages.append(f"Team report for {report_type} is disabled in Admin Settings.")
        else:
            sp_name = frappe.db.get_value("User", team_member, "full_name") or team_member
            stats = get_user_stats(team_member, report_type)
            
            user_res = export_crm_dashboard_excel(
                c_user=team_member, l_user=team_member, ea_user=team_member,
                export_type=report_type, send_email=0
            )
            user_bytes = base64.b64decode(user_res["filecontent"])
            user_fname = f"{sp_name.replace(' ', '_')}_{report_type.title()}_Report_{frappe.utils.today()}.xlsx"
            user_html = get_crm_report_html_template(report_type, is_individual=True, user_name=sp_name, stats_data=stats)
            
            frappe.sendmail(
                recipients=[team_member],
                subject=report_subject(report_type),
                message=user_html,
                attachments=[{"fname": user_fname, "fcontent": user_bytes}]
            )
            messages.append(f"Team member report for {report_type} sent successfully to {team_member}.")
        
    frappe.flags.ignore_permissions = False
    status_type = "success" if any("sent successfully" in m for m in messages) else "skipped"
    return {"status": status_type, "message": " ".join(messages) if messages else "No actions performed."}


def send_scheduled_crm_report(report_type="daily", force=False):
    """
    Generates and sends the daily, weekly, or monthly report to Admin and Individual team members.
    """
    import base64
    frappe.flags.ignore_permissions = True
    settings = frappe.get_single("Admin Settings")
    
    # Check if sending is enabled
    admin_field = f"send_{report_type}_crm_report"
    team_field = f"send_{report_type}_crm_report_to_team"

    admin_enabled = bool(int(getattr(settings, admin_field, 0) or 0))
    team_enabled = bool(int(getattr(settings, team_field, 0) or 0))

    if not admin_enabled and not team_enabled and not force:
        return {"status": "skipped", "message": f"{report_type.title()} reports are disabled in Admin Settings."}
        
    email_field = f"crm_{report_type}_report_emails"
    admin_emails = getattr(settings, email_field, "")
    
    # Get all active CRM users
    crm_users = frappe.db.sql("""
        SELECT u.name as user_id, u.full_name as sp_name
        FROM `tabUser` u
        JOIN `tabHas Role` hr ON hr.parent = u.name
        WHERE u.enabled = 1 AND u.user_type = 'System User'
          AND hr.role IN ('DAC CRM Head', 'DAC CRM')
        GROUP BY u.name
        ORDER BY u.full_name ASC
    """, as_dict=True)
    
    # Send to Admin overall if emails are configured
    if admin_emails and (admin_enabled or force):
        recipients = [e.strip() for e in admin_emails.split(",") if e.strip()]
        if recipients:
            stats_list = []
            for user in crm_users:
                stats = get_user_stats(user["user_id"], report_type)
                stats["sp_name"] = user["sp_name"]
                stats_list.append(stats)
                
            overall_res = export_crm_dashboard_excel(export_type=report_type, send_email=0)
            fname = overall_res["filename"]
            raw_bytes = base64.b64decode(overall_res["filecontent"])
            
            plabel = report_type.title()
            html_body = get_crm_report_html_template(report_type, is_individual=False, stats_data=stats_list)
            
            frappe.sendmail(
                recipients=recipients,
                subject=report_subject(report_type),
                message=html_body,
                attachments=[{"fname": fname, "fcontent": raw_bytes}]
            )
            
    # Send a personal report to each CRM team member — but only to those who
    # actually did something in the period. An all-zero report is noise, and it
    # trains people to ignore the mail.
    send_to_team = team_enabled or force

    skipped_no_activity = []
    for user in (crm_users if send_to_team else []):
        user_email = user["user_id"]
        # Generate user-filtered report
        try:
            stats = get_user_stats(user_email, report_type)

            if not _has_period_activity(stats):
                skipped_no_activity.append(user_email)
                continue

            user_res = export_crm_dashboard_excel(
                c_user=user_email, l_user=user_email, ea_user=user_email,
                export_type=report_type, send_email=0
            )
            user_bytes = base64.b64decode(user_res["filecontent"])
            user_fname = f"{user['sp_name'].replace(' ', '_')}_{report_type.title()}_Report_{frappe.utils.today()}.xlsx"
            
            user_html = get_crm_report_html_template(report_type, is_individual=True, user_name=user["sp_name"], stats_data=stats)
            
            frappe.sendmail(
                recipients=[user_email],
                subject=report_subject(report_type),
                message=user_html,
                attachments=[{"fname": user_fname, "fcontent": user_bytes}]
            )
        except Exception as e:
            frappe.log_error(title=f"Failed to send individual {report_type} report to {user_email}", message=frappe.get_traceback())
            
    frappe.flags.ignore_permissions = False

    msg = "{} report sent for {}.".format(report_type.title(), report_period_label(report_type))
    if not send_to_team:
        msg += " Team member reports are switched off in Admin Settings."
    elif skipped_no_activity:
        msg += " Skipped {} member(s) with no activity.".format(len(skipped_no_activity))
    return {
        "status": "success",
        "message": msg,
        "sent_to_team": send_to_team,
        "skipped": skipped_no_activity,
    }


def _has_period_activity(stats):
    """
    Did this person do anything in the period worth mailing about?
    Counts only work done, not all-time balances.
    """
    stats = stats or {}
    keys = ("leads_created", "contacts_created", "quotations_created",
            "act_lead", "act_contact", "act_done", "act_total")
    return any(flt(stats.get(k, 0)) for k in keys)


def get_all_time_stats(uid):
    leads_created = frappe.db.sql("SELECT COUNT(*) FROM `tabLead` WHERE docstatus < 2 AND lead_owner = %s", uid)[0][0] or 0
    newly_created_leads = frappe.db.sql("SELECT COUNT(*) FROM `tabLead` WHERE docstatus < 2 AND lead_owner = %s AND (custom_business_contacts IS NULL OR custom_business_contacts = '')", uid)[0][0] or 0
    converted_leads = frappe.db.sql("SELECT COUNT(*) FROM `tabLead` WHERE docstatus < 2 AND lead_owner = %s AND (custom_business_contacts IS NOT NULL AND custom_business_contacts != '')", uid)[0][0] or 0
    contacts_created = frappe.db.sql("SELECT COUNT(*) FROM `tabBusiness Contacts` WHERE docstatus < 2 AND assign_to = %s", uid)[0][0] or 0
    
    q_p_wh = "WHERE docstatus < 2 AND COALESCE(NULLIF(custom_lead_owner, ''), owner) = %s"
    quotations_created = frappe.db.sql(f"SELECT COUNT(*) FROM `tabQuotation` {q_p_wh}", uid)[0][0] or 0
    q_status_counts = _add_quotation_status_counts({}, q_p_wh, uid)
    q_created_draft = q_status_counts.get("q_st_draft", 0)
    q_created_open = q_status_counts.get("q_st_open", 0)
    q_created_ordered = q_status_counts.get("q_st_ordered", 0)
    
    # "Activities created" is deliberately not reported — only completed work is.
    
    act_lead = frappe.db.sql("SELECT COUNT(*) FROM `tabEvent Activity` WHERE docstatus < 2 AND assigned_to = %s AND status = 'Completed' AND reference_type = 'Lead'", uid)[0][0] or 0
    act_contact = frappe.db.sql("SELECT COUNT(*) FROM `tabEvent Activity` WHERE docstatus < 2 AND assigned_to = %s AND status = 'Completed' AND reference_type = 'Business Contacts'", uid)[0][0] or 0
    draft_quotations = frappe.db.sql("SELECT COUNT(*) FROM `tabQuotation` WHERE docstatus = 0 AND COALESCE(NULLIF(custom_lead_owner, ''), owner) = %s", uid)[0][0] or 0
    open_quotations = frappe.db.sql("SELECT COUNT(*) FROM `tabQuotation` WHERE docstatus = 1 AND status = 'Open' AND COALESCE(NULLIF(custom_lead_owner, ''), owner) = %s", uid)[0][0] or 0
    
    contacts_converted = frappe.db.sql("SELECT COUNT(*) FROM `tabBusiness Contacts` WHERE docstatus < 2 AND assign_to = %s AND status = 'Converted to Lead'", uid)[0][0] or 0
    orders_total = frappe.db.sql("SELECT COUNT(*) FROM `tabLead` WHERE docstatus < 2 AND lead_owner = %s AND custom_lead_category = 'Order'", uid)[0][0] or 0

    c_to_l = round(contacts_converted * 100.0 / contacts_created, 1) if contacts_created else 0.0
    l_to_o = round(orders_total * 100.0 / leads_created, 1) if leads_created else 0.0

    res = {
        "leads_created": leads_created,
        "newly_created_leads": newly_created_leads,
        "converted_leads": converted_leads,
        "contacts_created": contacts_created,
        "contacts_converted": contacts_converted,
        "orders_converted": orders_total,
        "quotations_created": quotations_created,
        "q_created_draft": q_created_draft,
        "q_created_open": q_created_open,
        "q_created_ordered": q_created_ordered,
        "act_lead": act_lead,
        "act_contact": act_contact,
        "draft_quotations": draft_quotations,
        "open_quotations": open_quotations,
        "c_to_l": c_to_l,
        "l_to_o": l_to_o,
        **q_status_counts
    }
    return res


def get_user_stats_by_date_range(uid, start_date=None, end_date=None):
    # Leads
    l_p_c = {"usr": uid}
    l_p_wh = "WHERE docstatus < 2 AND lead_owner = %(usr)s"
    if start_date and end_date:
        l_p_wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
        l_p_c["sd"] = start_date; l_p_c["ed"] = end_date
    leads_created = frappe.db.sql(f"SELECT COUNT(*) FROM `tabLead` {l_p_wh}", l_p_c)[0][0] or 0
    
    # Contacts
    c_p_c = {"usr": uid}
    c_p_wh = "WHERE docstatus < 2 AND assign_to = %(usr)s"
    if start_date and end_date:
        c_p_wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
        c_p_c["sd"] = start_date; c_p_c["ed"] = end_date
    contacts_created = frappe.db.sql(f"SELECT COUNT(*) FROM `tabBusiness Contacts` {c_p_wh}", c_p_c)[0][0] or 0
    
    # Quotations
    q_p_c = {"usr": uid}
    q_p_wh = "WHERE docstatus < 2 AND COALESCE(NULLIF(custom_lead_owner, ''), owner) = %(usr)s"
    if start_date and end_date:
        q_p_wh += " AND DATE(creation) BETWEEN %(sd)s AND %(ed)s"
        q_p_c["sd"] = start_date; q_p_c["ed"] = end_date
    quotations_created = frappe.db.sql(f"SELECT COUNT(*) FROM `tabQuotation` {q_p_wh}", q_p_c)[0][0] or 0
    
    # Activities (Completed)
    al_p = {"usr": uid}
    al_wh = "WHERE docstatus < 2 AND assigned_to = %(usr)s AND status = 'Completed' AND reference_type = 'Lead'"
    if start_date and end_date:
        al_wh += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
        al_p["sd"] = start_date; al_p["ed"] = end_date
    act_lead = frappe.db.sql(f"SELECT COUNT(*) FROM `tabEvent Activity` {al_wh}", al_p)[0][0] or 0
    
    ac_p = {"usr": uid}
    ac_wh = "WHERE docstatus < 2 AND assigned_to = %(usr)s AND status = 'Completed' AND reference_type = 'Business Contacts'"
    if start_date and end_date:
        ac_wh += " AND DATE(COALESCE(ends_on, starts_on)) BETWEEN %(sd)s AND %(ed)s"
        ac_p["sd"] = start_date; ac_p["ed"] = end_date
    act_contact = frappe.db.sql(f"SELECT COUNT(*) FROM `tabEvent Activity` {ac_wh}", ac_p)[0][0] or 0
    
    # Conversions
    p_cvd = {"usr": uid}
    wh_cvd = "WHERE docstatus < 2 AND assign_to = %(usr)s AND " + BC_CONVERTED_COND_SQL
    if start_date and end_date:
        wh_cvd += " AND ({d}) BETWEEN %(sd)s AND %(ed)s".format(d=BC_CONVERSION_DATE_SQL)
        p_cvd["sd"] = start_date; p_cvd["ed"] = end_date
    contacts_converted = frappe.db.sql(f"SELECT COUNT(*) FROM `tabBusiness Contacts` {wh_cvd}", p_cvd)[0][0] or 0
    
    p_ord = {"usr": uid}
    wh_ord = "WHERE l.docstatus < 2 AND l.lead_owner = %(usr)s AND " + LEAD_ORDER_COND_SQL
    if start_date and end_date:
        wh_ord += " AND ({d}) BETWEEN %(sd)s AND %(ed)s".format(d=LEAD_ORDER_CONV_DATE_SQL)
        p_ord["sd"] = start_date; p_ord["ed"] = end_date
    orders_total = frappe.db.sql(f"SELECT COUNT(DISTINCT l.name) FROM `tabLead` l {wh_ord}", p_ord)[0][0] or 0
    
    c_to_l = round(contacts_converted * 100.0 / contacts_created, 1) if contacts_created else 0.0
    l_to_o = round(orders_total * 100.0 / leads_created, 1) if leads_created else 0.0
    
    return {
        "leads_created": leads_created,
        "contacts_created": contacts_created,
        "quotations_created": quotations_created,
        "act_done": act_lead + act_contact,
        "c_to_l": c_to_l,
        "l_to_o": l_to_o,
        "contacts_converted": contacts_converted,
        "orders_converted": orders_total
    }


def get_comparison_matrix_html(uids):
    if not isinstance(uids, list):
        uids = [uids]
        
    import datetime as _dt
    import calendar as _cal
    _today = frappe.utils.getdate(frappe.utils.today())
    
    # 1. Fiscal Year (Overall) Dates
    if _today.month >= 4:
        fy_start_year = _today.year
        fy_end_year = _today.year + 1
    else:
        fy_start_year = _today.year - 1
        fy_end_year = _today.year
    fy_start_str = f"{fy_start_year}-04-01"
    fy_end_str = f"{fy_end_year}-03-31"
    
    # Format labels in dd-mmm-yyyy format
    fy_start_label = f"01-Apr-{fy_start_year}"
    fy_end_label = f"31-Mar-{fy_end_year}"
    
    # 2. Prev Month Dates
    first_this_month = _today.replace(day=1)
    last_prev_month = first_this_month - _dt.timedelta(days=1)
    first_prev_month = last_prev_month.replace(day=1)
    prev_month_start = first_prev_month.strftime("%Y-%m-%d")
    prev_month_end = last_prev_month.strftime("%Y-%m-%d")
    
    # 3. This Month Dates
    this_month_start = _today.replace(day=1).strftime("%Y-%m-%d")
    this_month_end = _today.strftime("%Y-%m-%d")
    
    # 4. Week (Start-Fri) Dates
    _wday = _today.weekday()
    week_start = _today - _dt.timedelta(days=_wday)
    week_friday = week_start + _dt.timedelta(days=4)
    week_start_str = week_start.strftime("%Y-%m-%d")
    week_friday_str = week_friday.strftime("%Y-%m-%d")
    
    # Helper to sum stats for all users in a period
    def get_combined_stats(start, end):
        combined = {
            "leads_created": 0, "contacts_created": 0, "quotations_created": 0, "act_done": 0,
            "contacts_converted": 0, "orders_converted": 0
        }
        for uid in uids:
            s = get_user_stats_by_date_range(uid, start, end)
            for k in combined:
                combined[k] += s.get(k, 0)
        # Calculate rates
        combined["c_to_l"] = round(combined["contacts_converted"] * 100.0 / combined["contacts_created"], 1) if combined["contacts_created"] else 0.0
        combined["l_to_o"] = round(combined["orders_converted"] * 100.0 / combined["leads_created"], 1) if combined["leads_created"] else 0.0
        return combined

    overall = get_combined_stats(fy_start_str, fy_end_str)
    prev_month = get_combined_stats(prev_month_start, prev_month_end)
    this_month = get_combined_stats(this_month_start, this_month_end)
    week = get_combined_stats(week_start_str, week_friday_str)
    
    # Return HTML table
    return """
    <h3 style="margin-top: 25px; margin-bottom: 10px; color: #0f172a; border-bottom: 2px solid #3b82f6; padding-bottom: 5px; font-size: 15px;">
        📊 CRM Performance Matrix
    </h3>
    <div style="overflow-x: auto; margin: 15px 0; border: 1px solid #e2e8f0; border-radius: 8px;">
        <table style="width: 100%; border-collapse: collapse; font-size: 12px; font-family: sans-serif;">
            <thead>
                <tr style="background-color: #0f172a; color: #ffffff;">
                    <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: left; font-weight: 600;">Metric</th>
                    <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Overall<br><span style="font-size: 10px; font-weight: normal; color: #cbd5e1;">{fy_start} to {fy_end}</span></th>
                    <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Prev Month</th>
                    <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">This Month</th>
                    <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Week (Start-Fri)</th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; font-weight: 600; color: #334155;">Leads Created</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;"><strong>{o[leads_created]}</strong></td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{pm[leads_created]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{tm[leads_created]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{wk[leads_created]}</td>
                </tr>
                <tr>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; font-weight: 600; color: #334155;">Contacts Created</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;"><strong>{o[contacts_created]}</strong></td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{pm[contacts_created]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{tm[contacts_created]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{wk[contacts_created]}</td>
                </tr>
                <tr>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; font-weight: 600; color: #334155;">Quotations Created</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;"><strong>{o[quotations_created]}</strong></td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{pm[quotations_created]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{tm[quotations_created]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{wk[quotations_created]}</td>
                </tr>
                <tr>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; font-weight: 600; color: #334155;">Activities Completed</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;"><strong>{o[act_done]}</strong></td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{pm[act_done]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{tm[act_done]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{wk[act_done]}</td>
                </tr>
                <tr>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; font-weight: 600; color: #334155;">Contact to Lead</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;"><strong>{o[contacts_converted]}</strong></td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{pm[contacts_converted]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{tm[contacts_converted]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{wk[contacts_converted]}</td>
                </tr>
                <tr>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; font-weight: 600; color: #334155;">Lead to Order</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;"><strong>{o[orders_converted]}</strong></td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{pm[orders_converted]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{tm[orders_converted]}</td>
                    <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{wk[orders_converted]}</td>
                </tr>
            </tbody>
        </table>
    </div>
    """.format(fy_start=fy_start_label, fy_end=fy_end_label, o=overall, pm=prev_month, tm=this_month, wk=week)


def get_crm_report_html_template(report_type, is_individual=False, user_name="", stats_data=None):
    plabel = report_type.title()
    title = "My {} Performance Report".format(plabel) if is_individual else "Overall {} CRM Executive Report".format(plabel)
    greeting = "Dear {},".format(user_name) if is_individual else ""
    today_str = _fmt_dmy(frappe.utils.today())
    period_range = report_period_label(report_type)
    
    stats_html = ""
    if is_individual and stats_data:
        user_email = frappe.db.get_value("User", {"full_name": user_name}, "name") or frappe.session.user
        overall_stats = get_all_time_stats(user_email)
        
        # Grid 1: Period
        grid_period = """
        <h3 style="margin-top: 25px; margin-bottom: 10px; color: #1e3a8a; border-bottom: 2px solid #3b82f6; padding-bottom: 5px; font-size: 15px;">
            📅 Performance ({period_range})
        </h3>
        <div style="margin: 15px 0; display: grid; grid-template-columns: repeat(2, 1fr); gap: 12px;">
            <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center;">
                <div style="font-size: 11px; font-weight: 600; color: #64748b; text-transform: uppercase;">Leads Created</div>
                <div style="font-size: 20px; font-weight: 700; color: #1e3a8a; margin-top: 4px;">{leads}</div>
                <div style="font-size: 11px; color: #64748b; margin-top: 3px;">New: {new_leads} | Conv: {converted_leads}</div>
            </div>
            <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center;">
                <div style="font-size: 11px; font-weight: 600; color: #64748b; text-transform: uppercase;">Contacts Created</div>
                <div style="font-size: 20px; font-weight: 700; color: #1e3a8a; margin-top: 4px;">{contacts}</div>
            </div>
            <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center;">
                <div style="font-size: 11px; font-weight: 600; color: #64748b; text-transform: uppercase;">Quotations Created</div>
                <div style="font-size: 20px; font-weight: 700; color: #1e3a8a; margin-top: 4px;">{quotations}</div>
                <div style="font-size: 11px; color: #64748b; margin-top: 3px;">{q_status_line}</div>
            </div>
            <div style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 12px; text-align: center;">
                <div style="font-size: 11px; font-weight: 600; color: #64748b; text-transform: uppercase;">Activities Completed</div>
                <div style="font-size: 20px; font-weight: 700; color: #16a34a; margin-top: 4px;">{act_done_total}</div>
                <div style="font-size: 11px; color: #64748b; margin-top: 3px;">Lead: {act_lead} | Contact: {act_contact}</div>
            </div>
            <div style="background-color: #fdf4ff; border: 1px solid #f5d0fe; border-radius: 8px; padding: 12px; text-align: center; grid-column: span 2;">
                <div style="font-size: 11px; font-weight: 600; color: #86198f; text-transform: uppercase;">Conversion Metrics</div>
                <div style="font-size: 13px; font-weight: 700; color: #701a75; margin-top: 6px;">
                    Contact to Lead: {c_conv} &nbsp;|&nbsp; Lead to Order: {o_conv}
                </div>
            </div>
        </div>
        """.format(
            period_range=period_range,
            leads=stats_data.get('leads_created', 0),
            new_leads=stats_data.get('newly_created_leads', 0),
            converted_leads=stats_data.get('converted_leads', 0),
            contacts=stats_data.get('contacts_created', 0),
            quotations=stats_data.get('quotations_created', 0),
            q_status_line=_quotation_status_line(stats_data),
            act_done_total=(stats_data.get('act_lead', 0) or 0) + (stats_data.get('act_contact', 0) or 0),
            act_lead=stats_data.get('act_lead', 0),
            act_contact=stats_data.get('act_contact', 0),
            c_to_l=stats_data.get('c_to_l', 0),
            l_to_o=stats_data.get('l_to_o', 0),
            c_conv=stats_data.get('contacts_converted', 0),
            o_conv=stats_data.get('orders_converted', 0)
        )
        
        # Grid 2: Overall
        grid_overall = """
        <h3 style="margin-top: 25px; margin-bottom: 10px; color: #0f172a; border-bottom: 2px solid #94a3b8; padding-bottom: 5px; font-size: 15px;">
            📊 Overall All Performance
        </h3>
        <div style="margin: 15px 0; display: grid; grid-template-columns: repeat(2, 1fr); gap: 12px;">
            <div style="background-color: #f1f5f9; border: 1px solid #cbd5e1; border-radius: 8px; padding: 12px; text-align: center;">
                <div style="font-size: 11px; font-weight: 600; color: #475569; text-transform: uppercase;">Total Leads Created</div>
                <div style="font-size: 20px; font-weight: 700; color: #0f172a; margin-top: 4px;">{leads}</div>
                <div style="font-size: 11px; color: #64748b; margin-top: 3px;">New: {new_leads} | Conv: {converted_leads}</div>
            </div>
            <div style="background-color: #f1f5f9; border: 1px solid #cbd5e1; border-radius: 8px; padding: 12px; text-align: center;">
                <div style="font-size: 11px; font-weight: 600; color: #475569; text-transform: uppercase;">Total Contacts Created</div>
                <div style="font-size: 20px; font-weight: 700; color: #0f172a; margin-top: 4px;">{contacts}</div>
            </div>
            <div style="background-color: #f1f5f9; border: 1px solid #cbd5e1; border-radius: 8px; padding: 12px; text-align: center;">
                <div style="font-size: 11px; font-weight: 600; color: #475569; text-transform: uppercase;">Total Quotations Created</div>
                <div style="font-size: 20px; font-weight: 700; color: #0f172a; margin-top: 4px;">{quotations}</div>
                <div style="font-size: 11px; color: #475569; margin-top: 3px;">{q_status_line}</div>
            </div>
            <div style="background-color: #f1f5f9; border: 1px solid #cbd5e1; border-radius: 8px; padding: 12px; text-align: center;">
                <div style="font-size: 11px; font-weight: 600; color: #475569; text-transform: uppercase;">Total Activities Completed</div>
                <div style="font-size: 20px; font-weight: 700; color: #16a34a; margin-top: 4px;">{act_done_total}</div>
                <div style="font-size: 11px; color: #475569; margin-top: 3px;">Lead: {act_lead} | Contact: {act_contact}</div>
            </div>
            <div style="background-color: #fffbeb; border: 1px solid #fef3c7; border-radius: 8px; padding: 12px; text-align: center; grid-column: span 2;">
                <div style="font-size: 11px; font-weight: 600; color: #b45309; text-transform: uppercase;">Conversion Metrics</div>
                <div style="font-size: 13px; font-weight: 700; color: #78350f; margin-top: 6px;">
                    Contact to Lead: {c_conv} &nbsp;|&nbsp; Lead to Order: {o_conv}
                </div>
            </div>
            <div style="background-color: #fffbeb; border: 1px solid #fef3c7; border-radius: 8px; padding: 12px; text-align: center; grid-column: span 2;">
                <div style="font-size: 11px; font-weight: 600; color: #b45309; text-transform: uppercase;">Quotations Status Tracking</div>
                <div style="font-size: 15px; font-weight: 700; color: #78350f; margin-top: 6px; display: flex; justify-content: space-around;">
                    <span>Draft: <strong>{draft}</strong></span>
                    <span>Open / Submitted: <strong>{open}</strong></span>
                </div>
            </div>
        </div>
        """.format(
            leads=overall_stats.get('leads_created', 0),
            new_leads=overall_stats.get('newly_created_leads', 0),
            converted_leads=overall_stats.get('converted_leads', 0),
            contacts=overall_stats.get('contacts_created', 0),
            quotations=overall_stats.get('quotations_created', 0),
            q_status_line=_quotation_status_line(overall_stats),
            act_done_total=(overall_stats.get('act_lead', 0) or 0) + (overall_stats.get('act_contact', 0) or 0),
            act_lead=overall_stats.get('act_lead', 0),
            act_contact=overall_stats.get('act_contact', 0),
            draft=overall_stats.get('draft_quotations', 0),
            open=overall_stats.get('open_quotations', 0),
            c_to_l=overall_stats.get('c_to_l', 0),
            l_to_o=overall_stats.get('l_to_o', 0),
            c_conv=overall_stats.get('contacts_converted', 0),
            o_conv=overall_stats.get('orders_converted', 0)
        )
        comparison_matrix_html = get_comparison_matrix_html(user_email)
        stats_html = grid_period + comparison_matrix_html + grid_overall
        
    elif not is_individual and stats_data:
        # Admin Overall: render Period Table AND Overall Table separately!
        
        # 1. Period Table
        rows_period = ""
        total_p_leads = total_p_new_leads = total_p_converted_leads = total_p_contacts = total_p_quotations = total_p_q_draft = total_p_q_open = total_p_q_ordered = 0
        total_p_act_lead = total_p_act_contact = 0
        total_p_rows = 0
        total_p_contacts_converted = total_p_orders_converted = 0
        total_p_status = {}
        for s in stats_data:
            if (s.get('leads_created', 0) == 0 and 
                s.get('contacts_created', 0) == 0 and 
                s.get('quotations_created', 0) == 0 and
                s.get('act_lead', 0) == 0 and 
                s.get('act_contact', 0) == 0):
                continue
            total_p_leads += s.get('leads_created', 0)
            total_p_new_leads += s.get('newly_created_leads', 0)
            total_p_converted_leads += s.get('converted_leads', 0)
            total_p_contacts += s.get('contacts_created', 0)
            total_p_contacts_converted += s.get('contacts_converted', 0)
            total_p_orders_converted += s.get('orders_converted', 0)
            total_p_quotations += s.get('quotations_created', 0)
            total_p_q_draft += s.get('q_created_draft', 0)
            total_p_q_open += s.get('q_created_open', 0)
            total_p_q_ordered += s.get('q_created_ordered', 0)
            for _st in QUOTATION_STATUSES:
                _k = _q_status_key(_st)
                total_p_status[_k] = total_p_status.get(_k, 0) + (s.get(_k, 0) or 0)
            total_p_rows += 1
            total_p_act_lead += s.get('act_lead', 0)
            total_p_act_contact += s.get('act_contact', 0)
            
            rows_period += """
            <tr>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; font-weight: 600; color: #334155;">{sp_name}</td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">
                    <strong>{leads}</strong><br>
                    <span style="font-size: 10px; color: #64748b;">(New: {new_leads} | Conv: {converted_leads})</span>
                </td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{contacts}</td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">
                    <strong>{quotations}</strong><br>
                    <span style="font-size: 10px; color: #64748b;">{q_status_line}</span>
                </td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">
                    <strong>{act_done}</strong><br>
                    <span style="font-size: 10px; color: #64748b;">(Lead: {act_lead} | Contact: {act_contact})</span>
                </td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">
                    <span style="font-size: 11px;">BC&#8594;Lead: <strong>{c_conv}</strong></span><br>
                    <span style="font-size: 11px;">Lead&#8594;Order: <strong>{o_conv}</strong></span>
                </td>
            </tr>
            """.format(
                sp_name=s.get('sp_name', '-'),
                leads=s.get('leads_created', 0),
                new_leads=s.get('newly_created_leads', 0),
                converted_leads=s.get('converted_leads', 0),
                contacts=s.get('contacts_created', 0),
                quotations=s.get('quotations_created', 0),
                q_status_line=_quotation_status_line(s),
                act_done=s.get('act_lead', 0) + s.get('act_contact', 0),
                act_lead=s.get('act_lead', 0),
                act_contact=s.get('act_contact', 0),
                c_to_l=s.get('c_to_l', 0),
                l_to_o=s.get('l_to_o', 0),
                c_conv=s.get('contacts_converted', 0),
                o_conv=s.get('orders_converted', 0)
            )
            
        rows_period += """
        <tr style="background-color: #f1f5f9; font-weight: bold;">
            <td style="padding: 10px; border: 1px solid #e2e8f0; color: #0f172a;">TOTAL</td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #0f172a;">
                <strong>{leads}</strong><br>
                <span style="font-size: 10px; color: #0f172a;">(New: {new_leads} | Conv: {converted_leads})</span>
            </td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #0f172a;">{contacts}</td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #0f172a;">
                <strong>{quotations}</strong><br>
                <span style="font-size: 10px; color: #0f172a;">{q_status_line}</span>
            </td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #0f172a;">
                <strong>{act_done}</strong><br>
                <span style="font-size: 10px; color: #0f172a;">(Lead: {act_lead} | Contact: {act_contact})</span>
            </td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #0f172a;">
                <span style="font-size: 11px;">BC&#8594;Lead: <strong>{c_conv}</strong></span><br>
                <span style="font-size: 11px;">Lead&#8594;Order: <strong>{o_conv}</strong></span>
            </td>
        </tr>
        """.format(
            leads=total_p_leads,
            new_leads=total_p_new_leads,
            converted_leads=total_p_converted_leads,
            contacts=total_p_contacts,
            quotations=total_p_quotations,
            q_status_line=_quotation_status_line(total_p_status),
            act_done=total_p_act_lead + total_p_act_contact,
            act_lead=total_p_act_lead,
            act_contact=total_p_act_contact,
            c_conv=total_p_contacts_converted,
            o_conv=total_p_orders_converted,
            c_to_l=round(total_p_contacts_converted * 100.0 / total_p_contacts, 1) if total_p_contacts else 0.0,
            l_to_o=round(total_p_orders_converted * 100.0 / total_p_leads, 1) if total_p_leads else 0.0
        )
        
        table_period = """
        <h3 style="margin-top: 25px; margin-bottom: 10px; color: #1e3a8a; border-bottom: 2px solid #3b82f6; padding-bottom: 5px; font-size: 15px;">
            📅 Performance Summary ({period_range})
        </h3>
        <div style="overflow-x: auto; margin: 15px 0; border: 1px solid #e2e8f0; border-radius: 8px;">
            <table style="width: 100%; border-collapse: collapse; font-size: 12px;">
                <thead>
                    <tr style="background-color: #1e3a8a; color: #ffffff;">
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: left; font-weight: 600;">Team Member</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Leads</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Contacts</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Quotations Created</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Act. Completed</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Conversion Rates</th>
                    </tr>
                </thead>
                <tbody>
                    {rows}
                </tbody>
            </table>
        </div>
        """.format(period_range=period_range, rows=rows_period)
        
        # 2. Overall Table
        crm_users = frappe.db.sql("""
            SELECT u.name as user_id, u.full_name as sp_name
            FROM `tabUser` u
            JOIN `tabHas Role` hr ON hr.parent = u.name
            WHERE u.enabled = 1 AND u.user_type = 'System User'
              AND hr.role IN ('DAC CRM Head', 'DAC CRM')
            GROUP BY u.name
            ORDER BY u.full_name ASC
        """, as_dict=True)
        
        rows_overall = ""
        total_o_leads = total_o_new_leads = total_o_converted_leads = total_o_contacts = total_o_quotations = total_o_q_draft = total_o_q_open = total_o_q_ordered = 0
        total_o_act_lead = total_o_act_contact = total_o_draft = total_o_open = 0
        total_o_rows = 0
        total_o_contacts_converted = total_o_orders_converted = 0
        total_o_status = {}
        for user in crm_users:
            o_stats = get_all_time_stats(user["user_id"])
            if (o_stats["leads_created"] == 0 and 
                o_stats["contacts_created"] == 0 and 
                o_stats["quotations_created"] == 0 and 
                o_stats["act_lead"] == 0 and 
                o_stats["act_contact"] == 0):
                continue
            total_o_leads += o_stats["leads_created"]
            total_o_new_leads += o_stats.get("newly_created_leads", 0)
            total_o_converted_leads += o_stats.get("converted_leads", 0)
            total_o_contacts += o_stats["contacts_created"]
            total_o_contacts_converted += o_stats.get("contacts_converted", 0)
            total_o_orders_converted += o_stats.get("orders_converted", 0)
            total_o_quotations += o_stats["quotations_created"]
            total_o_q_draft += o_stats["q_created_draft"]
            total_o_q_open += o_stats["q_created_open"]
            total_o_q_ordered += o_stats["q_created_ordered"]
            for _st in QUOTATION_STATUSES:
                _k = _q_status_key(_st)
                total_o_status[_k] = total_o_status.get(_k, 0) + (o_stats.get(_k, 0) or 0)
            total_o_rows += 1
            total_o_act_lead += o_stats["act_lead"]
            total_o_act_contact += o_stats["act_contact"]
            total_o_draft += o_stats["draft_quotations"]
            total_o_open += o_stats["open_quotations"]
            
            rows_overall += """
            <tr>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; font-weight: 600; color: #334155;">{sp_name}</td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">
                    <strong>{leads}</strong><br>
                    <span style="font-size: 10px; color: #64748b;">(New: {new_leads} | Conv: {converted_leads})</span>
                </td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">{contacts}</td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">
                    <strong>{quotations}</strong><br>
                    <span style="font-size: 10px; color: #64748b;">{q_status_line}</span>
                </td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">
                    <strong>{act_done}</strong><br>
                    <span style="font-size: 10px; color: #64748b;">(Lead: {act_lead} | Contact: {act_contact})</span>
                </td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #334155;">
                    <span style="font-size: 11px;">BC&#8594;Lead: <strong>{c_conv}</strong></span><br>
                    <span style="font-size: 11px;">Lead&#8594;Order: <strong>{o_conv}</strong></span>
                </td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #b45309; font-weight: 600;">{draft}</td>
                <td style="padding: 8px 10px; border: 1px solid #e2e8f0; text-align: center; color: #1e3a8a; font-weight: 600;">{open}</td>
            </tr>
            """.format(
                sp_name=user["sp_name"],
                leads=o_stats["leads_created"],
                new_leads=o_stats.get("newly_created_leads", 0),
                converted_leads=o_stats.get("converted_leads", 0),
                contacts=o_stats["contacts_created"],
                quotations=o_stats["quotations_created"],
                q_status_line=_quotation_status_line(o_stats),
                act_done=o_stats["act_lead"] + o_stats["act_contact"],
                act_lead=o_stats["act_lead"],
                act_contact=o_stats["act_contact"],
                c_to_l=o_stats.get("c_to_l", 0),
                l_to_o=o_stats.get("l_to_o", 0),
                c_conv=o_stats.get('contacts_converted', 0),
                o_conv=o_stats.get('orders_converted', 0),
                draft=o_stats["draft_quotations"],
                open=o_stats["open_quotations"]
            )
            
        rows_overall += """
        <tr style="background-color: #f1f5f9; font-weight: bold;">
            <td style="padding: 10px; border: 1px solid #e2e8f0; color: #0f172a;">TOTAL</td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #0f172a;">
                <strong>{leads}</strong><br>
                <span style="font-size: 10px; color: #0f172a;">(New: {new_leads} | Conv: {converted_leads})</span>
            </td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #0f172a;">{contacts}</td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #0f172a;">
                <strong>{quotations}</strong><br>
                <span style="font-size: 10px; color: #0f172a;">{q_status_line}</span>
            </td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #0f172a;">
                <strong>{act_done}</strong><br>
                <span style="font-size: 10px; color: #0f172a;">(Lead: {act_lead} | Contact: {act_contact})</span>
            </td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #0f172a;">
                <span style="font-size: 11px;">BC&#8594;Lead: <strong>{c_conv}</strong></span><br>
                <span style="font-size: 11px;">Lead&#8594;Order: <strong>{o_conv}</strong></span>
            </td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #b45309;">{draft}</td>
            <td style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; color: #1e3a8a;">{open}</td>
        </tr>
        """.format(
            leads=total_o_leads,
            new_leads=total_o_new_leads,
            converted_leads=total_o_converted_leads,
            contacts=total_o_contacts,
            quotations=total_o_quotations,
            q_status_line=_quotation_status_line(total_o_status),
            act_done=total_o_act_lead + total_o_act_contact,
            act_lead=total_o_act_lead,
            act_contact=total_o_act_contact,
            c_conv=total_o_contacts_converted,
            o_conv=total_o_orders_converted,
            c_to_l=round(total_o_contacts_converted * 100.0 / total_o_contacts, 1) if total_o_contacts else 0.0,
            l_to_o=round(total_o_orders_converted * 100.0 / total_o_leads, 1) if total_o_leads else 0.0,
            draft=total_o_draft,
            open=total_o_open
        )
        
        table_overall = """
        <h3 style="margin-top: 25px; margin-bottom: 10px; color: #0f172a; border-bottom: 2px solid #94a3b8; padding-bottom: 5px; font-size: 15px;">
            📊 Overall All Performance Summary
        </h3>
        <div style="overflow-x: auto; margin: 15px 0; border: 1px solid #e2e8f0; border-radius: 8px;">
            <table style="width: 100%; border-collapse: collapse; font-size: 12px;">
                <thead>
                    <tr style="background-color: #0f172a; color: #ffffff;">
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: left; font-weight: 600;">Team Member</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Leads</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Contacts</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Quotations Created</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Act. Completed</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Conversion Rates</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Draft Quotations</th>
                        <th style="padding: 10px; border: 1px solid #e2e8f0; text-align: center; font-weight: 600;">Open Quotations</th>
                    </tr>
                </thead>
                <tbody>
                    {rows}
                </tbody>
            </table>
        </div>
        """.format(rows=rows_overall)
        user_emails = [u["user_id"] for u in crm_users]
        comparison_matrix_html = get_comparison_matrix_html(user_emails)
        stats_html = table_period + comparison_matrix_html + table_overall

    base_url = frappe.utils.get_url()
    html_template = """<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
</head>
<body style="font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f7fa; color: #333; margin: 0; padding: 0;">
    <div class="wrapper" style="max-width: 750px; margin: 40px auto; background-color: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.08);">
        <div class="header" style="background-color: #1e3a8a; background: #1e3a8a linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%); padding: 30px; color: #ffffff; text-align: center;">
            <h1 style="margin: 0; font-size: 24px; font-weight: 700; color: #ffffff;">&#128202; __TITLE__</h1>
            <p style="margin: 8px 0 0; font-size: 14px; color: #dbeafe;">__PLABEL__ report &nbsp;&#183;&nbsp; __PERIOD__</p>
            <p style="margin: 4px 0 0; font-size: 12px; color: #bfdbfe;">Generated __TODAY__</p>
        </div>
        <div class="content" style="padding: 30px 40px; line-height: 1.6;">
            __GREETING_BLOCK__
            <p class="intro" style="font-size: 14px; color: #555; margin-bottom: 25px;">Please find below the performance counts and tracking matrix for this period. The detailed Excel workbook is also attached.</p>
            
            __STATS_HTML__
            
            <div style="margin-top: 30px; border-top: 1px solid #e2e8f0; padding-top: 20px; text-align: center;">
                <p style="font-size: 13px; color: #475569; margin-bottom: 12px; font-weight: bold;">🔗 Quick Links to CRM Entries:</p>
                <div style="text-align: center; margin-bottom: 20px;">
                    <a href="__BASE_URL__/app/crm-dashboard" style="display: inline-block; background-color: #2563eb; color: #ffffff; padding: 8px 16px; border-radius: 6px; text-decoration: none; font-size: 13px; font-weight: 600; margin: 4px;">CRM Dashboard</a>
                    <a href="__BASE_URL__/app/lead" style="display: inline-block; background-color: #1e3a8a; color: #ffffff; padding: 8px 16px; border-radius: 6px; text-decoration: none; font-size: 13px; font-weight: 600; margin: 4px;">Leads List</a>
                    <a href="__BASE_URL__/app/business-contacts" style="display: inline-block; background-color: #1e3a8a; color: #ffffff; padding: 8px 16px; border-radius: 6px; text-decoration: none; font-size: 13px; font-weight: 600; margin: 4px;">Contacts List</a>
                    <a href="__BASE_URL__/app/quotation" style="display: inline-block; background-color: #1e3a8a; color: #ffffff; padding: 8px 16px; border-radius: 6px; text-decoration: none; font-size: 13px; font-weight: 600; margin: 4px;">Quotations List</a>
                    <a href="__BASE_URL__/app/event-activity" style="display: inline-block; background-color: #1e3a8a; color: #ffffff; padding: 8px 16px; border-radius: 6px; text-decoration: none; font-size: 13px; font-weight: 600; margin: 4px;">Activities List</a>
                </div>
            </div>
            
            <p style="font-size: 14px; color: #555; margin-bottom: 12px;"><strong>What's inside the attached Excel workbook:</strong></p>
            <ul style="padding-left: 20px; font-size: 14px; color: #475569;">
                <li style="margin-bottom: 8px;"><strong>Summary Overview</strong>: Overall performance comparison matrix with team member breakdown.</li>
                <li style="margin-bottom: 8px;"><strong>Detailed Period Splits</strong>: Detailed period metrics with Draft, Open, and Ordered splits.</li>
                <li style="margin-bottom: 8px;"><strong>Overall Summary (All-Time)</strong>: Comprehensive all-time team performance comparison.</li>
                <li style="margin-bottom: 8px;"><strong>Pending Quotations (All-Time)</strong>: Draft and Open Quotations with direct link redirection.</li>
                <li style="margin-bottom: 8px;"><strong>All Leads, Contacts, Activities & Quotations</strong>: Detailed record sheets with direct links.</li>
            </ul>
            
            <p style="margin-top: 30px; font-size: 14px; color: #555;">Best regards,<br><strong style="color: #1e3a8a;">DAC CRM System</strong></p>
        </div>
        <div class="footer" style="background-color: #f8fafc; padding: 20px; text-align: center; font-size: 12px; color: #94a3b8; border-top: 1px solid #e2e8f0;">
            This is an automated system email. Please do not reply directly.
        </div>
    </div>
</body>
</html>"""
    greeting_block = (
        '<div class="greeting" style="font-size: 16px; font-weight: bold; '
        'color: #1e3a8a; margin-bottom: 15px;">{}</div>'.format(greeting)
        if greeting else ""
    )
    return html_template.replace("__TITLE__", title)\
                         .replace("__PERIOD__", period_range)\
                         .replace("__PLABEL__", plabel)\
                         .replace("__TODAY__", today_str)\
                         .replace("__GREETING_BLOCK__", greeting_block)\
                         .replace("__STATS_HTML__", stats_html)\
                         .replace("__BASE_URL__", base_url)




import frappe
from frappe import _

# @frappe.whitelist()
# def check_linked_quotation(lead_name):
#     has_quotation = frappe.db.exists(
#         "Quotation", 
#         {
#             "quotation_to": "Lead", 
#             "party_name": lead_name, 
#             "docstatus": ["<", 2] 
#         }
#     )
#     return {"has_quotation": bool(has_quotation)}


# @frappe.whitelist()
# def convert_lead_to_business_contact(lead_name):
    
#     if not frappe.db.exists("Lead", lead_name):
#         frappe.throw(_("Lead {0} not found").format(lead_name))
        
#     lead_doc = frappe.get_doc("Lead", lead_name)
    
#     # Strictly Check for quotation at DB level 
#     has_quotation = frappe.db.exists(
#         "Quotation", 
#         {
#             "quotation_to": "Lead", 
#             "party_name": lead_name, 
#             "docstatus": ["<", 2] 
#         }
#     )

#     # 1. Base Field Mappings
#     bc = frappe.new_doc("Business Contacts")
#     bc.contact_name = lead_doc.lead_name or lead_doc.first_name
#     bc.organization_name = lead_doc.company_name
#     bc.mobile_number = lead_doc.mobile_no or lead_doc.phone
#     bc.source = lead_doc.source
#     bc.assign_to = lead_doc.lead_owner
    
#     if has_quotation:
#         bc.status = "Converted to Lead"
#         bc.lead_id = lead_doc.name
#     else:
#         bc.status = "Open"
#         bc.lead_id = None          

#     bc.insert(ignore_permissions=True)
    
#     # 2. Scenarios 
#     if not has_quotation:
#         # NO QUOTATION = TRANSFER ALL ACTIVITIES TO CONTACT & DELETE LEAD
#         table_mappings = {
#             "Communication": "reference_doctype",
#             "Comment": "reference_doctype",
#             "ToDo": "reference_type",
#             "Event Activity": "reference_type" 
#         }
        
#         for dt, ref_col in table_mappings.items():
#             if frappe.db.exists("DocType", dt):
#                 try:
#                     frappe.db.sql(f"""
#                         UPDATE `tab{dt}` 
#                         SET `{ref_col}`='Business Contacts', reference_name=%s
#                         WHERE `{ref_col}`='Lead' AND reference_name=%s
#                     """, (bc.name, lead_doc.name))
#                 except Exception:
#                     continue
                    
#         frappe.db.commit()
        
#         frappe.delete_doc("Lead", lead_doc.name, force=True, ignore_permissions=True)
#         return {"action": "deleted", "business_contact_name": bc.name}

#     else:
#         # WITH QUOTATION = ACTIVITIES REMAIN IN THE LEAD AND LEAD STAYS ALIVE. WE JUST LINK IT!
#         frappe.db.set_value("Lead", lead_doc.name, "custom_business_contacts", bc.name)
#         frappe.db.commit()
#         return {"action": "linked", "business_contact_name": bc.name}




import frappe
from frappe import _

@frappe.whitelist()
def check_linked_quotation(lead_name):
    # Override permissions for checking linked quotation
    frappe.flags.ignore_permissions = True
    
    has_quotation = frappe.db.exists(
        "Quotation", 
        {
            "quotation_to": "Lead", 
            "party_name": lead_name, 
            "docstatus": ["<", 2] 
        }
    )
    return {"has_quotation": bool(has_quotation)}


@frappe.whitelist()
def convert_lead_to_business_contact(lead_name):
    # OVERRIDE ALL PERMISSIONS (Create, Read, Write, Delete, Attachments)
    frappe.flags.ignore_permissions = True
    
    if not frappe.db.exists("Lead", lead_name):
        frappe.throw(_("Lead {0} not found").format(lead_name))
        
    lead_doc = frappe.get_doc("Lead", lead_name)
    
    # Strictly Check for quotation at DB level 
    has_quotation = frappe.db.exists(
        "Quotation", 
        {
            "quotation_to": "Lead", 
            "party_name": lead_name, 
            "docstatus": ["<", 2] 
        }
    )

    # 1. Base Field Mappings
    bc = frappe.new_doc("Business Contacts")
    bc.contact_name = lead_doc.lead_name or lead_doc.first_name
    bc.organization_name = lead_doc.company_name
    bc.mobile_number = lead_doc.mobile_no or lead_doc.phone
    bc.source = lead_doc.source
    bc.assign_to = lead_doc.lead_owner
    
    # ----------------------------------------------------
    # MAP ADDRESS
    # ----------------------------------------------------
    if lead_doc.get("custom_address"):
        bc.address = lead_doc.custom_address
        if "custom_address" in [f.fieldname for f in bc.meta.fields]:
            bc.custom_address = lead_doc.custom_address

    # ----------------------------------------------------
    # MAP INDUSTRY
    # ----------------------------------------------------
    industry_value = lead_doc.get("industry") or lead_doc.get("custom_industry")
    if industry_value:
        bc_fields = [f.fieldname for f in bc.meta.fields]
        if "industry" in bc_fields:
            bc.industry = industry_value
        if "custom_industry" in bc_fields:
            bc.custom_industry = industry_value

    if has_quotation:
        bc.status = "Converted to Lead"
        bc.lead_id = lead_doc.name
    else:
        bc.status = "Open"
        bc.lead_id = None          

    bc.insert(ignore_permissions=True)

    # ----------------------------------------------------
    # MAP CREATION DATE FROM LEAD TO BUSINESS CONTACTS
    # ----------------------------------------------------
    frappe.db.set_value("Business Contacts", bc.name, "creation", lead_doc.creation, update_modified=False)
    bc.creation = lead_doc.creation
    
    # 2. Scenarios 
    if not has_quotation:
        # NO QUOTATION = TRANSFER ALL ACTIVITIES & LINKS TO CONTACT & DELETE LEAD
        table_mappings = {
            "Communication": "reference_doctype",
            "Comment": "reference_doctype",
            "ToDo": "reference_type",
            "Event Activity": "reference_type" 
        }
        
        for dt, ref_col in table_mappings.items():
            if frappe.db.exists("DocType", dt):
                try:
                    frappe.db.sql(f"""
                        UPDATE `tab{dt}` 
                        SET `{ref_col}`='Business Contacts', reference_name=%s
                        WHERE `{ref_col}`='Lead' AND reference_name=%s
                    """, (bc.name, lead_doc.name))
                except Exception:
                    continue

        # Re-link attached Addresses / Contacts from Lead to Business Contacts
        try:
            frappe.db.sql("""
                UPDATE `tabDynamic Link`
                SET link_doctype='Business Contacts', link_name=%s
                WHERE link_doctype='Lead' AND link_name=%s
            """, (bc.name, lead_doc.name))
        except Exception:
            pass
                    
        frappe.db.commit()
        
        # Delete Lead - Bypass all permission checks
        frappe.delete_doc("Lead", lead_doc.name, force=True, ignore_permissions=True)
        return {"action": "deleted", "business_contact_name": bc.name}

    else:
        # WITH QUOTATION = ACTIVITIES REMAIN IN THE LEAD. LINK IT!
        frappe.db.set_value("Lead", lead_doc.name, "custom_business_contacts", bc.name)
        frappe.db.commit()
        return {"action": "linked", "business_contact_name": bc.name}

@frappe.whitelist()
def check_leads_for_conversion(lead_names):
    import json
    if isinstance(lead_names, str):
        lead_names = json.loads(lead_names)

    already_converted = []
    has_quotations = []
    no_quotations = []

    for name in lead_names:
        custom_bc = frappe.db.get_value("Lead", name, "custom_business_contacts")
        if custom_bc:
            already_converted.append(name)
            continue

        has_q = frappe.db.exists(
            "Quotation", 
            {
                "quotation_to": "Lead", 
                "party_name": name, 
                "docstatus": ["<", 2] 
            }
        )
        if has_q:
            has_quotations.append(name)
        else:
            no_quotations.append(name)

    return {
        "already_converted": already_converted,
        "has_quotations": has_quotations,
        "no_quotations": no_quotations
    }

@frappe.whitelist()
def bulk_convert_leads(lead_names):
    import json
    if isinstance(lead_names, str):
        lead_names = json.loads(lead_names)

    converted_linked = []
    converted_deleted = []
    skipped_already_converted = []
    errors = []

    for name in lead_names:
        try:
            custom_bc = frappe.db.get_value("Lead", name, "custom_business_contacts")
            if custom_bc:
                skipped_already_converted.append(name)
                continue

            res = convert_lead_to_business_contact(name)
            if res.get("action") == "deleted":
                converted_deleted.append(name)
            else:
                converted_linked.append(name)
        except Exception as e:
            errors.append(f"Lead {name}: {str(e)}")

    return {
        "converted_linked": converted_linked,
        "converted_deleted": converted_deleted,
        "skipped_already_converted": skipped_already_converted,
        "errors": errors
    }

@frappe.whitelist()
def mark_lead_lost_backend(lead_name, category, lost_reason, lost_reason_description, current_activity_id=None, completion_note=None):
    # 1. Complete the current activity if passed
    if current_activity_id:
        frappe.db.set_value("Event Activity", current_activity_id, {
            "status": "Completed",
            "notes": completion_note or "Lead marked as Lost",
            "ends_on": frappe.utils.now_datetime()
        })
    
    # 2. Cancel all other Open activities for this Lead
    frappe.db.sql("""
        UPDATE `tabEvent Activity`
        SET status = 'Cancelled', notes = 'Lead marked as Lost'
        WHERE reference_type = 'Lead' AND reference_name = %s AND status = 'Open'
    """, lead_name)
    
    # 3. Update the Lead status/category and details
    lead = frappe.get_doc("Lead", lead_name)
    lead.custom_lead_type = "LOST"
    
    if category == "Enquiry":
        lead.custom_lost_enquiry_reason = lost_reason
        lead.custom_lost_enquiry_description = lost_reason_description
        action = "Lost Enquiry"
    else:
        lead.custom_lost_pipeline_reason = lost_reason
        lead.custom_lost_pipeline_description = lost_reason_description
        action = "Lost Pipeline"
        
    # Apply workflow action
    from frappe.model.workflow import apply_workflow
    apply_workflow(lead, action)
    
    return {"status": "success"}

CRM_DASHBOARD_SOURCE_FILE = "crm_dashboard_v31.html"
CRM_DASHBOARD_BLOCKS = ["CRM Dashboard", "New CRM Dashboard"]


@frappe.whitelist()
def sync_crm_dashboard_html_block(source_file=None, blocks=None):
    """Syncs the combined dashboard file into the target Custom HTML Blocks.

    The source file holds <style>, markup and <script> in one document; this splits
    them into the block's style / html / script columns.

    Frappe's create_shadow_element (dom.js) wraps the script column content inside
    a backtick template literal:
        script.textContent = `...(function(){ let root_element=...; ${js} })();`;

    Any backtick inside our JS (from ES6 template literals) breaks Frappe's outer
    template literal causing: "Failed to execute 'appendChild' on 'Node': Unexpected token"

    Fix: escape all backticks in js_only with backslash before storing.

    Args:
        source_file: file name inside crm_dashboard_block/ (defaults to CRM_DASHBOARD_SOURCE_FILE)
        blocks: list or comma-separated names of Custom HTML Blocks to update
    """
    import os
    import re
    source_file = source_file or CRM_DASHBOARD_SOURCE_FILE
    target_blocks = as_filter_list(blocks) or CRM_DASHBOARD_BLOCKS

    app_path = frappe.get_app_path("erp_dacsinc_custom")
    html_file = os.path.normpath(os.path.join(app_path, "..", "crm_dashboard_block", source_file))

    if not os.path.exists(html_file):
        return "Error: {} not found".format(source_file)

    with open(html_file, "r", encoding="utf-8") as f:
        full_content = f.read()

    # Split HTML, CSS, and JS: extract style and script content
    style_match = re.search(r'<style[^>]*>([\s\S]*?)</style>', full_content, re.IGNORECASE)
    if style_match:
        css_only = style_match.group(1).strip()
    else:
        css_only = ""

    script_match = re.search(r'<script[^>]*>([\s\S]*?)</script>', full_content, re.IGNORECASE)
    if script_match:
        js_only = script_match.group(1).strip()
    else:
        js_only = ""

    html_only = full_content
    if style_match:
        html_only = html_only.replace(style_match.group(0), "")
    if script_match:
        html_only = html_only.replace(script_match.group(0), "")
    html_only = html_only.strip()

    updated, missing = [], []
    for b_name in target_blocks:
        if frappe.db.exists("Custom HTML Block", b_name):
            frappe.db.set_value("Custom HTML Block", b_name, "html", html_only)
            frappe.db.set_value("Custom HTML Block", b_name, "style", css_only)
            frappe.db.set_value("Custom HTML Block", b_name, "script", js_only)
            frappe.clear_cache(doctype="Custom HTML Block")
            updated.append(b_name)
        else:
            missing.append(b_name)

    frappe.db.commit()
    msg = (f"Synced {source_file} (html: {len(html_only)} bytes, style: {len(css_only)} bytes, "
           f"script: {len(js_only)} bytes, raw backticks preserved) into: {', '.join(updated) or 'nothing'}")
    if missing:
        msg += f" | skipped (no such block): {', '.join(missing)}"
    return msg


@frappe.whitelist()
def sync_business_contacts_history():
	if "System Manager" not in frappe.get_roles():
		frappe.throw(frappe._("Not authorized"), frappe.PermissionError)

	import uuid

	contacts = frappe.db.get_all(
		"Business Contacts",
		filters={"status": "Converted to Lead"},
		fields=["name", "lead_id", "creation"]
	)

	updated_count = 0

	for c in contacts:
		# Check if there is any history for this Business Contact
		has_history = frappe.db.exists(
			"Lead Status Change History",
			{
				"parent": c.name,
				"parenttype": "Business Contacts",
				"parentfield": "contact_status_change_history"
			}
		)
		
		if not has_history:
			# Get Lead creation date, fallback to Business Contact creation date
			lead_creation_date = frappe.db.get_value("Lead", c.lead_id, "creation") or c.creation
			
			# Direct DB Insert to avoid overhead/timeout and bypass validation/hooks
			frappe.db.sql("""
				INSERT INTO `tabLead Status Change History`
					(name, parent, parenttype, parentfield, idx,
					 old_status, new_status, changed_by, updated_at, reason)
				VALUES
					(%s, %s, 'Business Contacts', 'contact_status_change_history', 1,
					 'Open', 'Converted to Lead', 'Administrator', %s, 'Converted to Lead')
			""", (f"auto-{uuid.uuid4()}", c.name, lead_creation_date))
			
			updated_count += 1

	if updated_count > 0:
		frappe.db.commit()

	return {"status": "success", "updated_count": updated_count}